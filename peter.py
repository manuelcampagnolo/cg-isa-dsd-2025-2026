######################################################
# Manuel Campagnolo (abril 2023)
# ISA/ULIsboa
# script para preparar resumo do ficheiro Excel DSD ISA 2023-2024
#######################################################

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.styles import Border, Side

from copy import copy
import os
from fuzzywuzzy import fuzz # compare strings
import numpy as np

import pandas as pd
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

########################################################################################################
SAVE_HTML_MAT=False

fnDSD="DSD_2324_12jun2023_CorrigidoML_desprotegido_editado_MLC.xlsx"  #"DSD_2324_27maio.xlsx" #  "DSD_v1_teste.xlsx"; 
fnResumo="resumo_DSD_2324_12junho2023__hf_hd.xlsx"
VALIDATION_VALUE='Inserir docente'
# worksheets de input
ws_name_preencher='DSD (para preencher)'
ws_name_info='DSD (informação UCs)'
ws_name_docentes='DocentesNovo' 
# ws preencher

coldoc='AK'
print(column_index_from_string(coldoc))
colunas_red=['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento',  'Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC', 'ano curricular', 'semestre', 'ECTS', 'semanas', 'Total horas da UC', 'Somatório', 'Horas em falta na UC']
column_codigo_UC_preencher='Código UC'
column_somatorio_preencher= 'Somatório'
column_preencher_first='Total Horas Teóricas'
column_preencher_last='Total Horas  Outras'
column_new_horas_totais='Horas totais docente'
column_new_horas_semanais='Horas semanais docente'
info_cols_to_unprotect=['H','M','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AD','AE']
DAA='Docente a atribuir'
# ws info (colocada para facilitar VLOOKUP)
column_total_horas_info='Total Horas previsto '
column_codigo_UC_info='Código UC'
# colunas DSD
columns_to_copy='Nome da UC' # 'Nome' # designação UC
column_key='Inserir docentes na UC ' #'docentes ' # cuidado: tem um espaço a mais
coluna_validacao=column_key
col_codigo_uc='Código UC'
column_horas_em_falta_preencher='Horas em falta na UC' #Horas em falta na UC
column_somatorio_preencher= 'Somatório'
colunas_nome_UC=['Nome da UC', column_codigo_UC_info, 'ciclo de estudos', 'ano curricular', 'semestre'] #Nome da UC
colunas_FOS=['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento']
colResponsavel='Responsável'
newcolResponsavel='nomeResponsavel'
colNomeDocente='NomeDocente'
colLinhaResp='linhaResp'
cols_horas=['Total Horas Teóricas', 'Horas Teóricas por semana','Total Horas Teórico-práticas', 'Horas Teórico-práticas por semana', 'Total Horas Laboratoriais', 'Horas Laboratoriais por semana', 'Total Horas  Trabalho de campo', 'Horas Trabalho de Campo por semana','Total Horas  Seminário', 'Horas Seminário por semana','Total Horas  Estágio', 'Total Horas  Outras']
# colunas DSD docentes
col_horas_semanais='Horas semanais docente'
col_horas_totais='Horas totais docente'
col_nome_completo='Nome completo'
col_posicao='Posição'
cols_horas_docentes=[col_horas_semanais,col_horas_totais,col_nome_completo,col_posicao]

soma_horas_docente_uc='Soma horas docente UC'
total_horas_docente='Total horas docente'
total_horas_docencia_uc='Total horas docentes para UC'
total_horas_uc_='Total horas UC'

# colunas a apagar em DSD info: funciona
col_apagar_info_1='Total Horas Somadas '
col_apagar_info_2='Horas em falta na UC'
col_apagar_info_3='Código UC'
col_apagar_info_4='Total Horas previsto ' #Total Horas previsto 
################################################################################################################# funcoes
def fill_empty_cells(df, col_name):
    # Create a copy of the dataframe to avoid modifying the original
    filled_df = df.copy()
    
    # Create a mask of empty cells in the column
    empty_mask = filled_df[col_name].isna()
    
    # Iterate over each row in the column
    for i, cell_value in enumerate(filled_df[col_name]):
        # If the cell is empty, fill it with the value of the closest non-empty cell above
        if pd.isna(cell_value):
            j = i - 1
            while j >= 0 and pd.isna(filled_df.loc[j, col_name]):
                j -= 1
            if j >= 0:
                filled_df.loc[i, col_name] = filled_df.loc[j, col_name]           
    return filled_df

# def df_to_excel_simple(df, ws, header=True, index=True):
#     for r in dataframe_to_rows(df, index=index, header=header):
#         ws.append(r)

def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def create_workbook_from_dataframe(df):
    """
    1. Create workbook from specified pandas.DataFrame
    2. Adjust columns width to fit the text inside
    3. Make the index column and the header row bold
    4. Fill background color for the header row

    Other beautification MUST be done by usage side.
    """
    workbook = openpyxl.Workbook()
    ws = workbook.active

    rows = dataframe_to_rows(df.reset_index(), index=False)
    col_widths = [0] * (len(df.columns) + 1)
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):

            if type(val) is str:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            elif hasattr(val, "sort"):
                cell = ws.cell(row=i, column=j, value=", ".join(list(map(lambda v: str(v), list(val)))))
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            else:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val)) + 1])

            # Make the index column and the header row bold
            if i == 1 or j == 1:
                cell.font = Font(bold=True)

    # Adjust column width
    for i, w in enumerate(col_widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = w

    return workbook

def set_border(ws):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows()
    for row in rows:
        for cell in row:
            cell.border = border


def df_to_excel_with_columns(df,ws,maxwidth=30,header=True,index=False):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = min(max_length+2, maxwidth) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_excel(df,ws,header=True,index=False)


############################################################ funcoes
# devolve letra da coluna com nome (1a linha) da worksheet ws
def nomeColuna2letter(ws,nome):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx=headers.index(nome)
    return get_column_letter(idx+1)

####################################################################################################################

dfinfo= pd.read_excel(fnDSD, sheet_name=ws_name_info) 
dfdsd= pd.read_excel(fnDSD, sheet_name=ws_name_preencher)
dfdsd.columns

# info docentes em DSD
dfaux=dfdsd[cols_horas_docentes]
#dfaux.columns
dfaux=dfaux.dropna()
dfaux=dfaux.drop(['Horas semanais docente', 'Horas totais docente'],axis=1)
len(dfaux) #153 # todos têm Posição

# limpar dfinfo
dfinfo=dfinfo.drop(col_apagar_info_1,axis=1)
dfinfo=dfinfo.drop(col_apagar_info_2,axis=1)
dfinfo=dfinfo.drop(col_apagar_info_3,axis=1)
dfinfo=dfinfo.drop(col_apagar_info_4,axis=1)
dfinfo = dfinfo.dropna(axis=1, how='all')

# criar coluna de Responsável
dfdsd.loc[dfdsd[colResponsavel]=='sim',newcolResponsavel]=dfdsd.loc[dfdsd[colResponsavel]=='sim',column_key]
#dfdsd.head(20)[newcolResponsavel]

# preencher algumas colunas com o valor acima
for col in colunas_FOS+colunas_nome_UC+[newcolResponsavel]:
    dfdsd=fill_empty_cells(dfdsd, col)
print(dfdsd.shape)

# check
#dfdsd.loc[dfdsd['Nome da UC']=='Estética e Ética da Paisagem',['Nome da UC', 'nomeResponsavel']]

# remover linhas com VALIDATION_VALUE ou em branco na coluna column_key: se não tiver sido indicado docente, a linha é descartada
dfdsd=dfdsd[dfdsd[column_key]!= VALIDATION_VALUE]
dfdsd=dfdsd[dfdsd[column_key]!= DAA] # novo junho 2023: se é 'docente a atribuir' é descartado
dfdsd=dfdsd[dfdsd[column_key].notna()]
print(dfdsd.shape)

# drop 2nd row
#dfdsd=dfdsd.drop(1)

# remover colunas DSD
dfdsd=dfdsd[colunas_FOS+[colResponsavel, newcolResponsavel,column_key]+colunas_nome_UC+cols_horas+[column_horas_em_falta_preencher]]

# converter horas para numeric
for col in cols_horas:
    dfdsd[col] = pd.to_numeric(dfdsd[col], errors='coerce')
#dfdsd.dtypes

# Contabilizar as horas em falta a partir de dfdsd (novo: junho 2023)
dfdsd[soma_horas_docente_uc]=dfdsd[cols_horas].sum(axis=1)

# com groupby, somar todas as linhas que correspondem ao mesmo docente
dfdsd[total_horas_docente] = dfdsd.groupby(column_key)[soma_horas_docente_uc].transform('sum')

if False: 
    #selecionar colunas dos docentes em DSD
    dfhd=dfdsd[cols_horas_docentes]
    dfhd=dfhd.dropna(axis=0)
    dfhd[col_horas_semanais]=dfhd[col_horas_semanais].round(2)


dfdsd.columns
dfinfo.columns

# Criar e preencher dfucs
# criar novas colunas em dfdsd e dfinfo com concatenação de nome UC, ciclo, area disciplinar, ano curricular, semestre
chaves_dfdsd=['Áreas Disicplinares', 'Nome da UC', 'ciclo de estudos', 'ano curricular', 'semestre']
chaves_dfinfo=['Área disicplinar',  'Nome', 'ciclo de estudos','ano curricular', 'semestre']
def concat_cols(df,cols):
    newcol=df[cols[0]].map(str)
    for i in range(1,len(cols)):
        newcol=newcol+df[cols[i]].map(str)
    return newcol

# criar chaves para cruzamento de dfdsd e dfinfo
dfdsd['keyUC']=concat_cols(dfdsd,chaves_dfdsd)
dfinfo['keyUC']=concat_cols(dfinfo,chaves_dfinfo)

# criar coluna com código da UC
dfdsd=pd.merge(dfdsd, dfinfo[['Nome','Código UC.1','Total Horas previsto .1','keyUC']], left_on='keyUC', right_on='keyUC',)

# re-nomear colunas dfdsd
dfdsd = dfdsd.rename(columns={'Código UC.1': 'codigoUC', 'Total Horas previsto .1': total_horas_uc_})

# com groupby, somar todas as linhas que correspondem à mesma UC
dfdsd[total_horas_docencia_uc] = dfdsd.groupby('codigoUC')[soma_horas_docente_uc].transform('sum')
dfdsd[column_horas_em_falta_preencher]=dfdsd[total_horas_uc_]-dfdsd[total_horas_docencia_uc]

# funciona:
#dfdsd.loc[dfdsd['codigoUC']==1749, cols_horas+[soma_horas_docente_uc]]
#dfdsd.loc[dfdsd[soma_horas_docente_uc]>0, cols_horas+[soma_horas_docente_uc]].iloc[100]

# selecionar as linhas dos responsáveis das UCs
dfucs=dfdsd[dfdsd[colResponsavel]=='sim']
# os códigos de UCs não devem ter repetições
if dfucs['codigoUC'].duplicated().sum() != 0: 
    stop
dfucs=dfucs[colunas_FOS+[newcolResponsavel]+colunas_nome_UC+[total_horas_uc_,total_horas_docencia_uc,column_horas_em_falta_preencher]]

# eliminar as linhas dos responsáveis
dfdsd=dfdsd[dfdsd[colResponsavel]!='sim']

#################
dfdsd.columns
# criar dfhd
horas_docentes=dfdsd.groupby(column_key)[soma_horas_docente_uc].sum()
dfhd=pd.merge(dfaux,horas_docentes,left_on='Nome completo',right_index=True, how='left')
dfhd['Horas semanais']=round(dfhd[soma_horas_docente_uc]/28,2)
len(dfhd) # 138


#################### validação

# # comparar dfhd e dfdsd quando se agrupa horas por docente
# dfdoc=dfdsd.groupby(column_key)['Soma horas docente UC'].sum()
# dfdoc=pd.merge(dfdoc, dfhd, on=column_key)
# dfdoc['diff']=dfdoc['Soma horas docente UC_x']-dfdoc['Soma horas docente UC_y']
# dfdoc.sort_values(by='diff')
# dfdoc.sum(axis=0)
# len(dfdoc)
# len(dfhd)
# dfhd

# validação
# número de horas totais docência
# em dfdsd
dfdsd[soma_horas_docente_uc].sum()
# em dfhd
dfhd[soma_horas_docente_uc].sum()
# número total de horas das UCs
dfucs[total_horas_uc_].sum()
# total horas em falta
dfucs[column_horas_em_falta_preencher].sum()
dfucs[total_horas_uc_].sum()-dfdsd[soma_horas_docente_uc].sum()

############### output
wbr=openpyxl.Workbook()
wsr_ucs=wbr.create_sheet('horas_UCs')
wsr_docentes=wbr.create_sheet('horas_docentes')
wsr_ucs_docentes=wbr.create_sheet('horas_UCs_docentes')
wsr_info=wbr.create_sheet('info_UCs')

#cols_dfucs_drop=['keyUC',soma_horas_docente_uc, 'Nome', 'codigoUC']
cols_dfdsd=['Áreas Disicplinares', 'Departamento', 'Responsável', 'nomeResponsavel','Inserir docentes na UC ', 'Nome da UC', 'Código UC',soma_horas_docente_uc]
cols_dfinfo_drop=['keyUC']

df_to_excel_with_columns(dfucs, wsr_ucs)
df_to_excel_with_columns(dfdsd[cols_dfdsd], wsr_ucs_docentes)
df_to_excel_with_columns(dfhd, wsr_docentes)
df_to_excel_with_columns(dfinfo.drop(cols_dfinfo_drop,axis=1), wsr_info)

# Freeze the top row, and add filter 
for ws in [wsr_ucs,wsr_ucs_docentes,wsr_docentes,wsr_info]:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

if 'Sheet' in wbr.sheetnames:  # remove default sheet
    wbr.remove(wbr['Sheet'])

wbr.save(fnResumo)
wbr.close

if SAVE_HTML_MAT:
    output_columns=['Código UC','Nome da UC', 'Áreas Disicplinares', 'Áreas Científicas (FOS)',  'ciclo de estudos','nomeResponsavel']
    # determinar UCs em que o responsável é da SM
    dfucs.columns
    dfucs.loc[dfucs['Áreas Disicplinares']=='MAT',['Nome da UC','nomeResponsavel']]
    docs_mat=list(dfucs.loc[dfucs['Áreas Disicplinares']=='MAT','nomeResponsavel'].unique())
    ucs_mat=list(dfucs.loc[dfucs['nomeResponsavel'].isin(docs_mat),'Nome da UC'].unique())
    # UCs coordenadas por docentes MAT
    df=dfucs.loc[dfucs['nomeResponsavel'].isin(docs_mat) ,output_columns]
    # determinar UCs em que o docente é da SM
    dfdsd.columns # 'Inserir docentes na UC '
    ucs_all_mat=dfdsd.loc[dfdsd['Inserir docentes na UC '].isin(list(docs_mat)) ,'Nome da UC'].unique()
    ucs_not_mat=list(set(ucs_all_mat).difference(set(ucs_mat)))
    df=pd.concat([df,dfucs.loc[dfucs['Nome da UC'].isin(ucs_not_mat) ,output_columns]],axis=0)
    df=df.sort_values(by=['ciclo de estudos','Nome da UC'])
    # corrigir nomes das colunas
    df=df.rename(columns={'nomeResponsavel': 'Nome Responsável'})
    df=df.rename(columns={'Áreas Disicplinares': 'Áreas Disciplinares'})
    html_table = df.to_html(index=False)
    # save HTML table to a file
    with open('UCs_SM.html', 'w',encoding='utf-8') as f:
        f.write(html_table)


# horas totais em cada ciclo
dfinfo.head()
dfinfo.duplicated

# número de UCs
len(dfinfo)
# número de horas de UCs pro ciclo
dfinfo.drop_duplicates().groupby('ciclo de estudos')['Total Horas previsto .1'].sum()

