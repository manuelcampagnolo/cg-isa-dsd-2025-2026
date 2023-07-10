######################################################
# Manuel Campagnolo (abril 2023)
# ISA/ULIsboa
# script para preparar resumo do ficheiro Excel DSD ISA 2023-2024
#######################################################

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.styles import Border, Side

from copy import copy
import os
from fuzzywuzzy import fuzz # compare strings
import numpy as np

from fuzzywuzzy import process
import pandas as pd

from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

########################################################################################################
SAVE_HTML_MAT=False

fnDSD="DSD_2324_12jun2023_CorrigidoML_desprotegido_editado_MLC_29junho.xlsx"  #"DSD_2324_27maio.xlsx" #  "DSD_v1_teste.xlsx"; 
fnExterno="DSD_2023_2024_servico_externo_v6_revisto_TF_DSD_28junho.xlsx" #"DSD_2023_2024_servico_externo_v5_revisto_TF_28junho.xlsx"  #"DSD_2324_27maio.xlsx" #  "DSD_v1_teste.xlsx"; 
#docente_extra=('Duarte Neiva', 'Química',5)
# Output
fnResumo="resumo_DSD_2324_1julho2023.xlsx"
#fnNomePosicao="nomes_docentes_codigos_RH_17maio2023_editado_MLC.xlsx"
fnNomePosicao="nomes_docentes_codigos_RH_17maio2023_editado_MLC.xlsx"

VALIDATION_VALUE='Inserir docente'
# worksheets de input
ws_name_preencher='DSD (para preencher)'
ws_name_info='DSD (informação UCs)'
ws_name_docentes='DocentesNovo' 
ws_horas_docentes_extra='horasDocentesExtra'
ws_mudanca_responsaveis='mudancaResponsavel'
# ws preencher

coldoc='AK'
print(column_index_from_string(coldoc))
colunas_red=['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento',  'Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC', 'ano curricular', 'semestre', 'ECTS', 'semanas', 'Total horas da UC', 'Somatório', 'Horas em falta na UC']
column_codigo_UC_preencher='Código UC'
column_somatorio_preencher= 'Somatório'
column_preencher_first='Total Horas Teóricas'
column_preencher_last='Total Horas  Outras'
column_new_horas_totais='Horas totais docente'
column_new_horas_semanais='Horas semanais docente'
info_cols_to_unprotect=['H','M','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AD','AE']
DAA='Docente a atribuir'
# ws info (colocada para facilitar VLOOKUP)
column_total_horas_info='Total Horas previsto '
column_codigo_UC_info='Código UC'
# colunas DSD
columns_to_copy='Nome da UC' # 'Nome' # designação UC
column_key='Inserir docentes na UC ' #'docentes ' # cuidado: tem um espaço a mais
coluna_validacao=column_key
col_codigo_uc='Código UC'
column_horas_em_falta_preencher='Horas em falta na UC' #Horas em falta na UC
column_somatorio_preencher= 'Somatório'
colunas_nome_UC=['Nome da UC', column_codigo_UC_info, 'ciclo de estudos', 'ano curricular', 'semestre'] #Nome da UC
colunas_FOS=['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento']
colResponsavel='Responsável'
newcolResponsavel='nomeResponsavel'
colNomeDocente='NomeDocente'
colLinhaResp='linhaResp'
cols_horas=['Total Horas Teóricas', 'Horas Teóricas por semana','Total Horas Teórico-práticas', 'Horas Teórico-práticas por semana', 'Total Horas Laboratoriais', 'Horas Laboratoriais por semana', 'Total Horas  Trabalho de campo', 'Horas Trabalho de Campo por semana','Total Horas  Seminário', 'Horas Seminário por semana','Total Horas  Estágio', 'Total Horas  Outras']
# colunas DSD docentes
col_horas_semanais='Horas semanais docente'
col_horas_totais='Horas totais docente'
col_nome_completo='Nome completo'
col_posicao='Posição'
cols_horas_docentes=[col_horas_semanais,col_horas_totais,col_nome_completo,col_posicao]

soma_horas_docente_uc='Soma horas docente UC'
total_horas_docente='Total horas docente'
total_horas_docencia_uc='Total horas docentes para UC'
total_horas_uc_='Total horas UC'

# horas externas
soma_horas_externo_docente='Horas docência externa'

# colunas a apagar em DSD info: funciona
col_apagar_info_1='Total Horas Somadas '
col_apagar_info_2='Horas em falta na UC'
#col_apagar_info_3='Código UC'
#col_apagar_info_4='Total Horas previsto ' #Total Horas previsto 
################################################################################################################# funcoes


def approximate_merge(df1, df2, key_column1, key_column2):
    merged_data = pd.merge(df1, df2, left_on=key_column1, right_on=key_column2, how='outer')

    highest_score = 0

    for i, row in merged_data.iterrows():
        key1 = row[key_column1]
        key2 = row[key_column2]
        similarity_score = fuzz.ratio(str(key1), str(key2))
        
        if similarity_score > highest_score:
            highest_score = similarity_score

    merged_data = merged_data[merged_data.apply(lambda row: fuzz.ratio(str(row[key_column1]), str(row[key_column2])) == highest_score, axis=1)]

    return merged_data

def replace_value(df, first_column_name, first_value, second_column_name, second_value):
    # Replace the value in the second column where the first column matches the first value
    df.loc[df[first_column_name] == first_value, second_column_name] = second_value
    return df


def inserir_docente(df, column_codigo_UC_preencher, input_value,soma_horas_docente_uc,horas,column_key,docente,colResponsavel):
    # Get the index of the first row where the column matches the input value
    first_match_index = df.index[df[column_codigo_UC_preencher] == input_value].min()
    # Get the first row where the column matches the input value
    row = df.loc[first_match_index].copy()
    # muda valores
    row.loc[soma_horas_docente_uc]=horas
    row.loc[column_key]=docente
    row.loc[colResponsavel]=''
    # Append the first match row to the DataFrame
    df.loc[df.index.max() + 1] = row
    return df


def fill_empty_cells(df, col_name):
    # Create a copy of the dataframe to avoid modifying the original
    filled_df = df.copy()
    
    # Create a mask of empty cells in the column
    empty_mask = filled_df[col_name].isna()
    
    # Iterate over each row in the column
    for i, cell_value in enumerate(filled_df[col_name]):
        # If the cell is empty, fill it with the value of the closest non-empty cell above
        if pd.isna(cell_value):
            j = i - 1
            while j >= 0 and pd.isna(filled_df.loc[j, col_name]):
                j -= 1
            if j >= 0:
                filled_df.loc[i, col_name] = filled_df.loc[j, col_name]           
    return filled_df

# def df_to_excel_simple(df, ws, header=True, index=True):
#     for r in dataframe_to_rows(df, index=index, header=header):
#         ws.append(r)

def create_workbook_from_dataframe(df):
    """
    1. Create workbook from specified pandas.DataFrame
    2. Adjust columns width to fit the text inside
    3. Make the index column and the header row bold
    4. Fill background color for the header row

    Other beautification MUST be done by usage side.
    """
    workbook = openpyxl.Workbook()
    ws = workbook.active

    rows = dataframe_to_rows(df.reset_index(), index=False)
    col_widths = [0] * (len(df.columns) + 1)
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):

            if type(val) is str:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            elif hasattr(val, "sort"):
                cell = ws.cell(row=i, column=j, value=", ".join(list(map(lambda v: str(v), list(val)))))
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            else:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val)) + 1])

            # Make the index column and the header row bold
            if i == 1 or j == 1:
                cell.font = Font(bold=True)

    # Adjust column width
    for i, w in enumerate(col_widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = w

    return workbook

def set_border(ws):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows()
    for row in rows:
        for cell in row:
            cell.border = border


def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def df_to_excel_with_columns(df,ws,maxwidth=30,header=True,index=False):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = max(len(column)-2,min(max_length+2, maxwidth)) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_excel(df,ws,header=True,index=False)


############################################################ funcoes
# devolve letra da coluna com nome (1a linha) da worksheet ws
def nomeColuna2letter(ws,nome):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx=headers.index(nome)
    return get_column_letter(idx+1)

####################################################################################################################

dfinfo= pd.read_excel(fnDSD, sheet_name=ws_name_info) 
dfdsd= pd.read_excel(fnDSD, sheet_name=ws_name_preencher)
dfexterno=pd.read_excel(fnExterno, sheet_name='Servico_externo') # adicionei coluna DSD
dfNomePosicao= pd.read_excel(fnNomePosicao)
dfhorasextra= pd.read_excel(fnDSD, sheet_name=ws_horas_docentes_extra)
dfalteraresp= pd.read_excel(fnDSD, sheet_name= ws_mudanca_responsaveis)

dfdsd.loc[dfdsd[column_key]=='Fernando Eduardo Lagos Costa', [ 'Inserir docentes na UC ','Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC']]

dfexterno.columns
dfdsd.columns # tem 'Código UC'
dfinfo.columns # tem 'Código UC'

# info docentes em DSD
#dfaux=dfdsd[cols_horas_docentes]
#dfaux.columns
#dfaux=dfaux.dropna() # para eliminar linhas sem informação
#dfaux=dfaux.drop(['Horas semanais docente', 'Horas totais docente'],axis=1)
dfaux=dfNomePosicao[[col_nome_completo,col_posicao]]
len(dfaux) #153 # todos têm Posição

# limpar dfinfo
dfinfo=dfinfo.drop(col_apagar_info_1,axis=1)
dfinfo=dfinfo.drop(col_apagar_info_2,axis=1)
#dfinfo=dfinfo.drop(col_apagar_info_3,axis=1) # 'Código UC'
# dfinfo=dfinfo.drop(col_apagar_info_4,axis=1) # 'Total Horas previsto '
dfinfo = dfinfo.dropna(axis=1, how='all')

# criar coluna de Responsável
dfdsd.loc[dfdsd[colResponsavel]=='sim',newcolResponsavel]=dfdsd.loc[dfdsd[colResponsavel]=='sim',column_key]
#dfdsd.head(20)[newcolResponsavel]

# preencher algumas colunas com o valor que lhe está acima na tabela
for col in colunas_FOS+colunas_nome_UC+[newcolResponsavel]:
    dfdsd=fill_empty_cells(dfdsd, col)
print(dfdsd.shape)

# check
#dfdsd.loc[dfdsd['Nome da UC']=='Estética e Ética da Paisagem',['Nome da UC', 'nomeResponsavel']]

# remover linhas com VALIDATION_VALUE ou em branco na coluna column_key: se não tiver sido indicado docente, a linha é descartada
dfdsd=dfdsd[dfdsd[column_key]!= VALIDATION_VALUE]
# as linhas seguintes eliminam todas as linhas de dfdsd que não tenham um docente escolhido no drop-down menu
dfdsd=dfdsd[dfdsd[column_key]!= DAA] # novo junho 2023: se é 'docente a atribuir' é descartado
dfdsd=dfdsd[dfdsd[column_key].notna()]  # se alguém apagou a célula fiva vazio e é lido como NA, e aqui é descartado
print(dfdsd.shape)

dfdsd.loc[dfdsd[column_key]=='Fernando Eduardo Lagos Costa', [ 'Inserir docentes na UC ','Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC']]

# Completar e organizar dfdsd (docentes extra, mudança de responsável)

# remover colunas DSD
dfdsd=dfdsd[colunas_FOS+[colResponsavel, newcolResponsavel,column_key]+colunas_nome_UC+cols_horas+[column_horas_em_falta_preencher]]

# converter horas para numeric
for col in cols_horas:
    dfdsd[col] = pd.to_numeric(dfdsd[col], errors='coerce')
#dfdsd.dtypes

# Contabilizar as horas em falta a partir de dfdsd (novo: junho 2023)
dfdsd[soma_horas_docente_uc]=dfdsd[cols_horas].sum(axis=1)
print(dfdsd.shape)


# inserir horas de docentes extra
for index, row in dfhorasextra.iterrows():
    dfdsd=inserir_docente(dfdsd, column_codigo_UC_preencher, row['codigoUC'],soma_horas_docente_uc,row['horas'],column_key,row['nomeDocenteExtra'],colResponsavel)

# alterar responsáveis
for index, row in dfalteraresp.iterrows():
    dfdsd=replace_value(dfdsd, column_codigo_UC_preencher, row['codigoUC'], newcolResponsavel, row['novoResp'])


dfdsd[[column_codigo_UC_preencher,soma_horas_docente_uc,column_key,newcolResponsavel]]
print(dfdsd.shape)

######################################### agrupar dfdsd por docente
# com groupby, somar todas as linhas que correspondem ao mesmo docente
dfdsd[total_horas_docente] = dfdsd.groupby(column_key)[soma_horas_docente_uc].transform('sum')

# OK: tem duas linhas
# dfdsd.loc[dfdsd[column_key]=='Fernando Eduardo Lagos Costa', [ 'Inserir docentes na UC ','Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC',total_horas_docente]]

dfdsd.columns
dfinfo.columns

# Criar e preencher dfucs
# criar novas colunas em dfdsd e dfinfo com concatenação de nome UC, ciclo, area disciplinar, ano curricular, semestre
if False:
    chaves_dfdsd=['Áreas Disicplinares', 'Nome da UC', 'ciclo de estudos', 'ano curricular', 'semestre']
    chaves_dfinfo=['Área disicplinar',  'Nome', 'ciclo de estudos','ano curricular', 'semestre']
    def concat_cols(df,cols):
        newcol=df[cols[0]].map(str)
        for i in range(1,len(cols)):
            newcol=newcol+df[cols[i]].map(str)
        return newcol

# criar chaves para cruzamento de dfdsd e dfinfo
#dfdsd['keyUC']=concat_cols(dfdsd,chaves_dfdsd)
#dfinfo['keyUC']=concat_cols(dfinfo,chaves_dfinfo)

# OK: 2 linhas
#dfdsd.loc[dfdsd[column_key]=='Fernando Eduardo Lagos Costa', ['keyUC','Inserir docentes na UC ','Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC',total_horas_docente]]

olddf=dfdsd.copy()
# Usar 'Código UC' para cruzar as tabelas e trazer as horas previstas para dfdsd
dfdsd=pd.merge(dfdsd, dfinfo[['Nome',col_codigo_uc,column_total_horas_info]], on=col_codigo_uc) #,how='left')
# approximate merge
# Bdfdsd=approximate_merge(dfdsd, dfinfo[['Nome','Código UC.1','Total Horas previsto .1','keyUC']], 'keyUC', 'keyUC')

dfdsd.columns
dfinfo.columns
# OK: 2 linhas
dfdsd.loc[dfdsd[column_key]=='Fernando Eduardo Lagos Costa', ['Inserir docentes na UC ','Responsável', 'Nome da UC', 'ciclo de estudos', col_codigo_uc,total_horas_docente]]

# re-nomear colunas dfdsd
dfdsd = dfdsd.rename(columns={ column_total_horas_info: total_horas_uc_})

# com groupby, somar todas as linhas que correspondem à mesma UC
dfdsd[total_horas_docencia_uc] = dfdsd.groupby(col_codigo_uc)[soma_horas_docente_uc].transform('sum')
dfdsd[column_horas_em_falta_preencher]=dfdsd[total_horas_uc_]-dfdsd[total_horas_docencia_uc]

# Ok: 2 linhas
# dfdsd.loc[dfdsd[column_key]=='Fernando Eduardo Lagos Costa', [ 'Inserir docentes na UC ','Responsável', 'Nome da UC', 'ciclo de estudos', 'codigoUC',total_horas_docente]]

# funciona:
#dfdsd.loc[dfdsd['codigoUC']==1749, cols_horas+[soma_horas_docente_uc]]
#dfdsd.loc[dfdsd[soma_horas_docente_uc]>0, cols_horas+[soma_horas_docente_uc]].iloc[100]

# selecionar as linhas dos responsáveis das UCs
dfucs=dfdsd[dfdsd[colResponsavel]=='sim']
# os códigos de UCs não devem ter repetições
if dfucs[col_codigo_uc].duplicated().sum() != 0: 
    stop
dfucs=dfucs[colunas_FOS+[newcolResponsavel]+colunas_nome_UC+[total_horas_uc_,total_horas_docencia_uc,column_horas_em_falta_preencher]]

# eliminar as linhas dos responsáveis
dfdsd=dfdsd[dfdsd[colResponsavel]!='sim']


################# docentes
dfdsd.columns
# criar dfhd
horas_docentes=dfdsd.groupby(column_key)[soma_horas_docente_uc].sum()
horas_docentes.sum()
df_horas_docentes=horas_docentes.reset_index().rename(columns={'index': 'Nome', 0: soma_horas_docente_uc})
#approximate_merge(dfaux, horas_docentes, key_column1, key_column2)
dfhd=pd.merge(dfaux,df_horas_docentes,left_on=col_nome_completo,right_on=column_key, how='outer')
dfhd[soma_horas_docente_uc].sum()
# tirar para verificar merge how=outer:
dfhd=dfhd[[col_nome_completo,col_posicao,soma_horas_docente_uc]]
dfhd=dfhd.dropna(axis=0, how='any')
len(dfhd) # 138

###################################### juntar informação serviço externo aos docentes
olddf=dfhd.copy()
dfhd=olddf.copy()
horas_externo=dfexterno[dfexterno['DSD']==1].groupby('Nome do docente')['Número horas letivas'].sum()
df_horas_externo=horas_externo.reset_index().rename(columns={'Número horas letivas': soma_horas_externo_docente})
# merge dfhd com df_horas_externo
dfhd=pd.merge(dfhd,df_horas_externo,left_on=col_nome_completo,right_on='Nome do docente', how='left')
dfhd.columns
dfhd=dfhd[['Nome completo', 'Posição', 'Soma horas docente UC', 'Horas docência externa']]
dfhd.loc[dfhd['Horas docência externa'].isna(),['Horas docência externa']]=0
dfhd['Horas totais']= dfhd['Soma horas docente UC']+dfhd[soma_horas_externo_docente].astype(float)
# por semana
dfhd['Horas semanais']=dfhd['Horas totais'].map(lambda x: round(x/28,2))

#################### validação

# # comparar dfhd e dfdsd quando se agrupa horas por docente
# dfdoc=dfdsd.groupby(column_key)['Soma horas docente UC'].sum()
# dfdoc=pd.merge(dfdoc, dfhd, on=column_key)
# dfdoc['diff']=dfdoc['Soma horas docente UC_x']-dfdoc['Soma horas docente UC_y']
# dfdoc.sort_values(by='diff')
# dfdoc.sum(axis=0)
# len(dfdoc)
# len(dfhd)
# dfhd


# validação
# número de horas totais docência
# em dfdsd
print('número de horas totais docência dfdsd',dfdsd[soma_horas_docente_uc].sum())
# em dfhd
print('número de horas totais docência dfhd',dfhd[soma_horas_docente_uc].sum())
# número total de horas das UCs
print('número de horas das UCs',dfucs[total_horas_uc_].sum())
# total horas em falta
print('total horas em falta',dfucs[column_horas_em_falta_preencher].sum())
print('total horas em falta',dfucs[total_horas_uc_].sum()-dfdsd[soma_horas_docente_uc].sum())

############### output
wbr=openpyxl.Workbook()
wsr_ucs=wbr.create_sheet('horas_UCs')
wsr_docentes=wbr.create_sheet('horas_docentes')
wsr_ucs_docentes=wbr.create_sheet('horas_UCs_docentes')
wsr_info=wbr.create_sheet('info_UCs')

#cols_dfucs_drop=['keyUC',soma_horas_docente_uc, 'Nome', 'codigoUC']
cols_dfdsd=['Áreas Disicplinares', 'Departamento', 'Responsável', 'nomeResponsavel','Inserir docentes na UC ', 'Nome da UC', col_codigo_uc,soma_horas_docente_uc,column_horas_em_falta_preencher]
cols_dfinfo_drop=[]

df_to_excel_with_columns(dfucs, wsr_ucs)
df_to_excel_with_columns(dfdsd[cols_dfdsd], wsr_ucs_docentes)
df_to_excel_with_columns(dfhd, wsr_docentes)
df_to_excel_with_columns(dfinfo.drop(cols_dfinfo_drop,axis=1), wsr_info)

# Freeze the top row, and add filter 
for ws in [wsr_ucs,wsr_ucs_docentes,wsr_docentes,wsr_info]:
    #ws.protection.sheet = True
    #ws.protection.password = 'resumo_DSD_2023_2024'
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

if 'Sheet' in wbr.sheetnames:  # remove default sheet
    wbr.remove(wbr['Sheet'])


wbr.save(fnResumo)
wbr.close

if SAVE_HTML_MAT:
    output_columns=[col_codigo_uc,'Nome da UC', 'Áreas Disicplinares', 'Áreas Científicas (FOS)',  'ciclo de estudos','nomeResponsavel']
    # determinar UCs em que o responsável é da SM
    dfucs.columns
    dfucs.loc[dfucs['Áreas Disicplinares']=='MAT',['Nome da UC','nomeResponsavel']]
    docs_mat=list(dfucs.loc[dfucs['Áreas Disicplinares']=='MAT','nomeResponsavel'].unique())
    ucs_mat=list(dfucs.loc[dfucs['nomeResponsavel'].isin(docs_mat),'Nome da UC'].unique())
    # UCs coordenadas por docentes MAT
    df=dfucs.loc[dfucs['nomeResponsavel'].isin(docs_mat) ,output_columns]
    # determinar UCs em que o docente é da SM
    dfdsd.columns # 'Inserir docentes na UC '
    ucs_all_mat=dfdsd.loc[dfdsd['Inserir docentes na UC '].isin(list(docs_mat)) ,'Nome da UC'].unique()
    ucs_not_mat=list(set(ucs_all_mat).difference(set(ucs_mat)))
    df=pd.concat([df,dfucs.loc[dfucs['Nome da UC'].isin(ucs_not_mat) ,output_columns]],axis=0)
    df=df.sort_values(by=['ciclo de estudos','Nome da UC'])
    # corrigir nomes das colunas
    df=df.rename(columns={'nomeResponsavel': 'Nome Responsável'})
    df=df.rename(columns={'Áreas Disicplinares': 'Áreas Disciplinares'})
    html_table = df.to_html(index=False)
    # save HTML table to a file
    with open('UCs_SM.html', 'w',encoding='utf-8') as f:
        f.write(html_table)


# horas totais em cada ciclo
dfinfo.head()
dfinfo.duplicated

# número de UCs
len(dfinfo)
# número de horas de UCs pro ciclo
dfinfo.drop_duplicates().groupby('ciclo de estudos')['Total Horas previsto '].sum()

