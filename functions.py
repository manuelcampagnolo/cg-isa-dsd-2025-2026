from pathlib import Path
import pandas as pd
from unidecode import unidecode
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors

def insert_row_at_beginning(df, new_row_dict):
    """
    Inserts a new row at the beginning of a DataFrame.

    Parameters:
    - df: pandas DataFrame
    - new_row_dict: dict, a dictionary representing the new row to be inserted

    Returns:
    - df_updated: pandas DataFrame, the DataFrame with the new row inserted
    """
    df.loc[-1] = new_row_dict
    df.index = df.index + 1
    df.sort_index(inplace=True)
    return df


def insert_row_at_end(df, new_row_dict):
    """
    Inserts a new row at the end of a DataFrame.

    Parameters:
    - df: pandas DataFrame
    - new_row_dict: dict, a dictionary representing the new row to be inserted

    Returns:
    - df_updated: pandas DataFrame, the DataFrame with the new row inserted
    """
    # Create a DataFrame from the dictionary, filling missing values with NaN
    new_row_df = pd.DataFrame([new_row_dict]).reindex(columns=df.columns)

    # Concatenate the original DataFrame and the new row DataFrame
    df_updated = pd.concat([df, new_row_df], ignore_index=True)

    return df_updated


def add_suffix_to_duplicates(df, column_name):
    """
    Adds suffixes to duplicate values in the specified column of a DataFrame.

    Parameters:
    - df: pandas DataFrame
    - column_name: str, the name of the column with potential duplicates

    Returns:
    - df_updated: pandas DataFrame, the updated DataFrame
    """

    # Get a boolean mask for duplicated values in the specified column
    mask = df.duplicated(subset=[column_name], keep=False)

    # Create a dictionary to store the count of each duplicated value
    suffix_count = {}

    # Iterate through the DataFrame and add suffixes to duplicates
    for index, row in df.iterrows():
        value = row[column_name]
        if mask[index]:
            # This value is a duplicate
            count = suffix_count.get(value, 1)
            suffix = f'_{count}'
            suffix_count[value] = count + 1
            df.at[index, column_name] = f'{value}{suffix}'

    return df
# Example usage:
# updated_df = add_suffix_to_duplicates(your_dataframe, 'your_column_name')

def reorder_and_filter_dataframe(df, column_name, value_list):
    """
    Reorders a DataFrame based on a list of values and drops rows with column values not in the list.

    Parameters:
    - df: pandas DataFrame
    - column_name: str, the name of the column to reorder and filter
    - value_list: list, the list of values to use for ordering and filtering

    Returns:
    - df_reordered_and_filtered: pandas DataFrame, the reordered and filtered DataFrame
    """
    # Ensure the column values are of the same type as the values in the list
    df[column_name] = df[column_name].astype(type(value_list[0]))

    # Reorder the DataFrame based on the specified column and list of values
    df_reordered = df[df[column_name].isin(value_list)].sort_values(by=column_name, key=lambda x: x.map({val: i for i, val in enumerate(value_list)}))

    return df_reordered

# Example usage:
# Replace 'your_dataframe', 'your_column_name', and 'your_value_list' with the actual DataFrame, column name, and value list you are working with
# reordered_and_filtered_df = reorder_and_filter_dataframe(your_dataframe, 'your_column_name', your_value_list)

def unlock_cells(ws, col_name, min_row=None, max_row=None,fill_color=None,border=None):
    if min_row is None: min_row=1
    if max_row is None: max_row=ws.max_row
    idx=column_index_from_string(col_name)
    for row in ws.iter_rows(min_row, max_row, min_col=idx, max_col=idx):
        for cell in row:
            if fill_color is not None and border is not None: 
                cell.fill=fill_color
                cell.border=border
            cell.protection = Protection(locked=False)

def stripe_cells(ws,fill_color,border):
    count=0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if count%2==0:
            for cell in row:
                if fill_color is not None: 
                    cell.fill=fill_color
                    cell.border=border
        count+=1

def get_letter_from_column_name(df,col_name):
    return df.columns.get_loc(col_name) + 1, get_column_letter(df.columns.get_loc(col_name) + 1)

def compact_excel_file(input_file, output_file):
    tabs = pd.ExcelFile(input_file).sheet_names 
    #print(tabs)
    with pd.ExcelWriter(output_file) as writer: 
        for sheet_name in tabs:
            print(sheet_name)
            # Read the sheet into a pandas DataFrame
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            df.columns=simplify_strings(df.columns) 
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print('sheet', sheet_name, 'created')

def compact_workbook(source_workbook, output_file):
    new_workbook = openpyxl.Workbook()
    # Iterate through sheets in the source workbook
    for sheet_name in source_workbook.sheetnames:
        print('Sheet:', sheet_name)
        # Copy the sheet
        source_sheet = source_workbook[sheet_name]
        new_sheet = new_workbook.create_sheet(title=sheet_name)

        # Copy each cell and its value
        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1, max_col=source_sheet.max_column):
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value

        # Copy styles
        for column in range(1, source_sheet.max_column + 1):
            for row in range(1, source_sheet.max_row + 1):
                source_cell = source_sheet.cell(row=row, column=column)
                new_cell = new_sheet.cell(row=row, column=column)

                # Copy font, fill, and border attributes
                new_cell.font = copy(source_cell.font)
                new_cell.fill = copy(source_cell.fill)
                new_cell.border = copy(source_cell.border)

                # Copy number format and alignment
                new_cell.number_format = copy(source_cell.number_format)
                new_cell.alignment = copy(source_cell.alignment)

    # Apply filters to the first row of each sheet
    for sheet in new_workbook.sheetnames:
        new_workbook[sheet].auto_filter.ref = new_workbook[sheet].dimensions

    # Save the new workbook
    new_workbook.save(output_file)


def prints_dictionary_of_sheet_and_column_names(workbook, input_file, ws_names=None):
    ''' ws_names is a dict where values are the names of the workbook sheets'''
    if ws_names is None: 
        ws_names={sheet_name: sheet_name for sheet_name in workbook.sheetnames}
    for sheet_name in workbook.sheetnames:
        # Read the sheet into a pandas DataFrame
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        df.columns=simplify_strings(df.columns)
        # determine dictionary key that has value sheet_name 
        key=list(filter(lambda x: ws_names[x] == sheet_name, ws_names))[0]
        print(key,'=', {key+'_'+x[0:6] : x for x in df.columns})

def simplify_strings(s):
    ''' removes accents and replaces spaces by _ ; requires unidecode'''
    return list(map(lambda x: x.replace(' ','_'),list(map(unidecode, s))))

def df_to_excel(df, ws, header, index, startrow, startcol):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def df_to_excel_with_columns(df,ws,maxwidth=30,header=True,index=False,startrow=0, startcol=0):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = max(len(column)-2,min(max_length+2, maxwidth)) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_excel(df,ws,header,index, startrow, startcol)

