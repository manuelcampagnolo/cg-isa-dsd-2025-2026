import openpyxl
from pathlib import Path

# Load the source workbook
#input_folder=Path(r'C:\Users\mlc\OneDrive - Universidade de Lisboa\Documents\profissional-isa-cv\cg-isa\DSD_2024_2025\backup_inputs_DSD')
try:
    working_dir=Path(__file__).parent.parent # working directory from script location: scripts are in 'scripts' folder
except:
    working_dir=Path().absolute()

FOLDER_SERVICOS= 'ficheiros_servicos_ISA' #'ficheiros_servicos_ISA'
DSD_INPUT_FICH='info_servicos_jan_2025.xlsx' #'2024_01_26 DSD_inform_202324_v6-1-1.xlsx (Dados MCaron e Carlos)_compact_ML3.xlsx'
DSD_INPUT_FICH='2024_01_26 DSD_inform_202324_v6-1-1.xlsx (Dados MCaron e Carlos)_compact_ML3.xlsx'
FOLDER_ANO = 'DSD_2024_2025'
input_folder= working_dir / FOLDER_ANO / FOLDER_SERVICOS
file_name= input_folder  / DSD_INPUT_FICH

def print_excel_info(file_name):

    print('File name:', file_name)

    # Load the workbook
    wb = openpyxl.load_workbook(file_name)
    
    # Print worksheet names
    print("Worksheet names:", wb.sheetnames)
    
    # Iterate through each worksheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nWorksheet: {sheet_name}")
        print("First row values:")
        
        # Print values of the first row
        for cell in sheet[1]:
            print(f"{cell.column_letter}: {cell.value}")

# Example usage
print_excel_info(file_name)
