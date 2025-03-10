# copyright Manuel Campagnolo 2025

# input:
# ficheiro dos serviços com info sobre RH, UCs etc e metadados
# parâmetros: data final contrato, ...
# output:
# ficheiro a partilhar para colocação dos nomes de responsáveis UCs e comentários: 'DSD_2024_2025_responsaveis_UCs.xlsx'


from copy import copy
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname # https://openpyxl.readthedocs.io/en/latest/validation.html?highlight=validation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from pathlib import Path
import pandas as pd
import re
from unidecode import unidecode
from functions import simplify_strings, df_to_excel_with_columns, compact_excel_file, get_letter_from_column_name,unlock_cells,stripe_cells
from functions import add_suffix_to_duplicates, reorder_and_filter_dataframe, insert_row_at_beginning, insert_row_at_end,sort_list
import sys


UNPROTECT_OUTPUT_CELLS=True # para desproteger células com função unlock_cells(ws, col_name, min_row=None, max_row=None):
PROTECT_WORKSHEET=True # tem que ser True para impedir escrita
PASSWORD='kathleen'

# Load the source workbook
#input_folder=Path(r'C:\Users\mlc\OneDrive - Universidade de Lisboa\Documents\profissional-isa-cv\cg-isa\DSD_2024_2025\backup_inputs_DSD')
try:
    working_dir=Path(__file__).parent.parent # working directory from script location: scripts are in 'scripts' folder
except:
    working_dir=Path().absolute()

ANO='2025-2026'
# FICHEIROS E PASTAS
FOLDER_ANO = 'DSD_2025_2026'
DSD_INPUT_FICH_stem='info_servicos_jan_2025_v3' #'info_servicos_jan_2025.xlsx' #'2024_01_26 DSD_inform_202324_v6-1-1.xlsx (Dados MCaron e Carlos)_compact_ML3.xlsx'
stem=DSD_OUTPUT_FICH_stem='DSD_2025_2026_coords_UCs' # 
suffix=DSD_OUTPUT_FICH_suffix='.xlsx'
FOLDER_SERVICOS= 'ficheiros_servicos_ISA' #'ficheiros_servicos_ISA'
FOLDER_OUTPUT='output_files'
# NOME FICHEIRO COMPACTO
COMPACT='_compact'
BLOQ='_bloq' # bloqueado com password
# list files in FOLDER_OUTPUT and create name of output file v1, v2, v3,...
folder_output=working_dir/FOLDER_ANO/FOLDER_OUTPUT
number_files=len(list(folder_output.rglob(stem+"*")))
DSD_OUTPUT_FICH_version=stem+'_v'+str(number_files+1)+suffix 
DSD_OUTPUT_FICH_version_bloqued=stem+'_v'+str(number_files+1)+BLOQ+suffix 

# CAMINHOS
input_folder= working_dir / FOLDER_ANO / FOLDER_SERVICOS
output_folder= working_dir / FOLDER_ANO / FOLDER_OUTPUT
input_file = input_folder/ (DSD_INPUT_FICH_stem+suffix)
compact_file= output_folder  / (DSD_INPUT_FICH_stem+COMPACT+suffix) # intermediate output (small xlsx input file)
if PROTECT_WORKSHEET: 
    output_file=output_folder / DSD_OUTPUT_FICH_version
else:
    output_file=output_folder / DSD_OUTPUT_FICH_version_bloqued

# worksheets and column names do ficheiro (único) dos serviços que já contém UCs e RHs para DSD 2024-2025
# prefixo=são as folhas do excel dos serviços DSD_INPUT_FICH: UC, UCMETA, etc
# sufixo _etc:= são as colunas que interessam da folha respetiva 
UC ='listagem_UCs' #'uc_2024-25'
N_extra_UCs=30 # para permitir introduzir novas UCs
UC_codigo='codigo_uc'
UC_uc = 'nome_uc_pt' # 'unidade_curricular'
#UC_area_cientifica='area_uc' #'area_cient' # drop
UC_ciclo_curso= 'ciclo_uc' # "2º Ciclo (M)"" ou "1º Ciclo (L)" ou "3º Ciclo (D)"
UC_ciclo_curso_curso='curso' # para filtrar cncg's
UC_tipo='tipo_externo'
UC_ects='ects'
UC_tipo_externo='externo' ## valor do atributo quando UC é externa: pode ser Normal, externa, 'Trabalho de projeto',...
UC_dept='dept_uc' # DCEB, DRAT, EXTISA
UC_dept_extisa='EXTISA' ## valor do atributo para filtar UC externas ao ISA
UC_atrib_select=[UC_codigo,UC_uc,UC_ciclo_curso,UC_ects,UC_tipo]
# novos atributos UC:
UC_resp='coordenador_UC'
UC_sugestoes='sugestões de modificação da info da UC'  # ver ACD 
UC_autor_sugestao='autor da sugestão'
#UC_numero_alunos='NumeroAlunos' # drop
#UCMETA='uc_meta'

# RH
N_extra_nomes=20
RH='docentes_mar_2025'#'docentes_dez_2024'
RH_nome='nome_doc' # nomes docentes
RH_numero = 'numero_doc'
RH_data_termino='data_termino'
RH_posicao='categ_doc'
RH_grupo='grupo_doc' # investigadores, docentes, não docente, ...
RH_dept_doc	= 'dept_doc' # CEF, DCEB, DRAT, CEABN
RH_seccao_doc= 'seccao_doc'
RH_contrato='contrato'
RH_obs = 'contrato' # poderia ser outro: escolhe-se o mais relevante para o DSD_coords
RN_contrato_convidado='Docente Convidado' # valor na colune 'contrato' 2025-2026
RH_atrib_select=[RH_nome,RH_numero,RH_posicao,RH_grupo,RH_dept_doc,RH_seccao_doc, RH_obs,RH_data_termino]
# values: Sheet_column_value
RH_nome_pro_bono='docente_PRO_BONO' 
RH_data_limite='data_limite'
# RH_data_fim_sem_termo='sem termo'
DATA_TERMINO='30/09/2025' #'30/09/2025'

# Bolseiros

############################### folhas do input a considerar
sheet_names=[RH,UC]

# desproteger células:
# responsável
# docentes extra + observações

'''
, 'UC_meta': 'uc_meta', 'AC': 'uc_AreaCient', 'AC_meta': 'AC_meta', 'CC': 'cod_curso', 'RH': 'RH', 'RH_meta': 'RH_meta', 'POS': 'RH_posicao'}
UC = {'UC_ciclo_': 'ciclo_curso', 'UC_uc_obr': 'uc_obrigatoria', 'UC_uc_opt': 'uc_optativa', 'UC_cod_uc': 'cod_uc', 'UC_unidad': 'unidade_curricular', 'UC_Respon': 'Responsavel', 'UC_não_fu': 'não_funciona_2024_2025', 'UC_area_c': 'area_cient', 'UC_ano': 'ano', 'UC_sem': 'sem', 'UC_h_trab': 'h_trab_totais', 'UC_T': 'T', 'UC_TP': 'TP', 'UC_PL': 'PL', 'UC_TC': 'TC', 'UC_S': 'S', 'UC_E': 'E', 'UC_OT': 'OT', 'UC_O': 'O', 'UC_h_cont': 'h_contacto_total', 'UC_ECTS': 'ECTS', 'UC_uc_int': 'uc_interna', 'UC_obs': 'obs'}
UC_meta = {'UC_meta_ciclo_': 'ciclo_curso', 'UC_meta_1 cicl': '1 ciclo', 'UC_meta_Licenc': 'Licenciatura'}
AC = {'AC_area_c': 'area_cientifica', 'AC_sigla': 'sigla', 'AC_classi': 'classificacao_area'}
AC_meta = {'AC_meta_classi': 'classificacao_area', 'AC_meta_CNAEF': 'CNAEF', 'AC_meta_Classi': 'Classificação Nacional das Áreas de Educação e Formação (CNAEF), Portaria n.o 256/2005 de 16 de Março. Classificação: nível 1, 2 e 3 (código: 3 dígitos)'}
CC = {'CC_cod_cu': 'cod_curso', 'CC_grau': 'grau', 'CC_curso': 'curso', 'CC_obs': 'obs'}
RH = {'RH_num_pe': 'num_pessoal', 'RH_nome': 'nome', 'RH_corpo': 'corpo', 'RH_posica': 'posicao', 'RH_ETI': 'ETI', 'RH_contra': 'contratacao_vinculo', 'RH_data_f': 'data_fim', 'RH_Obs': 'Obs'}
RH_meta = {'RH_meta_num_pe': 'num_pessoal', 'RH_meta_número': 'número de indentificação pessoal atribuído pelos RH', 'RH_meta_Unname': 'Unnamed: 4'}
RH_posicao = {'POS_posica': 'posicao', 'POS_h_min': 'h_min', 'POS_h_max': 'h_max', 'POS_obs': 'obs'}
''' 

# cor light red, light yellow
fill_red = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
fill_green = PatternFill(start_color="C0FFCB", end_color="C0FFCB", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # alpha: 1st 2 characters
fill_light_yellow = PatternFill(start_color="FFFFED", end_color="FFFFED", fill_type="solid") 
thin_border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

##################################################################### Load the source workbook
# source_workbook = load_workbook(input_file, read_only=True, data_only=True)
# Try to read smaller file; otherwise read original file and create smaller file
try:
    #source_workbook = load_workbook(compact_input_file)
    dummy = pd.ExcelFile(compact_file).sheet_names
except:
    print('ler ficheiro original não compactado')
    compact_excel_file(input_file, compact_file)
    print('ficheiro original compactado')
    existing_sheet_names = pd.ExcelFile(compact_file).sheet_names

#sys.exit('compact criado')

# Create a new workbook
#new_workbook = pd.ExcelWriter(output_file, engine='openpyxl')
new_workbook = Workbook()

# contar número de docentes em RH
#df = pd.read_excel(compact_file, sheet_name=RH)
#idx_RH_nome,letter_RH_nome=get_letter_from_column_name(df,RH_nome)  #df.columns.get_loc(RH_nome) + 1
#idx_RH_posicao,letter_RH_posicao=get_letter_from_column_name(df,RH_posicao)  #df.columns.get_loc(RH_nome) + 1
#idx_RH_obs,letter_RH_obs=get_letter_from_column_name(df,RH_obs)  #df.columns.get_loc(RH_nome) + 1
# adicionar N_extra_nomes nomes em branco em RH_home
# adicionar col
    
# contar número de Ucs
#df = pd.read_excel(compact_file, sheet_name=UC)
#numero_ucs=len(df[UC_uc])
#idx_UC_uc,letter_UC_uc=get_letter_from_column_name(df,UC_uc) 
#idx_uc=df.columns.get_loc(UC_uc) + 1

# Iterate through sheets in the source workbook
# Ensure that RH comes before UC and remove 'planos_estudos', etc
#sheet_names=[RH,RHMETA, UC,UCMETA]+list(set(sheet_names).difference(set([RH,RHMETA, UC,UCMETA,PE,AC,ACMETA,POS, RHPOSICAO, UCAREA])))
for sheet_name in sheet_names: #source_workbook.sheetnames:
    # for each sheet_name, we copy the contents of the input sheet, create a df, modify df, create validation, and write to workbook with df_to_excel_with_columns
    print(sheet_name) 
    # Create a new sheet in the new workbook
    new_worksheet = new_workbook.create_sheet(title=sheet_name)

    # Read the sheet into a pandas DataFrame
    df = pd.read_excel(compact_file, sheet_name=sheet_name)
    # removes accents and replaces ' ' by '_', convert to lowercase
    df.columns=simplify_strings(df.columns)

    # ordenar docentes por ordem alfabética, com os novos docentes à cabeça, mais "docente_PRO_BONO"; excluir docentes a termo, com termo antes de set 2024
    # Nota: 'docentes em contratação' podem ter o mesmo nome e por isso é preciso lidar com duplicações
    if sheet_name==RH:
        df=df[RH_atrib_select]
        # clean df
        df=insert_row_at_beginning(df,{RH_nome: RH_nome_pro_bono,  RH_obs: 'Docente ou especialista não do ISA que participa na docência sem receber pagamento do ISA: o nome do docente pode ser indicado na coluna de observações'})
        df=add_suffix_to_duplicates(df,RH_nome)
        #docentes_em_contratacao=list(df[df[RH_nome].str.contains(RH_nome_em_contratacao, case=False, na=False)][RH_nome])
        # remover docentes com contrato que acaba até 1 de setembro de 2024
        #sem_termo=list(df[df[RH_data_fim].str.contains(RH_data_fim_sem_termo, case=False, na=False)][RH_nome])
        #com_termo=list(set(list(df[RH_nome])).difference(set(sem_termo)))
        #df_com_termo=df[df[RH_nome].isin(com_termo)]
        #df_set_2024=df_com_termo[pd.to_datetime(df_com_termo[RH_data_fim]) > DATA_TERMO_CERTO]
        #com_termo_set_2024=list(df_set_2024[RH_nome])
        # criar lista em ordem alfabética de docentes que não estão em contratação
        #L=list(set(list(sem_termo+com_termo_set_2024)).difference(set(docentes_em_contratacao).union(set([RH_nome_pro_bono]))))
        #outros_docentes=sort_list(L, simplify_strings(L))
        # list de nomes de todos os potenciais docentes 
        #todos_docentes=[RH_nome_pro_bono]+docentes_em_contratacao+outros_docentes
        # verificar data de fim de contrato # anterior a 30/09/2025
        # usar RH_data_termino:
        df[RH_data_termino] = df[RH_data_termino].astype(str).str.replace('/', '-')
        df['extracted_date'] = df[RH_data_termino].str.extract(r'(\d{4}-\d{2}-\d{2})', expand=False)
        # Convert extracted dates to datetime
        df['date'] = pd.to_datetime(df['extracted_date'], format='%Y-%m-%d', errors='coerce')
        df[RH_data_limite] = df['date'].astype(str)
        # Create the boolean column
        df['before_DATA_TERMINO'] = df['date'] >= pd.to_datetime(DATA_TERMINO, format='%d/%m/%Y')
        # Set True for empty 'data_termino' values
        df.loc[df[RH_data_termino].isnull() | df[RH_data_termino].isna() | (df[RH_data_termino] == '') | (df[RH_data_termino].astype(str)=='nan'), 'before_DATA_TERMINO'] = True
        df.loc[df[RH_data_termino].isnull() | df[RH_data_termino].isna() | (df[RH_data_termino] == '') | (df[RH_data_termino].astype(str)=='nan'), RH_data_limite] = ''
        df['before_DATA_TERMINO'] = df['before_DATA_TERMINO'].fillna(True)
        df[RH_data_limite] = df[RH_data_limite].fillna('')
        print(df[[RH_data_termino,'extracted_date','date','before_DATA_TERMINO',RH_data_limite]].head(20))
        potenciais_docentes=df[df['before_DATA_TERMINO']][RH_nome]
        # Drop temporary columns
        df = df.drop(['extracted_date', 'date',RH_data_termino,'before_DATA_TERMINO'], axis=1)
        # end 
        
        # ordenar df segundo lista todos_docentes
        df=reorder_and_filter_dataframe(df, RH_nome, potenciais_docentes)
        # quando df está na forma final contar número docentes
        numero_docentes=df.shape[0]
        print('numero docentes:',numero_docentes)
        # colunas relevantes da tabela
        if RH_nome in df.columns:
            idx_RH_nome,letter_RH_nome=get_letter_from_column_name(df,RH_nome)  #df.columns.get_loc(RH_nome) + 1
        if RH_posicao in df.columns:
            idx_RH_posicao,letter_RH_posicao=get_letter_from_column_name(df,RH_posicao)  #df.columns.get_loc(RH_nome) + 1
        if RH_obs in df.columns:
            idx_RH_obs,letter_RH_obs=get_letter_from_column_name(df,RH_obs)  #df.columns.get_loc(RH_nome) + 1
        print('nome', idx_RH_nome, letter_RH_nome)
        print('posicao', idx_RH_posicao, letter_RH_posicao)
        print('obs', idx_RH_obs, letter_RH_obs)
        # acrescentar novas linhas em docentes para eventuais docentes não listados se N_extra_nomes>0
        for i in range(N_extra_nomes):
            df=insert_row_at_end(df,{RH_nome: 'Outro_docente_'+str(i+1), RH_posicao:'Categoria/Departamento', RH_obs: 'Justificação. Estas linhas devem ser preenchidas apenas pelos presidentes dos Dpts.' })
    
    # Criar drop-down menu para inserir nome responsável da UC
    if sheet_name==UC:
        df=df[UC_atrib_select]
        idx_UC_uc,letter_UC_uc=get_letter_from_column_name(df,UC_uc) 
        # remover coluna área cientifica, etc
        # df=df.drop(columns=[UC_area_cientifica,UC_numero_alunos])
        # re-ordenar UCs pelo nome, mas com cursos CNCG no fim
        df=add_suffix_to_duplicates(df,UC_uc)
        #cncg=list(df[df[UC_ciclo_curso].str.contains(UC_ciclo_curso_curso, case=False, na=False)][UC_uc])
        #cncg=sort_list(cncg, simplify_strings(cncg))
        #uc_ciclos=list(set(list(df[UC_uc])).difference(set(cncg)))
        #uc_ciclos=sort_list(uc_ciclos, simplify_strings(uc_ciclos))
        df=reorder_and_filter_dataframe(df, UC_uc, df[UC_uc]) # re-orders by UC_uc, uses all values (since 3rd arg is also UC_uc)
        # Criar coluna responsável
        if UC_resp not in df.columns: 
            df.insert(idx_UC_uc-1,UC_resp,'') # automatizar .A largura da coluna tem que ser grande para se verem os nomes
        idx_UC_resp,letter_UC_resp=get_letter_from_column_name(df,UC_resp)
        print(df, idx_UC_uc, UC_resp,idx_UC_resp,letter_UC_resp) # B
        # Criar coluna 'sugestões'
        if UC_sugestoes not in df.columns: 
            df.insert(df.shape[1],UC_sugestoes,'') 
        idx_UC_sugestoes,letter_UC_sugestoes=get_letter_from_column_name(df,UC_sugestoes) 
        # Criar coluna 'autor_sugestão'
        if UC_autor_sugestao not in df.columns: 
            df.insert(df.shape[1],UC_autor_sugestao,'') 
        idx_UC_autor_sugestao,letter_UC_autor_sugestao=get_letter_from_column_name(df,UC_autor_sugestao) 
        # UC_ciclo_curso
        idx_UC_ciclo_curso,letter_UC_ciclo_curso=get_letter_from_column_name(df,UC_ciclo_curso) 
        idx_UC_uc,letter_UC_uc=get_letter_from_column_name(df,UC_uc) 
        # quando df está na forma final contar número UCs
        numero_ucs=df.shape[0]
        print('numero UCs:',numero_ucs) 
        # para drop-down, ante sde acrescentar novas linhas 
        lista_ciclos=sorted(list(set(df[UC_ciclo_curso])))
        # dar a possibilidade de adicionar outras UCs
        for i in range(N_extra_UCs):
            df=insert_row_at_end(df,{UC_uc: 'Nome_UC_em_falta_'+str(i+1), UC_ciclo_curso:'Indicar ciclo UC em falta', UC_sugestoes: 'Linha a preencher apenas pelos presidentes dos Dpts' })
        # validation responsáveis - precisa de aceder a RH_nome
        dv = DataValidation(type="list", formula1=f"{quote_sheetname(RH)}!${letter_RH_nome}$2:${letter_RH_nome}${numero_docentes+1+N_extra_nomes}") # 10 extra
        new_worksheet.add_data_validation(dv)
        dv.add(f"${letter_UC_resp}$2:{letter_UC_resp}${numero_ucs+N_extra_UCs+1}") # creates drop-down menu. #automatizar
        # validation autor_sugestao
        dv.add(f"${letter_UC_autor_sugestao}$2:{letter_UC_autor_sugestao}${numero_ucs+N_extra_UCs+1}") # creates drop-down menu.
        # validation ciclo nova UC
        formula = '"{}"'.format(','.join(lista_ciclos))
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        new_worksheet.add_data_validation(dv)
        dv.add(f"${letter_UC_ciclo_curso}${numero_ucs+2}:{letter_UC_ciclo_curso}${numero_ucs+N_extra_UCs+1}") # creates drop-down menu.


    # Write the DataFrame to the new workbook
    #df.to_excel(new_workbook, sheet_name=sheet_name, index=False, startrow=0, header=True)
    df_to_excel_with_columns(df,new_worksheet,maxwidth=20,header=True,index=False,startrow=0, startcol=0)

    # Apply filters to the first row exceto se a sheet_name contém '_meta'
    if '_meta' not in sheet_name:
        new_worksheet.auto_filter.ref = new_worksheet.dimensions
        new_worksheet.freeze_panes = "A2"
    
    # a ideia é desbloquear algumas células e a seguir bloquear toda a worksheet
    # desbloquear células dos responsáveis e das sugestões das UC
    # células que vão ficar a verde:
    if sheet_name==UC and UNPROTECT_OUTPUT_CELLS:
        stripe_cells(new_worksheet, fill_color=fill_light_yellow,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_resp,fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_sugestoes,fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_autor_sugestao,fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_ciclo_curso, min_row=numero_ucs+2, max_row=numero_ucs+N_extra_UCs+1, fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_uc, min_row=numero_ucs+2, max_row=numero_ucs+N_extra_UCs+1, fill_color=fill_green,border=thin_border)
    
    # Dar possibilidade de criar novos docentes
    # apenas se N_extra_nomes>0
    if sheet_name==RH and UNPROTECT_OUTPUT_CELLS:
        stripe_cells(new_worksheet, fill_color=fill_light_yellow,border=thin_border)
        # docentes extra (a poderem ser adicionados)
        unlock_cells(new_worksheet,letter_RH_nome, min_row=numero_docentes+2, max_row=numero_docentes+N_extra_nomes+1, fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_RH_posicao, min_row=numero_docentes+2, max_row=numero_docentes+N_extra_nomes+1, fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_RH_obs, min_row=numero_docentes+2, max_row=numero_docentes+N_extra_nomes+1, fill_color=fill_green,border=thin_border)

    if PROTECT_WORKSHEET:
        new_worksheet.protection.sheet = True
        new_worksheet.protection.enable()
        new_worksheet.protection = SheetProtection(sheet=True, objects=False, scenarios=False, formatCells=False, formatRows=False, formatColumns=False, insertColumns=True, insertRows=True, insertHyperlinks=True, deleteColumns=True, deleteRows=True, selectLockedCells=False, selectUnlockedCells=False, sort=False, autoFilter=False, pivotTables=True, password=None, algorithmName=None, saltValue=None, spinCount=None, hashValue=None) #(autoFilter=True, formatColumns=True)

# eliminar a worksheet 'Sheet' que foi criada automaticamente
if 'Sheet' in new_workbook.sheetnames:  # remove default sheet
    new_workbook.remove(new_workbook['Sheet'])

# trocar ordem 
new_workbook._sheets = [new_workbook._sheets[i] for i in [1,0]]

# workbook protection
new_workbook.security.workbookPassword = PASSWORD
new_workbook.security.lockStructure = True

# Save the new workbook
try: 
    new_workbook.save(output_file)
except PermissionError:
    print('File must be in use: close if first please')
new_workbook.close