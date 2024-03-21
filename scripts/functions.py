from pathlib import Path
import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz # compare strings
from copy import copy
from datetime import datetime
from unidecode import unidecode
import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.styles import Border, Side


def approximate_merge(df1, df2, key_column1, key_column2):
    merged_data = pd.merge(df1, df2, left_on=key_column1, right_on=key_column2, how='outer')

    highest_score = 0

    for i, row in merged_data.iterrows():
        key1 = row[key_column1]
        key2 = row[key_column2]
        similarity_score = fuzz.ratio(str(key1), str(key2))
        
        if similarity_score > highest_score:
            highest_score = similarity_score

    merged_data = merged_data[merged_data.apply(lambda row: fuzz.ratio(str(row[key_column1]), str(row[key_column2])) == highest_score, axis=1)]

    return merged_data

def replace_value(df, first_column_name, first_value, second_column_name, second_value):
    # Replace the value in the second column where the first column matches the first value
    df.loc[df[first_column_name] == first_value, second_column_name] = second_value
    return df


def inserir_docente(df, column_codigo_UC_preencher, input_value,soma_horas_docente_uc,horas,column_key,docente,colResponsavel):
    # Get the index of the first row where the column matches the input value
    first_match_index = df.index[df[column_codigo_UC_preencher] == input_value].min()
    # Get the first row where the column matches the input value
    row = df.loc[first_match_index].copy()
    # muda valores
    row.loc[soma_horas_docente_uc]=horas
    row.loc[column_key]=docente
    row.loc[colResponsavel]=''
    # Append the first match row to the DataFrame
    df.loc[df.index.max() + 1] = row
    return df


def fill_empty_cells(df, col_name):
    # Create a copy of the dataframe to avoid modifying the original
    filled_df = df.copy()
    
    # Create a mask of empty cells in the column
    empty_mask = filled_df[col_name].isna()
    
    # Iterate over each row in the column
    for i, cell_value in enumerate(filled_df[col_name]):
        # If the cell is empty, fill it with the value of the closest non-empty cell above
        if pd.isna(cell_value):
            j = i - 1
            while j >= 0 and pd.isna(filled_df.loc[j, col_name]):
                j -= 1
            if j >= 0:
                filled_df.loc[i, col_name] = filled_df.loc[j, col_name]           
    return filled_df

# def df_to_excel_simple(df, ws, header=True, index=True):
#     for r in dataframe_to_rows(df, index=index, header=header):
#         ws.append(r)

def create_workbook_from_dataframe(df):
    """
    1. Create workbook from specified pandas.DataFrame
    2. Adjust columns width to fit the text inside
    3. Make the index column and the header row bold
    4. Fill background color for the header row

    Other beautification MUST be done by usage side.
    """
    workbook = openpyxl.Workbook()
    ws = workbook.active

    rows = dataframe_to_rows(df.reset_index(), index=False)
    col_widths = [0] * (len(df.columns) + 1)
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):

            if type(val) is str:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            elif hasattr(val, "sort"):
                cell = ws.cell(row=i, column=j, value=", ".join(list(map(lambda v: str(v), list(val)))))
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            else:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val)) + 1])

            # Make the index column and the header row bold
            if i == 1 or j == 1:
                cell.font = Font(bold=True)

    # Adjust column width
    for i, w in enumerate(col_widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = w

    return workbook

def set_border(ws):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows()
    for row in rows:
        for cell in row:
            cell.border = border


def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def df_to_excel_with_columns(df,ws,maxwidth=30,header=True,index=False):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = max(len(column)-2,min(max_length+2, maxwidth)) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_excel(df,ws,header=True,index=False)

# devolve letra da coluna com nome (1a linha) da worksheet ws
def nomeColuna2letter(ws,nome):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx=headers.index(nome)
    return get_column_letter(idx+1)


def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def create_workbook_from_dataframe(df):
    """
    1. Create workbook from specified pandas.DataFrame
    2. Adjust columns width to fit the text inside
    3. Make the index column and the header row bold
    4. Fill background color for the header row

    Other beautification MUST be done by usage side.
    """
    workbook = openpyxl.Workbook()
    ws = workbook.active

    rows = dataframe_to_rows(df.reset_index(), index=False)
    col_widths = [0] * (len(df.columns) + 1)
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):

            if type(val) is str:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            elif hasattr(val, "sort"):
                cell = ws.cell(row=i, column=j, value=", ".join(list(map(lambda v: str(v), list(val)))))
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            else:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val)) + 1])

            # Make the index column and the header row bold
            if i == 1 or j == 1:
                cell.font = Font(bold=True)

    # Adjust column width
    for i, w in enumerate(col_widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = w

    return workbook

def set_border(ws):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows()
    for row in rows:
        for cell in row:
            cell.border = border

# devolve letra da coluna com nome (1a linha) da worksheet ws
def nomeColuna2letter(ws,nome):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx=headers.index(nome)
    return get_column_letter(idx+1)

# copy column named nameCol to a a column located at colLetter
def copyColumn(ws,nameCol,colLetter,colType=None):
    count=0
    for cell in ws['{}:{}'.format(nomeColuna2letter(ws,nameCol),nomeColuna2letter(ws,nameCol))]:
        count+=1
        if colType=='numeric': # tenta converter para numérico
            try: 
                new_cell=ws.cell(row=cell.row, column=column_index_from_string(colLetter))
                new_cell.value=int(cell.value)
                new_cell.number_format='0'
            except:
                ws.cell(row=cell.row, column=column_index_from_string(colLetter), value=cell.value)
        else:
            ws.cell(row=cell.row, column=column_index_from_string(colLetter), value=cell.value)

def copy_sheet(source_sheet, target_sheet, idx):
    rows_resp=copy_cells(source_sheet, target_sheet, idx)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)
    return rows_resp

def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = False #copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

# idx indica a coluna em que estão os responsáveis das UCs
def copy_cells(source_sheet, target_sheet,idx):
    delta=0
    rows_resp=[]
    for r, row in enumerate(source_sheet.iter_rows()):
        #print('main r',r)
        # linhas em que há docente indicado
        if source_sheet.cell(r+1,idx+1).value: 
            #print(source_sheet.cell(r+1,idx+1).value)
            # ou 1a linha ou linhas dos responsáveis:
            if r==0: # 1a linha do ficheiro
                copy_first_row(r+delta,row, source_sheet, target_sheet)
            else:
                # linhas dos responsáveis
                copy_row(r+delta,row, source_sheet, target_sheet,fill=True)
                rows_resp.append(r+delta)
                for k in range(N):
                    delta+=1
                    partial_copy_row(r+delta,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao) # row comes from source; r+delta is the index in target
        else:
            # linhas em branco
            copy_row(r+delta,row, source_sheet, target_sheet, fill=False)
        #if r>40: break #<----------------------------------------------------------------------  remover
    print('delta: ',delta)
    return rows_resp

def copy_first_row(r,row, source_sheet, target_sheet):
     for c, cell in enumerate(row):
        source_cell = cell
        if c <= headers.index(colunas_red[-1]): # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                #target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def copy_row(r,row, source_sheet, target_sheet, fill):
    for c, cell in enumerate(row):
        if c==headers.index(colunas_red[-1]):
            # define default style
            red_cell=cell
    for c, cell in enumerate(row):
        source_cell = cell
        if c <= headers.index(colunas_red[-1]): # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            if fill:
                target_cell._value = source_cell._value
                target_cell.data_type = source_cell.data_type
                target_cell.font = copy(red_cell.font)
                target_cell.border = copy(red_cell.border)
                target_cell.fill = copy(red_cell.fill)
                target_cell.number_format = copy(red_cell.number_format)
                #target_cell.protection = copy(red_cell.protection)
                target_cell.alignment = copy(red_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def partial_copy_row(r,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao):
    for c, cell in enumerate(row):
        if c==headers.index(colunas_red[-1]):
            # define default style
            red_cell=cell
    for c, cell in enumerate(row):
        source_cell = cell
        if headers[c] in colunas_red: # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell.font = copy(red_cell.font)
            target_cell.border = copy(red_cell.border)
            target_cell.fill = copy(red_cell.fill)
            target_cell.number_format = copy(red_cell.number_format)
            #target_cell.protection = copy(red_cell.protection)
            target_cell.alignment = copy(red_cell.alignment)
        if headers[c] in columns_to_copy:  
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                #target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
        if headers[c]==coluna_validacao: # drop-down menus
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = VALIDATION_VALUE
            target_cell.data_type='s'
            #target_cell.protection = Protection(locked=False)


def sort_list(list1, list2):
    zipped_pairs = zip(list2, list1)
    z = [x for _, x in sorted(zipped_pairs)]
    return z

def insert_row_at_beginning(df, new_row_dict):
    """
    Inserts a new row at the beginning of a DataFrame.

    Parameters:
    - df: pandas DataFrame
    - new_row_dict: dict, a dictionary representing the new row to be inserted

    Returns:
    - df_updated: pandas DataFrame, the DataFrame with the new row inserted
    """
    df.loc[-1] = new_row_dict
    df.index = df.index + 1
    df.sort_index(inplace=True)
    return df


def insert_row_at_end(df, new_row_dict):
    """
    Inserts a new row at the end of a DataFrame.

    Parameters:
    - df: pandas DataFrame
    - new_row_dict: dict, a dictionary representing the new row to be inserted

    Returns:
    - df_updated: pandas DataFrame, the DataFrame with the new row inserted
    """
    # Create a DataFrame from the dictionary, filling missing values with NaN
    new_row_df = pd.DataFrame([new_row_dict]).reindex(columns=df.columns)

    # Concatenate the original DataFrame and the new row DataFrame
    df_updated = pd.concat([df, new_row_df], ignore_index=True)

    return df_updated


def add_suffix_to_duplicates(df, column_name):
    """
    Adds suffixes to duplicate values in the specified column of a DataFrame.

    Parameters:
    - df: pandas DataFrame
    - column_name: str, the name of the column with potential duplicates

    Returns:
    - df_updated: pandas DataFrame, the updated DataFrame
    """

    # Get a boolean mask for duplicated values in the specified column
    mask = df.duplicated(subset=[column_name], keep=False)

    # Create a dictionary to store the count of each duplicated value
    suffix_count = {}

    # Iterate through the DataFrame and add suffixes to duplicates
    for index, row in df.iterrows():
        value = row[column_name]
        if mask[index]:
            # This value is a duplicate
            count = suffix_count.get(value, 1)
            suffix = f'_{count}'
            suffix_count[value] = count + 1
            df.at[index, column_name] = f'{value}{suffix}'

    return df
# Example usage:
# updated_df = add_suffix_to_duplicates(your_dataframe, 'your_column_name')

def reorder_and_filter_dataframe(df, column_name, value_list):
    """
    Reorders a DataFrame based on a list of values and drops rows with column values not in the list.

    Parameters:
    - df: pandas DataFrame
    - column_name: str, the name of the column to reorder and filter
    - value_list: list, the list of values to use for ordering and filtering

    Returns:
    - df_reordered_and_filtered: pandas DataFrame, the reordered and filtered DataFrame
    """
    # Ensure the column values are of the same type as the values in the list
    df[column_name] = df[column_name].astype(type(value_list[0]))

    # Reorder the DataFrame based on the specified column and list of values
    df_reordered = df[df[column_name].isin(value_list)].sort_values(by=column_name, key=lambda x: x.map({val: i for i, val in enumerate(value_list)}))

    return df_reordered

# Example usage:
# Replace 'your_dataframe', 'your_column_name', and 'your_value_list' with the actual DataFrame, column name, and value list you are working with
# reordered_and_filtered_df = reorder_and_filter_dataframe(your_dataframe, 'your_column_name', your_value_list)

def unlock_cells(ws, col_name, min_row=None, max_row=None,fill_color=None,border=None):
    if min_row is None: min_row=1
    if max_row is None: max_row=ws.max_row
    idx=column_index_from_string(col_name)
    for row in ws.iter_rows(min_row, max_row, min_col=idx, max_col=idx):
        for cell in row:
            if fill_color is not None and border is not None: 
                cell.fill=fill_color
                cell.border=border
            cell.protection = Protection(locked=False)

def stripe_cells(ws,fill_color,border):
    count=0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if count%2==0:
            for cell in row:
                if fill_color is not None: 
                    cell.fill=fill_color
                    cell.border=border
        count+=1

def get_letter_from_column_name(df,col_name):
    return df.columns.get_loc(col_name) + 1, get_column_letter(df.columns.get_loc(col_name) + 1)

def get_next_letter_from_column_name(df,col_name):
    return df.columns.get_loc(col_name) + 2, get_column_letter(df.columns.get_loc(col_name) + 2)

def compact_excel_file(input_file, output_file):
    tabs = pd.ExcelFile(input_file).sheet_names 
    #print(tabs)
    with pd.ExcelWriter(output_file) as writer: 
        for sheet_name in tabs:
            print(sheet_name)
            # Read the sheet into a pandas DataFrame
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            df.columns=simplify_strings(df.columns) 
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print('sheet', sheet_name, 'created')

def compact_workbook(source_workbook, output_file):
    new_workbook = openpyxl.Workbook()
    # Iterate through sheets in the source workbook
    for sheet_name in source_workbook.sheetnames:
        print('Sheet:', sheet_name)
        # Copy the sheet
        source_sheet = source_workbook[sheet_name]
        new_sheet = new_workbook.create_sheet(title=sheet_name)

        # Copy each cell and its value
        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1, max_col=source_sheet.max_column):
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value

        # Copy styles
        for column in range(1, source_sheet.max_column + 1):
            for row in range(1, source_sheet.max_row + 1):
                source_cell = source_sheet.cell(row=row, column=column)
                new_cell = new_sheet.cell(row=row, column=column)

                # Copy font, fill, and border attributes
                new_cell.font = copy(source_cell.font)
                new_cell.fill = copy(source_cell.fill)
                new_cell.border = copy(source_cell.border)

                # Copy number format and alignment
                new_cell.number_format = copy(source_cell.number_format)
                new_cell.alignment = copy(source_cell.alignment)

    # Apply filters to the first row of each sheet
    for sheet in new_workbook.sheetnames:
        new_workbook[sheet].auto_filter.ref = new_workbook[sheet].dimensions

    # Save the new workbook
    new_workbook.save(output_file)


def prints_dictionary_of_sheet_and_column_names(workbook, input_file, ws_names=None):
    ''' ws_names is a dict where values are the names of the workbook sheets'''
    if ws_names is None: 
        ws_names={sheet_name: sheet_name for sheet_name in workbook.sheetnames}
    for sheet_name in workbook.sheetnames:
        # Read the sheet into a pandas DataFrame
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        df.columns=simplify_strings(df.columns)
        # determine dictionary key that has value sheet_name 
        key=list(filter(lambda x: ws_names[x] == sheet_name, ws_names))[0]
        print(key,'=', {key+'_'+x[0:6] : x for x in df.columns})

def simplify_strings(s):
    ''' removes accents and replaces spaces by _ ; requires unidecode; s is a list of strings'''
    return list(map(lambda x: x.replace(' ','_'),list(map(unidecode, s))))

def df_to_excel(df, ws, header, index, startrow, startcol):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def df_to_excel_with_columns(df,ws,maxwidth=30,header=True,index=False,startrow=0, startcol=0):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = max(len(column)-2,min(max_length+2, maxwidth)) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_excel(df,ws,header,index, startrow, startcol)

def replace_values_in_string(s, df):
    '''
    This function iterates over each row of the DataFrame df, replacing occurrences of values from the 'old' column with their corresponding values from the 'new' column in the string s, and returns the modified string.
    '''
    modified_s = str(s)
    for index, row in df.iterrows():
        modified_s = modified_s.replace(row[0], row[1])
    return modified_s

# Define a function to generate unique codes for values with length less than 4
def generate_code(value,L):
    if len(value) < L:
        return '00000' + str(hash(value))
    else:
        return value