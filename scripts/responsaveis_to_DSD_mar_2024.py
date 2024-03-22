######################################################
# Manuel Campagnolo (mar 2024)
# ISA/ULIsboa
# script para preparar ficheiro Excel DSD ISA 2024-2025
#######################################################

from pathlib import Path
import pandas as pd
from copy import copy
from unidecode import unidecode
import random
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname # https://openpyxl.readthedocs.io/en/latest/validation.html?highlight=validation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from functions import simplify_strings, df_to_excel_with_columns, compact_excel_file, get_letter_from_column_name,unlock_cells,stripe_cells
from functions import add_suffix_to_duplicates, reorder_and_filter_dataframe, insert_row_at_beginning, insert_row_at_end,sort_list
from functions import replace_values_in_string, generate_code, get_next_letter_from_column_name
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
pd.options.mode.copy_on_write = True
############################################# constants

UNPROTECT_OUTPUT_CELLS=True # para desproteger células com função unlock_cells(ws, col_name, min_row=None, max_row=None):
PROTECT_WORKSHEET=True # tem que ser True para impedir escrita
PASSWORD='kathleen'

# root (working directory)
try:
    working_dir=Path(__file__).parent.parent # working directory from script location: scripts are in 'scripts' folder
except:
    working_dir=Path().absolute()


# folders
FOLDER_SERVICOS= 'ficheiros_servicos_ISA'
FOLDER_OUTPUT='output_files'
FOLDER_FICH_RESPONSAVEIS_UCs='ficheiros_responsaveis_ucs'
FOLDER_2425='DSD_2024_2025'
FOLDER_2324='DSD_2023_2024'
FN_externo='DSD_2023_2024_servico_externo_v6_revisto_TF_DSD_28junho.xlsx'
FN_responsaveis='DSD_2024_2025_responsaveis_UCs_fechado_12_mar_2024_desprotegido_ML_15MAR2024_corr_SIG2ciclo.xlsx'
FN_resumo='resumo_DSD_2324_feito_em_fev_2024.xlsx'

# list files in FOLDER_OUTPUT and create name of output file
folder_output=working_dir/FOLDER_2425/FOLDER_OUTPUT
number_files=len(list(folder_output.rglob("DSD_2024_2025*")))
DSD_OUTPUT_FICH='DSD_2024_2025_v'+str(number_files+1)+'.xlsx' 

# main files
fn_resp= working_dir / FOLDER_2425 / FOLDER_FICH_RESPONSAVEIS_UCs / FN_responsaveis
fn_resumo= working_dir / FOLDER_2324 / FN_resumo
fn_externo= working_dir / FOLDER_2324 / FN_externo
fn_output = folder_output / DSD_OUTPUT_FICH

# worksheets (upper case) and column names (lower case) from fn_resp

# a coluna das horas de 2023_2024 vem de resumo
UC ='uc_2024-25'
UC_uc = 'unidade_curricular'
UC_obrigatoria = 'uc_obrigatoria'
UC_optativa = 'uc_optativa'
UC_codigo='cod_uc'
UC_resp='responsavel_unidade_curricular' #
UC_resp_justif='justicação novo responsável UC'
UC_resp_cc='responsável confirmado CC'
UC_mudanca_resp='Houve mudança'
UC_sugestoes='sugestões de modificação da info da UC'
UC_autor_sugestao='autor da sugestão'
UC_ciclo_curso = 'ciclo_curso'
UC_horas_contacto = 'h_contacto_total'
UC_ano='ano'
UC_sem='sem'
UC_horas_totais='horas totais'
UC_horas_totais_sugeridas='horas totais revistas'
UC_dif_horas='diferença horas'
UC_horas_justif='justificação horas totais'
# as outras colunas vem de UC:
DSD='DSD_2024-25'
DSD_horas_totais='horas totais UC'
DSD_horas_docente='horas docente'
DSD_resp='docente                '
DSD_obs='observação              '
DSD_autor='autor da observação'
#
UCMETA='uc_meta'
RH='RH'
RH_nome='nome' # nomes docentes
RH_numero = 'num_pessoal'
RH_data_fim='data_fim'
RH_posicao='posicao'
RH_obs = 'Obs'
RH_horas='total horas'
RH_horas_ext='total horas externas'
RH_horas_semana='horas por semana'
RH_autor_sugestao='autor da sugestão'
RHPOSICAO='RH_posicao' # drop
RHMETA='RH_meta'
PE='planos_estudos' # drop
AC='AC' #drop
ACMETA='AC_meta'#drop
POS='POS' #drop
UCAREA='uc_AreaCient' # drop
CODCURSO='cod_curso'
CODCURSO_cod_curso='cod_curso'
CODCURSO_sigla='siglaCurso'
EXT='Servico_externo'
EXT_docente='Nome do docente'
EXT_horas='Número horas letivas'

# sheet from fn_resumo
HORASUCS='horas_UCs'
HORASUCS_codigo='Código UC'
HORASUCS_nome_uc='Nome da UC'
HORASUCS_total_horas='Total horas UC'
HORASUCS_total_horas_novo = 'Total horas docência'

# values: Sheet_column_value
RH_nome_pro_bono='docente_PRO_BONO' 
RH_nome_em_contratacao='Docente em contratação'
RH_data_fim_sem_termo='sem termo'
DATA_TERMO_CERTO='2024-09-01'
N_extra_nomes=0 # docentes adicionais a poder criar em RH
N_docentes=18
N_ext_plus=30 # linhas adicionais a poder criar em serviço externo

# desproteger células:
# responsável
# docentes extra + observações

'''
, 'UC_meta': 'uc_meta', 'AC': 'uc_AreaCient', 'AC_meta': 'AC_meta', 'CC': 'cod_curso', 'RH': 'RH', 'RH_meta': 'RH_meta', 'POS': 'RH_posicao'}
UC = {'UC_ciclo_': 'ciclo_curso', 'UC_uc_obr': 'uc_obrigatoria', 'UC_uc_opt': 'uc_optativa', 'UC_cod_uc': 'cod_uc', 'UC_unidad': 'unidade_curricular', 'UC_Respon': 'Responsavel', 'UC_não_fu': 'não_funciona_2024_2025', 'UC_area_c': 'area_cient', 'UC_ano': 'ano', 'UC_sem': 'sem', 'UC_h_trab': 'h_trab_totais', 'UC_T': 'T', 'UC_TP': 'TP', 'UC_PL': 'PL', 'UC_TC': 'TC', 'UC_S': 'S', 'UC_E': 'E', 'UC_OT': 'OT', 'UC_O': 'O', 'UC_h_cont': 'h_contacto_total', 'UC_ECTS': 'ECTS', 'UC_uc_int': 'uc_interna', 'UC_obs': 'obs'}
UC_meta = {'UC_meta_ciclo_': 'ciclo_curso', 'UC_meta_1 cicl': '1 ciclo', 'UC_meta_Licenc': 'Licenciatura'}
AC = {'AC_area_c': 'area_cientifica', 'AC_sigla': 'sigla', 'AC_classi': 'classificacao_area'}
AC_meta = {'AC_meta_classi': 'classificacao_area', 'AC_meta_CNAEF': 'CNAEF', 'AC_meta_Classi': 'Classificação Nacional das Áreas de Educação e Formação (CNAEF), Portaria n.o 256/2005 de 16 de Março. Classificação: nível 1, 2 e 3 (código: 3 dígitos)'}
CC = {'CC_cod_cu': 'cod_curso', 'CC_grau': 'grau', 'CC_curso': 'curso', 'CC_obs': 'obs'}
RH = {'RH_num_pe': 'num_pessoal', 'RH_nome': 'nome', 'RH_corpo': 'corpo', 'RH_posica': 'posicao', 'RH_ETI': 'ETI', 'RH_contra': 'contratacao_vinculo', 'RH_data_f': 'data_fim', 'RH_Obs': 'Obs'}
RH_meta = {'RH_meta_num_pe': 'num_pessoal', 'RH_meta_número': 'número de indentificação pessoal atribuído pelos RH', 'RH_meta_Unname': 'Unnamed: 4'}
POS = {'POS_posica': 'posicao', 'POS_h_min': 'h_min', 'POS_h_max': 'h_max', 'POS_obs': 'obs'}
'''
# colunas relevantes EXT:
col_EXT=[EXT_docente,'Nome da UC', 'Nome do curso', 'Nível',
       'Semestre de funcionamento', 'Parceria institucional estabelecida',
       'Com centro de custos ISA ou de empresas do ISA', 'Ocorrência',
       'Instituição responsável', 'Link do curso', 'Responsável do Curso',
       'Responsável de UC', 'Email responsável UC', 'Funcionamento',
       'Funciona dependente do nº alunos',
       'Número mínimo de alunos para funcionar', 'Número horas letivas',
       'observação']

####################################### dados para validação EXT
SN=['Sim','Não']
mydict={EXT_docente: None,
         'Nome da UC': None,
         'Nome do curso': None,
         'Nível': ['1º ciclo','2º ciclo','3º ciclo','Não conferente de grau'],
         'Semestre de funcionamento': ['1º','2º','extra-semestre'],
         'Parceria institucional estabelecida': SN,
         'Com centro de custos ISA ou de empresas do ISA': SN,
         'Ocorrência': ['Todos os anos', 'Não regular'],
         'Instituição responsável': None,
         'Link do curso': None,
         'Responsável do Curso': None,
         'Responsável de UC': None,
         'Email responsável UC': None,
         'Funcionamento': ['obrigatória','optativa'],
         'Funciona dependente do nº alunos': SN,
         'Número mínimo de alunos para funcionar': None,
         'Número horas letivas':None
}
####################################

# cor light red, light yellow
fill_red = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
fill_red = PatternFill(bgColor="FFC0CB",fill_type="solid")
fill_green = PatternFill(start_color="C0FFCB", end_color="C0FFCB", fill_type="solid")
fill_updated=PatternFill(start_color="FFDBE9",fill_type="solid") # light pink
fill_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # alpha: 1st 2 characters
fill_light_yellow = PatternFill(start_color="FFFFED", end_color="FFFFED", fill_type="solid") 
thin_border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

###################################################################### read inputs to dataframes
# read sheet names
sheets_resp = pd.ExcelFile(fn_resp).sheet_names
sheets_resumo = pd.ExcelFile(fn_resumo).sheet_names
sheets_externo = pd.ExcelFile(fn_externo).sheet_names

# read data from inputs
#RH
df_rh = pd.read_excel(fn_resp, sheet_name=RH)
# RH: acrescentar N_extra_nomes novas linhas e também coluna horas letivas do docente
N_rh=df_rh.shape[0]
for i in range(N_extra_nomes):
    df_rh=insert_row_at_end(df_rh,{RH_nome: 'indicar nome '+str(i+1), RH_posicao:'indicar se possível', RH_obs: 'nomes em falta na lista: NÃO SÃO contratações' })
df_rh[RH_horas]=''
df_rh[RH_horas_ext]=''
df_rh[RH_horas_semana]=''
df_rh_meta = pd.read_excel(fn_resp, sheet_name=RHMETA,header=None)
df_rh_meta.columns=['coluna','descricao','descricao_2','descricao_3','descricao_4']
# CODCURSO
df_cod_curso = pd.read_excel(fn_resp, sheet_name=CODCURSO)
df_cod_curso_aux=df_cod_curso[df_cod_curso[CODCURSO_cod_curso].notna()] # remover cursos sem cod_curso
df_cod_curso[CODCURSO_cod_curso]=df_cod_curso[CODCURSO_cod_curso].astype(str).str.strip()
df_cod_curso_aux[CODCURSO_cod_curso]=df_cod_curso_aux[CODCURSO_cod_curso].astype(str).str.strip()
#RESUMO
df_resumo = pd.read_excel(fn_resumo, sheet_name=HORASUCS)
df_resumo[HORASUCS_codigo]=df_resumo[HORASUCS_codigo].astype(str).str.strip()
df_resumo = df_resumo.rename(columns={HORASUCS_total_horas: HORASUCS_total_horas_novo})
del(HORASUCS_total_horas)
# UC
df_uc = pd.read_excel(fn_resp, sheet_name=UC)
df_uc[UC_codigo]=df_uc[UC_codigo].astype(str).str.strip()
df_uc=df_uc.fillna('')
# adicionar código às UCs sem código: a probabilidade de 2 terem o mesmo código é muito baixa
df_uc.loc[df_uc[UC_codigo].str.len()<4,UC_codigo]=df_uc.loc[df_uc[UC_codigo].str.len()<4,UC_codigo].apply(lambda x: 'cod'+str(random.randint(10000000, 99999999)))
#print(df_uc[df_uc[UC_codigo].str.len()>7][UC_codigo])
# criar coluna que será protegida do resp confirmado pelo CC; UC_resp será desprotegido e validado
df_uc[UC_resp_cc]=df_uc[UC_resp]
df_uc_meta = pd.read_excel(fn_resp, sheet_name=UCMETA,header=None)
df_uc_meta.columns=['coluna','descricao','observacoes']
# trocar abreviaturas em UC_obrigatoria,UC_optativa
df_replace=df_cod_curso_aux[[CODCURSO_cod_curso,CODCURSO_sigla]]
df_uc[UC_obrigatoria] = df_uc[UC_obrigatoria].apply(lambda x: replace_values_in_string(x, df_replace))
df_uc[UC_optativa] = df_uc[UC_optativa].apply(lambda x: replace_values_in_string(x, df_replace))
# merge with df_resumo: para ir buscar o total de horas da UC (2023/2024)
df_uc=pd.merge(df_uc,df_resumo[[HORASUCS_codigo,HORASUCS_total_horas_novo]],left_on=UC_codigo,right_on=HORASUCS_codigo,how='left')
# substituir HORASUCS_total_horas_novo por horas contacto quando nan
df_uc[UC_horas_totais] = df_uc[HORASUCS_total_horas_novo].fillna(df_uc[UC_horas_contacto])
# adicionar colunas para horas revistas e justificação da mudança
df_uc[UC_horas_totais_sugeridas] = df_uc[UC_horas_totais]
df_uc[UC_dif_horas]=0
df_uc[UC_horas_justif]=''
df_uc[UC_resp_justif]=''
df_uc[UC_mudanca_resp]=''
# re-ordenar colunas UC
df_uc=df_uc[[UC_ciclo_curso, UC_obrigatoria,UC_optativa, UC_codigo, 'uc_interna', UC_ano,UC_sem,'ECTS',UC_uc,
             'h_trab_totais', 'T', 'TP','PL', 'TC', 'S', 'E', 'OT', 'O',
             UC_horas_contacto, UC_horas_totais, UC_horas_totais_sugeridas,UC_dif_horas,UC_horas_justif,'obs',  
             UC_autor_sugestao, UC_sugestoes,UC_resp_cc,UC_resp,UC_mudanca_resp, UC_resp_justif]]

# EXTERNO
df_ext = pd.read_excel(fn_externo, sheet_name=EXT)
df_ext=df_ext[col_EXT]
L=list(df_rh[RH_nome])
df_ext=reorder_and_filter_dataframe(df_ext, EXT_docente, L)
N_ext=df_ext.shape[0]
# renomear
df_ext = df_ext.rename(columns={'Número horas letivas': 'Número horas letivas 2022_2023'})
# trocar colunas
col_EXT=list(set(df_ext.columns)-set(['observação','Número horas letivas 2022_2023']))+['observação','Número horas letivas 2022_2023']
df_ext=df_ext[col_EXT]
df_ext[EXT_horas]=0
col_EXT=df_ext.columns

# main dataframe for DSD
# select columns to be shown: os nomes das colunas de DSD e de UC são iguais nesta altura
df_dsd=df_uc[[UC_ciclo_curso,UC_codigo,UC_ano, UC_sem,UC_uc,UC_resp]]
df_dsd = df_dsd.rename(columns={UC_resp: DSD_resp})
#df_dsd[DSD_horas_totais]=''
df_dsd[DSD_horas_docente]=0
df_dsd[DSD_autor]=''
df_dsd[DSD_obs]=''
df_dsd=df_dsd.fillna('')
# duplicar linhas N_docentes vezes
IDXREP='idx_rep'
N_ucs=df_dsd.shape[0]
df_dsd=df_dsd.reindex(df_dsd.index.repeat(N_docentes)).reset_index(drop=True)
# create indices: 0 corresponde à linha principal do UC; 1 é o docente que é o responsável; 2 a N-1 é para os restantes docentes
series=[]
for _ in range(N_ucs): series.extend(range(N_docentes))
df_dsd[IDXREP]=series
# limpar células para idx_rep>0, para que sejam indicados os docentes na coluna DSD_resp
for j in range(df_dsd.shape[0]):
    if df_dsd.iloc[j, df_dsd.columns.get_loc(IDXREP)]==0:
        df_dsd.iloc[j,[df_dsd.columns.get_loc(c) for c in [DSD_resp]]]='Horas totais da UC:'
    if df_dsd.iloc[j, df_dsd.columns.get_loc(IDXREP)]>1:
        df_dsd.iloc[j,[df_dsd.columns.get_loc(c) for c in [DSD_resp]]]=''
    if df_dsd.iloc[j, df_dsd.columns.get_loc(IDXREP)]==N_docentes-2:
        df_dsd.iloc[j,[df_dsd.columns.get_loc(c) for c in [DSD_resp]]]='Em falta (positivo); A mais (negativo):'
    if df_dsd.iloc[j, df_dsd.columns.get_loc(IDXREP)]==N_docentes-1:
        df_dsd.iloc[j,[df_dsd.columns.get_loc(c) for c in df_dsd.columns]]=''
df_dsd.drop(columns=[IDXREP],inplace=True)

# select columns for printing
# dfaux=df[[UC_ciclo_curso,UC_codigo,UC_uc, UC_horas_contacto,HORASUCS_codigo,HORASUCS_total_horas_novo, UC_horas_totais_sugeridas, UC_horas_justif]]
# print(dfaux[dfaux[HORASUCS_total_horas].notna()])

# Criar novos registos em _meta
df_uc_meta=insert_row_at_end(df_uc_meta, {'coluna':HORASUCS_total_horas_novo, 'descricao':'horas totais de docência da UC no DSD 2023-2024; caso não exista, é o número d ehoras de contacto da UC'})
df_uc_meta=insert_row_at_end(df_uc_meta, {'coluna':UC_resp, 'descricao':'Novo responsável'})
df_uc_meta=insert_row_at_end(df_uc_meta, {'coluna':UC_resp_justif, 'descricao':'Justificação de alteração de responsável'})
df_uc_meta=insert_row_at_end(df_uc_meta, {'coluna':UC_horas_totais_sugeridas, 'descricao':'Horas totais sugeridas para a UC'})
df_uc_meta=insert_row_at_end(df_uc_meta, {'coluna':UC_horas_justif, 'descricao':'Justificação para as horas totais sugeridas para a UC'})
df_rh_meta=insert_row_at_end(df_rh_meta, {'coluna':RH_horas, 'descricao':'horas totais de docência no ano'})
df_rh_meta=insert_row_at_end(df_rh_meta, {'coluna':RH_horas_semana, 'descricao':'horas médias docência por semana'})

############################################################################## converter dataframes em sheets
new_workbook = Workbook()
sheet_names=[DSD,UC,UCMETA,RH,RHMETA,EXT,CODCURSO]
for sheet_name in sheet_names:
    # for each sheet_name, 1) create worksheet; 2) create validation; 3)  df_to_excel_with_columns; 4) stripe_cells; 5) unlock_cells
    print(sheet_name) 
    # visit all sheets
    if sheet_name==DSD: df=df_dsd.copy()
    if sheet_name==UC: df=df_uc.copy()
    if sheet_name==UCMETA: df=df_uc_meta.copy()
    if sheet_name==RH: df=df_rh.copy()
    if sheet_name==RHMETA: df=df_rh_meta.copy()
    if sheet_name==EXT: df=df_ext.copy()
    if sheet_name==CODCURSO: df=df_cod_curso.copy()

    # Create a new sheet in the new workbook
    new_worksheet = new_workbook.create_sheet(title=sheet_name)
    
    if sheet_name==DSD or sheet_name==UC or sheet_name==EXT: 
        # add datavalidation to worksheet
        idx,letter=get_letter_from_column_name(df_rh,RH_nome)
        # dv_resp inclui o 1o valor de 'Coordenação externa ao ISA'
        dv_resp = DataValidation(type="list", formula1=f"{quote_sheetname(RH)}!${letter}$2:${letter}${N_rh+1+N_extra_nomes}") # 10 extra
        dv_docente = DataValidation(type="list", formula1=f"{quote_sheetname(RH)}!${letter}$3:${letter}${N_rh+1+N_extra_nomes}") # 10 extra
        dv_docente_nao_pro_bono = DataValidation(type="list", formula1=f"{quote_sheetname(RH)}!${letter}$4:${letter}${N_rh+1+N_extra_nomes}") # 10 extra
        new_worksheet.add_data_validation(dv_resp)
        new_worksheet.add_data_validation(dv_docente)
        new_worksheet.add_data_validation(dv_docente_nao_pro_bono)

    # apply validation: create drop-down menu
    # validação docente e autor observação em DSD
    if sheet_name==DSD: 
        idx_docente,letter_docente=get_letter_from_column_name(df_dsd,DSD_resp) # é onde se coloca o nome do docente
        idx_autor,letter_autor=get_letter_from_column_name(df_dsd,DSD_autor)
        for i in range(N_ucs):
            dv_docente.add(f"${letter_docente}${i*N_docentes+4}:{letter_docente}${(i+1)*N_docentes-1}") # creates drop-down menu.
            dv_docente_nao_pro_bono.add(f"${letter_autor}${i*N_docentes+2}:{letter_autor}${(i+1)*N_docentes-1}") # creates drop-down menu.
    # validation autor_sugestao
    if sheet_name==UC: 
        idx,letter=get_letter_from_column_name(df_uc,UC_resp)
        dv_resp.add(f"${letter}$2:{letter}${N_ucs+1}") # creates drop-down menu.
    
    # validação EXT
    if sheet_name==EXT: 
        idx,letter=get_letter_from_column_name(df_ext,EXT_docente)
        dv_docente_nao_pro_bono.add(f"${letter}$2:{letter}${N_ext+N_ext_plus+1}") # creates drop-down menu.
        # outras colunas com validação (listas)
        for key, value in mydict.items(): # key é o nome da coluna, value é a lista para validação
            if value is not None:
                #print(key,value)
                idx,letter=get_letter_from_column_name(df_ext,key)
                dv = DataValidation(type="list", formula1='"' + ','.join(value) + '"')
                new_worksheet.add_data_validation(dv)
                dv.add(f"${letter}$2:{letter}${N_ext+N_ext_plus+1}") # creates drop-down menu.

    # Write the DataFrame to the new workbook: determines max width of columns
    if sheet_name==EXT:
        df_to_excel_with_columns(df,new_worksheet,maxwidth=15,header=True,index=False,startrow=0, startcol=0)
    else:
        df_to_excel_with_columns(df,new_worksheet,maxwidth=30,header=True,index=False,startrow=0, startcol=0)

    # Apply filters to the first row
    if '_meta' not in sheet_name:
        new_worksheet.auto_filter.ref = new_worksheet.dimensions
        if sheet_name==UC:
            idx,letter=get_next_letter_from_column_name(df_uc,UC_uc)
            new_worksheet.freeze_panes = f"{letter}{2}"
        else:
            new_worksheet.freeze_panes = "A2"
    # stripes
    stripe_cells(new_worksheet, fill_color=fill_light_yellow,border=thin_border)    
    
    ############################################################################# específico de cada "sheet"
    # a ideia é desbloquear algumas células e a seguir bloquear toda a worksheet
    if sheet_name==DSD and UNPROTECT_OUTPUT_CELLS:
        idx,letter=get_letter_from_column_name(df,DSD_horas_docente)
        for i in range(N_ucs):
            unlock_cells(new_worksheet,letter,min_row=i*N_docentes+3,max_row=(i+1)*N_docentes-1,fill_color=fill_green,border=thin_border)
        # desbloquear os nomes dos docentes, observaçoes e autores
        idx_docente,letter_docente=get_letter_from_column_name(df,DSD_resp)
        idx_autor,letter_autor=get_letter_from_column_name(df,DSD_autor)
        idx_obs,letter_obs=get_letter_from_column_name(df,DSD_obs)
        for i in range(N_ucs):
            unlock_cells(new_worksheet,letter_docente,min_row=i*N_docentes+4,max_row=(i+1)*N_docentes-1,fill_color=fill_green,border=thin_border)
            unlock_cells(new_worksheet,letter_autor,min_row=i*N_docentes+2,max_row=(i+1)*N_docentes-1,fill_color=fill_green,border=thin_border)
            unlock_cells(new_worksheet,letter_obs,min_row=i*N_docentes+2,max_row=(i+1)*N_docentes-1,fill_color=fill_green,border=thin_border)
        # células para número de horas docência (VLOOKUP)
        idx_docente,letter_docente=get_letter_from_column_name(df_dsd,DSD_resp)
        idx_horas,letter_horas=get_letter_from_column_name(df_dsd,DSD_horas_docente)
        idx_codigo,letter_codigo=get_letter_from_column_name(df_dsd,UC_codigo)
        idx_codigo_UC,letter_codigo_UC=get_letter_from_column_name(df_uc,UC_codigo)
        idx_horas_UC,letter_horas_UC=get_letter_from_column_name(df_uc,UC_horas_totais_sugeridas)
        for i in range(N_ucs):
            # célula onde vai estar o número total de horas da UC
            c=new_worksheet.cell(column=idx_horas,row=i*N_docentes+2)  #.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            if idx_codigo_UC < idx_horas_UC:
                ncols = idx_horas_UC - idx_codigo_UC + 1 
                c.value=f"=VLOOKUP({letter_codigo}{i*N_docentes+2},'{UC}'!{letter_codigo_UC}:{letter_horas_UC}, {ncols}, FALSE)" 
                c.fill=fill_updated
            else:
                ncols =  idx_codigo_UC - idx_horas_UC + 1 
                c.value=f"=VLOOKUP({letter_codigo}{i*N_docentes+2},'{UC}'!{letter_horas_UC}:{letter_codigo_UC}, {ncols}, FALSE)" 
                c.fill=fill_updated
            # célula onde está o número de horas em falta
            c=new_worksheet.cell(column=idx_horas,row=(i+1)*N_docentes)
            c.value=f"={letter_horas}{i*N_docentes+2}-SUMIF({letter_docente}{i*N_docentes+3}:{letter_docente}{(i+1)*N_docentes-1},\"<>\",{letter_horas}{i*N_docentes+3}:{letter_horas}{(i+1)*N_docentes-1})"
            c.fill=fill_updated
        # com VLOOKUP ir buscar o nome do responsável à folha UC, coluna UC_resp e colocar na 2a célula do docente, em DSD_resp
        idx_docente,letter_docente=get_letter_from_column_name(df_dsd,DSD_resp)
        idx_codigo,letter_codigo=get_letter_from_column_name(df_dsd,UC_codigo)
        idx_codigo_UC,letter_codigo_UC=get_letter_from_column_name(df_uc,UC_codigo)
        idx_resp_UC,letter_resp_UC=get_letter_from_column_name(df_uc,UC_resp)
        for i in range(N_ucs):
            # célula onde vai estar o responsável da UC
            c=new_worksheet.cell(column=idx_docente,row=i*N_docentes+3)  
            if idx_codigo_UC < idx_resp_UC:
                ncols = idx_resp_UC - idx_codigo_UC + 1 
                c.value=f"=VLOOKUP({letter_codigo}{i*N_docentes+2},'{UC}'!{letter_codigo_UC}:{letter_resp_UC}, {ncols}, FALSE)" 
                c.fill=fill_updated
            else:
                ncols =  idx_codigo_UC - idx_resp_UC + 1 
                c.value=f"=VLOOKUP({letter_codigo}{i*N_docentes+2},'{UC}'!{letter_resp_UC}:{letter_codigo_UC}, {ncols}, FALSE)" 
                c.fill=fill_updated

    if sheet_name==UC and UNPROTECT_OUTPUT_CELLS:
        idx,letter=get_letter_from_column_name(df,UC_horas_totais_sugeridas)
        unlock_cells(new_worksheet,letter,fill_color=fill_green,border=thin_border)
        idx,letter=get_letter_from_column_name(df,UC_horas_justif)
        unlock_cells(new_worksheet,letter,fill_color=fill_green,border=thin_border)
        idx,letter=get_letter_from_column_name(df,UC_resp)
        unlock_cells(new_worksheet,letter,fill_color=fill_green,border=thin_border)
        idx,letter=get_letter_from_column_name(df,UC_resp_justif)
        unlock_cells(new_worksheet,letter,fill_color=fill_green,border=thin_border)
        # calcular diferenças entre horas e horas sugeridas
        idx_horas_UC,letter_horas_UC=get_letter_from_column_name(df_uc,UC_horas_totais)
        idx_horas_UC_sug,letter_horas_UC_sug=get_letter_from_column_name(df_uc,UC_horas_totais_sugeridas)
        idx_horas_UC_dif,letter_horas_UC_dif=get_letter_from_column_name(df_uc,UC_dif_horas)
        for i in range(N_ucs):
            # célula onde vai estar a diferença
            c=new_worksheet.cell(column=idx_horas_UC_dif,row=i+2)
            c.value=f"={letter_horas_UC_sug}{i+2}-{letter_horas_UC}{i+2}"  
            c.fill=fill_updated
        # identificar alteração de responsável
        idx_resp_cc_UC,letter_resp_cc_UC=get_letter_from_column_name(df_uc,UC_resp_cc)
        idx_resp_UC,letter_resp_UC=get_letter_from_column_name(df_uc,UC_resp)
        idx_mudanca_UC,letter_mudanca_UC=get_letter_from_column_name(df_uc,UC_mudanca_resp)
        for i in range(N_ucs):
            # célula onde vai estar a indicação da mudança
            c=new_worksheet.cell(column=idx_mudanca_UC,row=i+2)
            c.value=f"=IF({letter_resp_cc_UC}{i+2}<>{letter_resp_UC}{i+2},\"alterado\",\"---\")"  
            c.fill=fill_updated
    
    # EXT: criar linhas para mais serviço externo
    if sheet_name==EXT and UNPROTECT_OUTPUT_CELLS:
        stripe_cells(new_worksheet, fill_color=fill_light_yellow,border=thin_border)
        # desproteger linhas para novos docentes
        for col in col_EXT:
            idx,letter=get_letter_from_column_name(df_ext,col)
            unlock_cells(new_worksheet,letter, min_row=+2, max_row=N_ext+N_ext_plus+1, fill_color=fill_green,border=thin_border)
        # change row height
        #for k in range(N_ext+N_ext_plus+1):
        #    new_worksheet.row_dimensions[k+1].height = 25 # mudar altura da linha
    
    # Dar possibilidade de criar novos docentes
    if sheet_name==RH and UNPROTECT_OUTPUT_CELLS:
        stripe_cells(new_worksheet, fill_color=fill_light_yellow,border=thin_border)
        # docentes extra (a poderem ser adicionados)
        idx,letter=get_letter_from_column_name(df,RH_nome)
        unlock_cells(new_worksheet,letter, min_row=N_rh+2, max_row=N_rh+N_extra_nomes+1, fill_color=fill_green,border=thin_border)
        idx,letter=get_letter_from_column_name(df,RH_posicao)
        unlock_cells(new_worksheet,letter, min_row=N_rh+2, max_row=N_rh+N_extra_nomes+1, fill_color=fill_green,border=thin_border)
        idx,letter=get_letter_from_column_name(df,RH_obs)
        unlock_cells(new_worksheet,letter, min_row=N_rh+2, max_row=N_rh+N_extra_nomes+1, fill_color=fill_green,border=thin_border)
        # criar validação
        idx_nome,letter_nome=get_letter_from_column_name(df_rh,RH_nome)
        idx_horas,letter_horas=get_letter_from_column_name(df_rh,RH_horas)
        idx_horas_ext,letter_horas_ext=get_letter_from_column_name(df_rh,RH_horas_ext)
        idx_horas_semana,letter_horas_semana=get_letter_from_column_name(df_rh,RH_horas_semana)
        idx_docente,letter_docente=get_letter_from_column_name(df_dsd,DSD_resp)
        idx_docente_EXT,letter_docente_EXT=get_letter_from_column_name(df_ext,EXT_docente)
        idx_horas_DSD,letter_horas_DSD=get_letter_from_column_name(df_dsd,DSD_horas_docente)
        idx_horas_EXT,letter_horas_EXT=get_letter_from_column_name(df_ext,EXT_horas)
        for i in range(N_rh+N_extra_nomes+1):
            c=new_worksheet.cell(column=idx_horas,row=i+2)
            c.value=f"=SUMIF('{DSD}'!{letter_docente}:{letter_docente},'{RH}'!{letter_nome}{i+2},'{DSD}'!{letter_horas_DSD}:{letter_horas_DSD})"
            c.fill=fill_updated
            c=new_worksheet.cell(column=idx_horas_ext,row=i+2)
            c.value=f"=SUMIF('{EXT}'!{letter_docente_EXT}:{letter_docente_EXT},'{RH}'!{letter_nome}{i+2},'{EXT}'!{letter_horas_EXT}:{letter_horas_EXT})"
            c.fill=fill_updated
            c=new_worksheet.cell(column=idx_horas_semana,row=i+2)
            c.value=f"=ROUND(({letter_horas}{i+2}+{letter_horas_ext}{i+2})/28,2)"
            c.fill=fill_updated

    ############################################## protect worksheet
    if PROTECT_WORKSHEET:
        new_worksheet.protection.sheet = True
        new_worksheet.protection.enable()
        new_worksheet.protection = SheetProtection(sheet=True, objects=False, scenarios=False, formatCells=False, formatRows=False, formatColumns=False, insertColumns=True, insertRows=True, insertHyperlinks=True, deleteColumns=True, deleteRows=True, selectLockedCells=False, selectUnlockedCells=False, sort=False, autoFilter=False, pivotTables=True, password=None, algorithmName=None, saltValue=None, spinCount=None, hashValue=None) #(autoFilter=True, formatColumns=True)

# eliminar a worksheet 'Sheet' que foi criada automaticamente
if 'Sheet' in new_workbook.sheetnames:  # remove default sheet
    new_workbook.remove(new_workbook['Sheet'])

# workbook protection
new_workbook.security.workbookPassword = PASSWORD
new_workbook.security.lockStructure = True

# Save the new workbook
try: 
    new_workbook.save(fn_output)
except PermissionError:
    print('File must be in use: close if first please')
new_workbook.close