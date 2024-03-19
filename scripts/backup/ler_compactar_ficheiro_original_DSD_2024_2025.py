# ler ficheiro Excel e criar ficheiro mais simples, com prefixo '_compact.xlsx'
# possivelmente o script poderia ser simplificado usando pandas
# ver função compact_excel_file(input_file, output_file) que usa pandas

from copy import copy
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Border, Side
from pathlib import Path

"""
Ler input

"""

DSD_INPUT_FICH='DSD_inform_2024_2025_v5.xlsx'
PROTECT_OUTPUT_CELLS=True

stem=Path(DSD_INPUT_FICH).stem
suffix=Path(DSD_INPUT_FICH).suffix
COMPACT='_compact'
BLOQ='_bloq'

# Load the source workbook
#input_folder=Path(r'C:\Users\mlc\OneDrive - Universidade de Lisboa\Documents\profissional-isa-cv\cg-isa\DSD_2024_2025\backup_inputs_DSD')
working_dir=Path(__file__).parent.parent # working directory from script location
input_folder= working_dir / 'DSD_2024_2025' / 'input_files'
output_folder= working_dir / 'DSD_2024_2025' / 'output_files'
# Try to read smaller file; otherwise read original file and create smaller file
try:
    source_workbook = load_workbook(output_folder  / (stem+COMPACT+suffix))
except:
    source_workbook = load_workbook(input_folder  / DSD_INPUT_FICH)  # demora
    source_workbook.save(output_folder  / (stem+COMPACT+suffix))
output_file=output_folder / (stem+BLOQ+suffix)

# Create a new workbook
new_workbook = openpyxl.Workbook()

#fill_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # alpha: 1st 2 characters

def UseOpenpyxl(file_name):
    wb = openpyxl.load_workbook(file_name, read_only=True)
    sheet = wb.active
    rows = sheet.rows
    first_row = [cell.value for cell in next(rows)]
    data = []
    for row in rows:
        record = {}
        for key, cell in zip(first_row, row):
            if cell.data_type == 's':
                record[key] = cell.value.strip()
            else:
                record[key] = cell.value
        data.append(record)
    return data


# Iterate through sheets in the source workbook
for sheet_name in source_workbook.sheetnames:
    print('Sheet:', sheet_name)
    # Copy the sheet
    source_sheet = source_workbook[sheet_name]
    new_sheet = new_workbook.create_sheet(title=sheet_name)

    # Copy each cell and its value
    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1, max_col=source_sheet.max_column):
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value

            # Apply protection to all cells except the first row
            if cell.row > 1 and PROTECT_OUTPUT_CELLS:
                new_cell.protection = Protection(locked=True)

    # Copy styles
    for column in range(1, source_sheet.max_column + 1):
        for row in range(1, source_sheet.max_row + 1):
            source_cell = source_sheet.cell(row=row, column=column)
            new_cell = new_sheet.cell(row=row, column=column)

            # Copy font, fill, and border attributes
            new_cell.font = copy(source_cell.font)
            new_cell.fill = copy(source_cell.fill)
            new_cell.border = copy(source_cell.border)

            # Copy number format and alignment
            new_cell.number_format = copy(source_cell.number_format)
            new_cell.alignment = copy(source_cell.alignment)

# Apply filters to the first row of each sheet
for sheet in new_workbook.sheetnames:
    new_workbook[sheet].auto_filter.ref = new_workbook[sheet].dimensions

# Save the new workbook
new_workbook.save(output_file)
