######################################################
# Manuel Campagnolo (abril 2023)
# ISA/ULIsboa
# script para preparar ficheiro Excel DSD ISA 2023-2024
#######################################################

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.styles import Border, Side

from copy import copy
import os
from fuzzywuzzy import fuzz # compare strings
import numpy as np

import pandas as pd
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

#fuzz.ratio('abscerg', 'abcderfg')

# folder not necessary when working in a folder
#folder=r"C:\Users\mlc\OneDrive - Universidade de Lisboa\Documents\profissional-isa-cv\cg-isa"
#pathIn=os.path.join(folder,"DSD_2324_.xlsx")
#pathOut=os.path.join(folder,"DSD_2324_new.xlsx")
#fnIn="DSD_2324_.xlsx"
fnDSDv1="DSD_2324_28abril.xlsx" #  "DSD_v1_teste.xlsx"; 
DSDv1=True # para copiar a informação já inserida 'preencher'+ folha 'info'
#fnIn="DSD_2324_ML_12abr2023__.xlsx"
fnIn="DSD_2324_ML_12abr2023_acrescentar_UC_no_final__.xlsx" # agora tem folha 'DocentesNovo'
#fnOut="DSD_2324_ML_12abr2023_new.xlsx"
fnOut="DSD_2324.xlsx"
fnResumo="resumo_DSD_2324.xlsx"

RESUMO=True

########################################################################### functions (não usado)

def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

def create_workbook_from_dataframe(df):
    """
    1. Create workbook from specified pandas.DataFrame
    2. Adjust columns width to fit the text inside
    3. Make the index column and the header row bold
    4. Fill background color for the header row

    Other beautification MUST be done by usage side.
    """
    workbook = openpyxl.Workbook()
    ws = workbook.active

    rows = dataframe_to_rows(df.reset_index(), index=False)
    col_widths = [0] * (len(df.columns) + 1)
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):

            if type(val) is str:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            elif hasattr(val, "sort"):
                cell = ws.cell(row=i, column=j, value=", ".join(list(map(lambda v: str(v), list(val)))))
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val))])
            else:
                cell = ws.cell(row=i, column=j, value=val)
                col_widths[j - 1] = max([col_widths[j - 1], len(str(val)) + 1])

            # Make the index column and the header row bold
            if i == 1 or j == 1:
                cell.font = Font(bold=True)

    # Adjust column width
    for i, w in enumerate(col_widths):
        letter = get_column_letter(i + 1)
        ws.column_dimensions[letter].width = w

    return workbook

def set_border(ws):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows()
    for row in rows:
        for cell in row:
            cell.border = border

############################################################ funcoes
# devolve letra da coluna com nome (1a linha) da worksheet ws
def nomeColuna2letter(ws,nome):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx=headers.index(nome)
    return get_column_letter(idx+1)

# copy column named nameCol to a a column located at colLetter
def copyColumn(ws,nameCol,colLetter,colType=None):
    count=0
    for cell in ws['{}:{}'.format(nomeColuna2letter(ws,nameCol),nomeColuna2letter(ws,nameCol))]:
        count+=1
        if colType=='numeric': # tenta converter para numérico
            try: 
                new_cell=ws.cell(row=cell.row, column=column_index_from_string(colLetter))
                new_cell.value=int(cell.value)
                new_cell.number_format='0'
            except:
                ws.cell(row=cell.row, column=column_index_from_string(colLetter), value=cell.value)
        else:
            ws.cell(row=cell.row, column=column_index_from_string(colLetter), value=cell.value)

def copy_sheet(source_sheet, target_sheet, idx):
    rows_resp=copy_cells(source_sheet, target_sheet, idx)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)
    return rows_resp

def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = False #copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

# idx indica a coluna em que estão os responsáveis das UCs
def copy_cells(source_sheet, target_sheet,idx):
    delta=0
    rows_resp=[]
    for r, row in enumerate(source_sheet.iter_rows()):
        #print('main r',r)
        # linhas em que há docente indicado
        if source_sheet.cell(r+1,idx+1).value: 
            #print(source_sheet.cell(r+1,idx+1).value)
            # ou 1a linha ou linhas dos responsáveis:
            if r==0: # 1a linha do ficheiro
                copy_first_row(r+delta,row, source_sheet, target_sheet)
            else:
                # linhas dos responsáveis
                copy_row(r+delta,row, source_sheet, target_sheet,fill=True)
                rows_resp.append(r+delta)
                for k in range(N):
                    delta+=1
                    partial_copy_row(r+delta,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao) # row comes from source; r+delta is the index in target
        else:
            # linhas em branco
            copy_row(r+delta,row, source_sheet, target_sheet, fill=False)
        #if r>40: break #<----------------------------------------------------------------------  remover
    print('delta: ',delta)
    return rows_resp

def copy_first_row(r,row, source_sheet, target_sheet):
     for c, cell in enumerate(row):
        source_cell = cell
        if c <= headers.index(colunas_red[-1]): # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                #target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def copy_row(r,row, source_sheet, target_sheet, fill):
    for c, cell in enumerate(row):
        if c==headers.index(colunas_red[-1]):
            # define default style
            red_cell=cell
    for c, cell in enumerate(row):
        source_cell = cell
        if c <= headers.index(colunas_red[-1]): # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            if fill:
                target_cell._value = source_cell._value
                target_cell.data_type = source_cell.data_type
                target_cell.font = copy(red_cell.font)
                target_cell.border = copy(red_cell.border)
                target_cell.fill = copy(red_cell.fill)
                target_cell.number_format = copy(red_cell.number_format)
                #target_cell.protection = copy(red_cell.protection)
                target_cell.alignment = copy(red_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def partial_copy_row(r,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao):
    for c, cell in enumerate(row):
        if c==headers.index(colunas_red[-1]):
            # define default style
            red_cell=cell
    for c, cell in enumerate(row):
        source_cell = cell
        if headers[c] in colunas_red: # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell.font = copy(red_cell.font)
            target_cell.border = copy(red_cell.border)
            target_cell.fill = copy(red_cell.fill)
            target_cell.number_format = copy(red_cell.number_format)
            #target_cell.protection = copy(red_cell.protection)
            target_cell.alignment = copy(red_cell.alignment)
        if headers[c] in columns_to_copy:  
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                #target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
        if headers[c]==coluna_validacao: # drop-down menus
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = VALIDATION_VALUE
            target_cell.data_type='s'
            #target_cell.protection = Protection(locked=False)


########################################################################################################
# number max docentes a inserir por UC e informação a passar
# input
# fnDSDv1="DSD_2324_old.xlsx" # copiar 'DSD (para preencher)'
# fnIn="DSD_2324_ML_12abr2023_acrescentar_UC_no_final__.xlsx" # Copiar 'DSD (informação UCs)' +  'DocentesNovo'
# output
N=15 # número máximo docentes
VALIDATION_VALUE='Inserir docente'
# worksheets de input
ws_name_preencher='DSD (para preencher)'
ws_name_info='DSD (informação UCs)'
if DSDv1: 
    ws_name_docentes='DocentesNovo' 
else: 
    ws_name_docentes='docentes'

# ws preencher
columns_to_copy='Nome da UC' # 'Nome' # designação UC
column_key='Inserir docentes na UC ' #'docentes ' # cuidado: tem um espaço a mais
coluna_validacao=column_key
coldoc='AK'
print(column_index_from_string(coldoc))
colunas_red=['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento',  'Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC', 'ano curricular', 'semestre', 'ECTS', 'semanas', 'Total horas da UC', 'Somatório', 'Horas em falta na UC']
column_codigo_UC_preencher='Código UC'
column_horas_em_falta_preencher='Horas em falta na UC'
column_somatorio_preencher= 'Somatório'
column_preencher_first='Total Horas Teóricas'
column_preencher_last='Total Horas  Outras'
column_new_horas_totais='Horas totais docente'
column_new_horas_semanais='Horas semanais docente'

info_cols_to_unprotect=['H','M','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AD','AE']
DAA='Docente a atribuir'

# ws info (colocada para facilitar VLOOKUP)
column_total_horas_info='Total Horas previsto '
column_codigo_UC_info='Código UC'

# match strings
MIN_MATCH=80

# para resumo
column_horas_em_falta_preencher='Horas em falta na UC' #Horas em falta na UC
column_somatorio_preencher= 'Somatório'
colunas_nome_UC=['Nome da UC', column_codigo_UC_info, 'ciclo de estudos', 'ano curricular', 'semestre']
colunas_FOS=['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento']
colResponsavel='Responsável'
colNomeDocente='NomeDocente'
colLinhaResp='linhaResp'

# colunas a apagar em DSD info: funciona
col_apagar_info_1='Total Horas Somadas '
col_apagar_info_2='Horas em falta na UC'

#################################################################################################
# workbooks
########################################################### criar workbook (target)
# target work book and sheet
wb_target = openpyxl.Workbook()
target_sheet = wb_target.create_sheet(ws_name_preencher)
########################################################### current DSD
if DSDv1: 
    if RESUMO:
        wb1 = openpyxl.load_workbook(fnDSDv1,data_only=True) # with data_only, it will only read values
    else:
        wb1 = openpyxl.load_workbook(fnDSDv1) #,data_only=True) # with data_only, it will only read values
# worksheet input # wb_source is "DSD_2324_ML_12abr2023_acrescentar_UC_no_final__.xlsx" # Copiar 'DSD (informação UCs)' +  'DocentesNovo'
if RESUMO:
    wb_source = openpyxl.load_workbook(fnIn, data_only=True) # with data_only, it will only read values
else:
    wb_source = openpyxl.load_workbook(fnIn) #, data_only=True) # with data_only, it will only read values
wsnames=wb_source.sheetnames
print('ficheiro Madalena: ', wsnames)
# conteudo worksheet docentes
ws=wb_source[ws_name_preencher]
print([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])
ws=wb_source[ws_name_docentes]
print([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])
########################################################### workbook resumo_DSD_2324
wbr=openpyxl.Workbook()
wsr_ucs=wbr.create_sheet('UCs')
wsr_docentes=wbr.create_sheet('docentes')

###########################################################

# test if sheet names are right
# print(wsnames)
# for name in [ws_name_preencher,ws_name_info,ws_name_docentes]:
#     if name not in wsnames:
#         stop
#     source_sheet=wb_source[name]
#     print(source_sheet.max_row)
#     print(source_sheet.max_column)



################################################################################# criar main worksheet 'DSD (para preencher)'
source_sheet=wb_source[ws_name_preencher]

# coluna column_key -> idx
headers = [c.value for c in next(source_sheet.iter_rows(min_row=1, max_row=1))]
idx=headers.index(column_key)
idxred=headers.index(colunas_red[-1])
red_cell=source_sheet.cell(1,idxred)
colunas_horas_a_preencher=[]
acc=[]
for h in headers:
    acc.append(h)
    if column_preencher_first in acc and column_preencher_last not in acc:
        colunas_horas_a_preencher.append(h)
colunas_horas_a_preencher.append(column_preencher_last)

print(colunas_horas_a_preencher)


# Criar cópia
rows_resp=copy_sheet(source_sheet, target_sheet,idx)

# create drop-down validation list
# lista de docentes para validação (drop-down menu)
# lista de 'Posições': coluna B
val_sheet=wb_source[ws_name_docentes]
mylist = [c.value for c in val_sheet['A']]
mylistB = [c.value for c in val_sheet['B']]
# to remove None values in list
mylist = [i for i in mylist if i is not None]
mylistB = [i for i in mylistB if i is not None]
nomes_docentes=list(map(lambda x:x.strip(),mylist)) # eliminar \n
posicoes_docentes=list(map(lambda x:x.strip(),mylistB)) # eliminar \n
if len(mylist)!=len(mylistB): stop

nomes_docentes.append(DAA)
posicoes_docentes.append(' ')

max([len(x) for x in posicoes_docentes])

#docentes a remover
#nomes_docentes.remove('Adelino Mendes da Silva Paiva')
#nomes_docentes.remove('João António Ribeiro Ferreira Nunes')
len(nomes_docentes)


# escrever docentes em coluna coldoc e outras 2 colunas adicionais
for i in range(len(nomes_docentes)):
    # horas semanais
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc)-2)
    if i==0: 
        target_cell.value=column_new_horas_semanais
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else:
        target_cell.value=''
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    # horas totais
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc)-1)
    if i==0: 
        target_cell.value=column_new_horas_totais
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else:
        target_cell.value=''
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    # docentes
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc))
    if i==0: 
        target_cell.value=nomes_docentes[i]
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else: 
        target_cell.value=nomes_docentes[i]
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    # posições (categorias)
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc)+1)
    if i==0: 
        target_cell.value=posicoes_docentes[i]
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else: 
        target_cell.value=posicoes_docentes[i]
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    #if i <5 : print(target_sheet['AZ{}'.format(i+2)].value)
#str1 = ','.join(nomes_docentes)                                    
#str1 = '"'+str1+'"'
#str1= '"Y,N"' # funciona!
#str1='"Vítor Manuel Delgado Alves,Docente a atribuir"' # funciona!
#str1='"Susete Maria Gonçalves Marques,Teresa de Jesus da Silva Matos Nolasco Crespo,Teresa Maria Gonçalves Quilhó Marques dos Santos,Teresa Paula Gonçalves Cruz (EU),Vítor Manuel Delgado Alves,Docente a atribuir"'
#data_val = DataValidation(type='list',formula1=str1) #, allow_blank=False)
data_val = DataValidation(type='list',formula1='={}{}:{}{}'.format(coldoc,'$2',coldoc,'$'+str(1+len(nomes_docentes))), allow_blank=True) #, allow_blank=False)

# criar validação em target_sheet
target_sheet.add_data_validation(data_val)

# ler DSD existente
if DSDv1: 
    #wb1 = openpyxl.load_workbook(fnDSDv1) #, data_only=True) # with data_only, it will only read values
    names1=wb1.sheetnames
    # test if sheet names are right
    print(names1)
        
low_matches=[]
# subsituir os nomes já preenchidos pelo novo nome da lista da RH (nomes_docentes)
def most_similar_name(nome):
    if nome==DAA:
        return DAA
    if nome==VALIDATION_VALUE: 
        return VALIDATION_VALUE
    else:
        L=[fuzz.ratio(nome,x) for x in nomes_docentes]
        if np.max(L) < MIN_MATCH: 
            low_matches.append((nome,np.max(L)))
            return(DAA)
        else: 
            return nomes_docentes[np.argmax(L)]

#fuzz.ratio('Jorge Gominho',"Jorge Manuel Barros D'Almeida Gominho")
#nome='Ana Cristina Delaunay Caperta'
#most_similar_name('Jorge Gominho')

# depois de fazer a cópia de source_sheet para target_sheet:
# indicar as células em que ficam os drop-down menus e desproteger as células a preencher
# if DSDv1: copiar valores das células de DSDv_1 para DSD_v2 (target sheet):
# df para resumo

df=pd.DataFrame(columns=colunas_FOS+colunas_nome_UC+['linhaResp',colResponsavel,colNomeDocente]+colunas_horas_a_preencher+[column_somatorio_preencher,column_horas_em_falta_preencher])

# headers = [c.value for c in next(source_sheet.iter_rows(min_row=1, max_row=1))]
# colocar drop-down menu para colunas em coluna_validacao (que têm valor VALIDATION_VALUE='VAL')
# desproteger células de coluna_validacao e de colunas_horas_a_preencher
if DSDv1:
    ws1=wb1[ws_name_preencher]
    idxval=headers.index(coluna_validacao) # where drop-manu will be
    idx_not_empty=-1
    for k in range(target_sheet.max_row):
        ct=target_sheet.cell(k+1,idxval+1)
        if ct.value!=VALIDATION_VALUE: # para as linhas que não têm drop-down menu em DSD
            # for resumo_DSD2324 
            idx_not_empty +=1
            for colpreencher in colunas_FOS+colunas_nome_UC+[column_horas_em_falta_preencher]:
                idx=headers.index(colpreencher)
                dt=target_sheet.cell(k+1,idx+1)
                #if ct.value is not None: 
                #cr=wsr_ucs.cell(idx_not_empty+1,idx+1)
                #cr.value=dt.value
                df.loc[idx_not_empty,colpreencher]=dt.value
                # responsável
                df.loc[idx_not_empty,colResponsavel]=ct.value
                df.loc[idx_not_empty,colLinhaResp]=True
                # para completar abaixo
                df_aux=df.iloc[idx_not_empty]
        else:  # linhas que têm drop-down menu
            data_val.add(ct)
            # só coluna dos drop-down menus
            c1=ws1.cell(k+1,idxval+1) # mesma célula de ws1 e target_sheet<---------------------------------------
            if not c1.protection.locked and c1.value is not None and c1.value != '':
                # if False: 
                #     newname=most_similar_name(c1.value) # only if list of names is changed
                # else: 
                #     newname=c1.value
                ct.value=c1.value
                # for resumo_DSD2324 só escreve quando o nome do docente já tiver sido selecionado
                idx_not_empty +=1
                # para preencher df resumo, igual a df_aux definida acima
                for colpreencher in colunas_FOS+colunas_nome_UC+[column_horas_em_falta_preencher]:
                    df.loc[idx_not_empty,colpreencher]=df_aux.loc[colpreencher]
                df.loc[idx_not_empty,colResponsavel]=df_aux.loc[colResponsavel]
                #cr=wsr_ucs.cell(idx_not_empty+1,idxval+1)
                #cr.value=c1.value
                df.loc[idx_not_empty,colNomeDocente]=c1.value
                #df.loc[idx_not_empty,colResponsavel]=nomeResponsavel
                df.loc[idx_not_empty,colLinhaResp]=False  
            ct.protection = Protection(locked=False)
            # encher outras colunas em target cell
            for colpreencher in colunas_horas_a_preencher+[column_somatorio_preencher,column_horas_em_falta_preencher]:
                idx=headers.index(colpreencher)
                dt=target_sheet.cell(k+1,idx+1)
                d1=ws1.cell(k+1,idx+1) # mesma célula
                # for resumo_DSD2324
                df.loc[idx_not_empty,colpreencher]=d1.value # including [column_somatorio_preencher,column_horas_em_falta_preencher]
                if not d1.protection.locked and d1.value is not None and d1.value != '':
                    dt.value=d1.value
                    # for resumo_DSD2324
                    # if d1.value!=VALIDATION_VALUE: # and d1.value!=DAA:
                        #cr=wsr_ucs.cell(idx_not_empty+1,idx+1)
                        #cr.value=d1.value
                dt.protection = Protection(locked=False)


# copy data frame to Excel, do the wsr_docentes worksheet
df=df.iloc[1:] #2?!
df=df.dropna(subset=[colResponsavel])
dfref = df[df[colNomeDocente] != VALIDATION_VALUE]

df_to_excel(dfref[dfref[colLinhaResp]], wsr_ucs, header=True)
df_to_excel(dfref[dfref[colLinhaResp]==False], wsr_docentes, header=True)

# """ ct=target_sheet['E3']
# data_val.add(ct)
# ct.value
# c1=ws1['E3']
# c1.value
# ct.value=c1.value
# ct.value """
#print(d1.value)
wb1.close

# ajustar tamanho das colunas
target_sheet.column_dimensions[coldoc].width = max([len(x) for x in nomes_docentes])
target_sheet.column_dimensions[get_column_letter(column_index_from_string(coldoc)+1)].width = max([len(x) for x in posicoes_docentes])

#dim_holder = DimensionHolder(worksheet=target_sheet)
#dim_holder[coldoc] = ColumnDimension(target_sheet, min=col, max=col, width=30)
#target_sheet.column_dimensions = dim_holder


####################################################################### copiar worksheet 'DSD (informação UCs)'
if DSDv1: 
    source_sheet=wb1[ws_name_info]  # lê no ficheiro DSD já prépreenchido
else:
    source_sheet=wb_source[ws_name_info] # lê no ficheiro DA

# copiar DSD info para target
info_sheet = wb_target.create_sheet(ws_name_info)
#headers = [c.value for c in next(source_sheet.iter_rows(min_row=1, max_row=1))]

cols_a_apagar=[nomeColuna2letter(source_sheet,col_apagar_info_1),nomeColuna2letter(source_sheet,col_apagar_info_2)]

copy_sheet_attributes(source_sheet, info_sheet)
for r, row in enumerate(source_sheet.iter_rows()):
    for c, cell in enumerate(row):
        if get_column_letter(c+1) not in cols_a_apagar:
            if c==1: print(r)
            source_cell = cell
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = info_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            if get_column_letter(c+1) in info_cols_to_unprotect:
                target_cell.protection = Protection(locked=False)
            #target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

copyColumn(info_sheet,column_codigo_UC_info,'AY','numeric')
copyColumn(info_sheet,column_total_horas_info,'AZ','numeric')

# copiar DSD info para resumo
wsr_info = wbr.create_sheet(ws_name_info)
copy_sheet_attributes(source_sheet, wsr_info)
for r, row in enumerate(source_sheet.iter_rows()):
    for c, cell in enumerate(row):
        if get_column_letter(c+1) not in cols_a_apagar:
            if c==1: print(r)
            source_cell = cell
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = wsr_info.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            if get_column_letter(c+1) in info_cols_to_unprotect:
                target_cell.protection = Protection(locked=False)
            #target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


#headers = [c.value for c in next(info_sheet.iter_rows(min_row=1, max_row=1))]

# column_total_horas

# ##################################################################### criar fórmulas em ws_name_preencher
# '''
# column_codigo_UC_preencher='Código UC'
# column_horas_em_falta_preencher='Horas em falta na UC'
# column_key='Inserir docentes na UC ' #'docentes ' # cuidado: tem um espaço a mais
# coluna_validacao=column_key
# column_somatorio_preencher
# coldoc='AK'
# # ws info
# column_total_horas_info='Total Horas previsto '
# column_codigo_UC_info='Código UC'
# column_preencher_first='Total Horas Teóricas'
# column_preencher_last='Total Horas  Outras'
# '''

for r in rows_resp:
    d=target_sheet.cell(r+1,column_index_from_string(nomeColuna2letter(target_sheet, column_somatorio_preencher)))
    d.value='' #"=SUM({}{}:{}{})".format(nomeColuna2letter(target_sheet, column_preencher_first),r+2,nomeColuna2letter(target_sheet, column_preencher_last),r+1+N)
    # total horas de docência por docente para a UC
    for delta in range(N):
        k=r+1+delta+1
        e=target_sheet.cell(k,column_index_from_string(nomeColuna2letter(target_sheet, column_somatorio_preencher)))
        e.value="=SUM({}{}:{}{})".format(nomeColuna2letter(target_sheet, column_preencher_first),k,nomeColuna2letter(target_sheet, column_preencher_last),k)
    # horas em falta: total de horas que vem de info menos soma das horas dos docentes
    c=target_sheet.cell(r+1,column_index_from_string(nomeColuna2letter(target_sheet, column_horas_em_falta_preencher)))
    # código 1564 -- Sistemas de Produção Animal nos Trópicos
    if r+1<4948: # com as UC originais
        c.value="=VLOOKUP(int({}{}),'{}'!AY:AZ, 2, FALSE)-SUM({}{}:{}{})".format(nomeColuna2letter(target_sheet, column_codigo_UC_preencher),r+1,ws_name_info,nomeColuna2letter(target_sheet, column_preencher_first),r+2,nomeColuna2letter(target_sheet, column_preencher_last),r+1+N)
    else: # UC adicionada no fim do ficheiro fnIn
        c.value="=VLOOKUP(1564,'{}'!AY:AZ, 2, FALSE)-SUM({}{}:{}{})".format(ws_name_info,nomeColuna2letter(target_sheet, column_preencher_first),r+2,nomeColuna2letter(target_sheet, column_preencher_last),r+1+N)

#copyColumn(target_sheet,column_somatorio_preencher,'AZ')
for i in range(len(nomes_docentes)): #target_sheet.max_row):
    if i>0:
        g=target_sheet.cell(i+1,column_index_from_string(coldoc)-1)
        g.value="=SUMIF({}:{},{}{},{}:{})".format(nomeColuna2letter(target_sheet, column_key),nomeColuna2letter(target_sheet, column_key),coldoc,i+1,nomeColuna2letter(target_sheet, column_somatorio_preencher),nomeColuna2letter(target_sheet, column_somatorio_preencher))
        h=target_sheet.cell(i+1,column_index_from_string(coldoc)-2)
        h.value="=SUMIF({}:{},{}{},{}:{})/28".format(nomeColuna2letter(target_sheet, column_key),nomeColuna2letter(target_sheet, column_key),coldoc,i+1,nomeColuna2letter(target_sheet, column_somatorio_preencher),nomeColuna2letter(target_sheet, column_somatorio_preencher))
        h.number_format='0.00'

# =SUMIF(E:E,K2,AA) # adaptar

# column_index_from_string
#nomeColuna2letter(target_sheet, column_codigo_UC_preencher)
#=VLOOKUP(int(I2);'DSD (informação UCs)'!AY:AZ; 2; FALSE)
# =SUM(O3:Z12)


###################################################################### gravar wb_target
# eliminar a worksheet original
if 'Sheet' in wb_target.sheetnames:  # remove default sheet
    wb_target.remove(wb_target['Sheet'])
if 'Sheet' in wbr.sheetnames:  # remove default sheet
    wbr.remove(wbr['Sheet'])


# proteger
target_sheet.protection.sheet = True
target_sheet.protection.password = 'kathleen'
info_sheet.protection.sheet = True
info_sheet.protection.password = 'kathleen'

# gravar para ficheiro
wb_target.save(fnOut)
wbr.save(fnResumo)

wb_target.close
wbr.close

# what is this?
# isinstance(c, openpyxl.cell.read_only.EmptyCell)

print(low_matches)


