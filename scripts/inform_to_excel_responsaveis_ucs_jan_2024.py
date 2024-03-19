# input:
# ficheiro dos serviços com info sobre RH, UCs etc e metadados
# parâmetros: data final contrato, ...
# output:
# ficheiro a partilhar para colocação dos nomes de responsáveis UCs e comentários: 'DSD_2024_2025_responsaveis_UCs.xlsx'


from copy import copy
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname # https://openpyxl.readthedocs.io/en/latest/validation.html?highlight=validation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from pathlib import Path
import pandas as pd
from unidecode import unidecode
from functions import simplify_strings, df_to_excel_with_columns, compact_excel_file, get_letter_from_column_name,unlock_cells,stripe_cells
from functions import add_suffix_to_duplicates, reorder_and_filter_dataframe, insert_row_at_beginning, insert_row_at_end,sort_list


UNPROTECT_OUTPUT_CELLS=True # para desproteger células com função unlock_cells(ws, col_name, min_row=None, max_row=None):
PROTECT_WORKSHEET=True # tem que ser True para impedir escrita
PASSWORD='kathleen'

# input DSD file
FOLDER_SERVICOS= 'ficheiros_servicos_ISA'
FOLDER_OUTPUT='output_files'
FOLDER_FICH_RESPONSAVEIS_UCs='ficheiros_responsaveis_ucs'
FORCE_OUTPUT_NAME=True # para forçar nome do output; caso contrário, é derivado do nome do ficheiro de input
DSD_OUTPUT_FICH='DSD_2024_2025_responsaveis_UCs.xlsx' # if FORCE_OUTPUT_NAME
#DSD_INPUT_FICH='DSD_inform_2024_2025_v5.xlsx'
#DSD_INPUT_FICH='2024_01_26 DSD_inform_202324_v6-1-1.xlsx (Dados MCaron e Carlos).xlsx'
DSD_INPUT_FICH='2024_01_26 DSD_inform_202324_v6-1-1.xlsx (Dados MCaron e Carlos)_compact_ML3.xlsx'
stem=Path(DSD_INPUT_FICH).stem
suffix=Path(DSD_INPUT_FICH).suffix
COMPACT='_compact'
BLOQ='_bloq'

# Load the source workbook
#input_folder=Path(r'C:\Users\mlc\OneDrive - Universidade de Lisboa\Documents\profissional-isa-cv\cg-isa\DSD_2024_2025\backup_inputs_DSD')
try:
    working_dir=Path(__file__).parent.parent # working directory from script location: scripts are in 'scripts' folder
except:
    working_dir=Path().absolute()


input_folder= working_dir / 'DSD_2024_2025' / FOLDER_SERVICOS
output_folder= working_dir / 'DSD_2024_2025' / FOLDER_OUTPUT
compact_input_file= output_folder  / (stem+COMPACT+suffix)
if FORCE_OUTPUT_NAME: 
    output_file=output_folder / DSD_OUTPUT_FICH
else:
    output_file=output_folder / (stem+BLOQ+suffix)

# worksheets and column names
UC ='uc_2024-25'
UC_uc = 'unidade_curricular'
UC_resp='responsavel_unidade_curricular'
UC_sugestoes='sugestões de modificação da info da UC'
UC_autor_sugestao='autor da sugestão'
UC_area_cientifica='area_cient' # drop
UC_numero_alunos='NumeroAlunos' # drop
UC_ciclo_curso= 'ciclo_curso'
UC_ciclo_curso_curso='curso' # para filtrar cncg's
UCMETA='uc_meta'
RH='RH'
RH_nome='nome' # nomes docentes
N_extra_nomes=0
RH_numero = 'num_pessoal'
RH_data_fim='data_fim'
RH_posicao='posicao'
RH_obs = 'Obs'
RHPOSICAO='RH_posicao' # drop
RHMETA='RH_meta'
PE='planos_estudos' # drop
AC='AC' #drop
ACMETA='AC_meta'#drop
POS='POS' #drop
UCAREA='uc_AreaCient' # drop

# values: Sheet_column_value
RH_nome_pro_bono='docente_PRO_BONO' 
RH_nome_em_contratacao='Docente em contratação'
RH_data_fim_sem_termo='sem termo'
DATA_TERMO_CERTO='2024-09-01'

# desproteger células:
# responsável
# docentes extra + observações

'''
, 'UC_meta': 'uc_meta', 'AC': 'uc_AreaCient', 'AC_meta': 'AC_meta', 'CC': 'cod_curso', 'RH': 'RH', 'RH_meta': 'RH_meta', 'POS': 'RH_posicao'}
UC = {'UC_ciclo_': 'ciclo_curso', 'UC_uc_obr': 'uc_obrigatoria', 'UC_uc_opt': 'uc_optativa', 'UC_cod_uc': 'cod_uc', 'UC_unidad': 'unidade_curricular', 'UC_Respon': 'Responsavel', 'UC_não_fu': 'não_funciona_2024_2025', 'UC_area_c': 'area_cient', 'UC_ano': 'ano', 'UC_sem': 'sem', 'UC_h_trab': 'h_trab_totais', 'UC_T': 'T', 'UC_TP': 'TP', 'UC_PL': 'PL', 'UC_TC': 'TC', 'UC_S': 'S', 'UC_E': 'E', 'UC_OT': 'OT', 'UC_O': 'O', 'UC_h_cont': 'h_contacto_total', 'UC_ECTS': 'ECTS', 'UC_uc_int': 'uc_interna', 'UC_obs': 'obs'}
UC_meta = {'UC_meta_ciclo_': 'ciclo_curso', 'UC_meta_1 cicl': '1 ciclo', 'UC_meta_Licenc': 'Licenciatura'}
AC = {'AC_area_c': 'area_cientifica', 'AC_sigla': 'sigla', 'AC_classi': 'classificacao_area'}
AC_meta = {'AC_meta_classi': 'classificacao_area', 'AC_meta_CNAEF': 'CNAEF', 'AC_meta_Classi': 'Classificação Nacional das Áreas de Educação e Formação (CNAEF), Portaria n.o 256/2005 de 16 de Março. Classificação: nível 1, 2 e 3 (código: 3 dígitos)'}
CC = {'CC_cod_cu': 'cod_curso', 'CC_grau': 'grau', 'CC_curso': 'curso', 'CC_obs': 'obs'}
RH = {'RH_num_pe': 'num_pessoal', 'RH_nome': 'nome', 'RH_corpo': 'corpo', 'RH_posica': 'posicao', 'RH_ETI': 'ETI', 'RH_contra': 'contratacao_vinculo', 'RH_data_f': 'data_fim', 'RH_Obs': 'Obs'}
RH_meta = {'RH_meta_num_pe': 'num_pessoal', 'RH_meta_número': 'número de indentificação pessoal atribuído pelos RH', 'RH_meta_Unname': 'Unnamed: 4'}
POS = {'POS_posica': 'posicao', 'POS_h_min': 'h_min', 'POS_h_max': 'h_max', 'POS_obs': 'obs'}
''' 

# cor light red, light yellow
fill_red = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
fill_green = PatternFill(start_color="C0FFCB", end_color="C0FFCB", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # alpha: 1st 2 characters
fill_light_yellow = PatternFill(start_color="FFFFED", end_color="FFFFED", fill_type="solid") 
thin_border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# # Load the source workbook
# source_workbook = load_workbook(input_file, read_only=True, data_only=True)
# Try to read smaller file; otherwise read original file and create smaller file
try:
    #source_workbook = load_workbook(compact_input_file)
    sheet_names = pd.ExcelFile(compact_input_file).sheet_names
except:
    print('ler ficheiro original não compactado')
    #source_workbook = load_workbook(input_folder  / DSD_INPUT_FICH)  # demora
    compact_excel_file(input_folder  / DSD_INPUT_FICH, compact_input_file)
    print('ficheiro original compactado')
    sheet_names = pd.ExcelFile(compact_input_file).sheet_names

# Create a new workbook
#new_workbook = pd.ExcelWriter(output_file, engine='openpyxl')
new_workbook = Workbook()

# contar número de docentes em RH
df = pd.read_excel(compact_input_file, sheet_name=RH)
idx_RH_nome,letter_RH_nome=get_letter_from_column_name(df,RH_nome)  #df.columns.get_loc(RH_nome) + 1
idx_RH_posicao,letter_RH_posicao=get_letter_from_column_name(df,RH_posicao)  #df.columns.get_loc(RH_nome) + 1
idx_RH_obs,letter_RH_obs=get_letter_from_column_name(df,RH_obs)  #df.columns.get_loc(RH_nome) + 1
# adicionar N_extra_nomes nomes em branco em RH_home
# adicionar col
    
# contar número de Ucs em UC_2024_2025
df = pd.read_excel(compact_input_file, sheet_name=UC)
numero_ucs=len(df[UC_uc])
idx_UC_uc,letter_UC_uc=get_letter_from_column_name(df,UC_uc) 
#idx_uc=df.columns.get_loc(UC_uc) + 1

# Iterate through sheets in the source workbook
# Ensure that RH comes before UC and remove 'planos_estudos', etc
sheet_names=[RH,RHMETA, UC,UCMETA]+list(set(sheet_names).difference(set([RH,RHMETA, UC,UCMETA,PE,AC,ACMETA,POS, RHPOSICAO, UCAREA])))
for sheet_name in sheet_names: #source_workbook.sheetnames:
    # for each sheet_name, we copy the contents of the input sheet, create a df, modify df, create validation, and write to workbook with df_to_excel_with_columns
    print(sheet_name) 
    # Create a new sheet in the new workbook
    new_worksheet = new_workbook.create_sheet(title=sheet_name)

    # Read the sheet into a pandas DataFrame
    df = pd.read_excel(compact_input_file, sheet_name=sheet_name)
    # removes accents and replaces ' ' by '_'
    df.columns=simplify_strings(df.columns)

    # ordenar docentes por ordem alfabética, com os novos docentes à cabeça, mais "docente_PRO_BONO"; excluir docentes a termo, com termo antes de set 2024
    # Nota: 'docentes em contratação' podem ter o mesmo nome e por isso é preciso lidar com duplicações
    if sheet_name==RH:
        df=insert_row_at_beginning(df,{RH_nome: RH_nome_pro_bono, RH_data_fim: RH_data_fim_sem_termo, RH_obs: 'Docente ou especialista não do ISA que participa na docência sem receber pagamento do ISA: o nome do docente pode ser indicado na coluna de observações'})
        df=add_suffix_to_duplicates(df,RH_nome)
        docentes_em_contratacao=list(df[df[RH_nome].str.contains(RH_nome_em_contratacao, case=False, na=False)][RH_nome])
        # remover docentes com contrato que acaba até 1 de setembro de 2024
        sem_termo=list(df[df[RH_data_fim].str.contains(RH_data_fim_sem_termo, case=False, na=False)][RH_nome])
        com_termo=list(set(list(df[RH_nome])).difference(set(sem_termo)))
        df_com_termo=df[df[RH_nome].isin(com_termo)]
        df_set_2024=df_com_termo[pd.to_datetime(df_com_termo[RH_data_fim]) > DATA_TERMO_CERTO]
        com_termo_set_2024=list(df_set_2024[RH_nome])
        # criar lista em ordem alfabética de docentes que não estão em contratação
        L=list(set(list(sem_termo+com_termo_set_2024)).difference(set(docentes_em_contratacao).union(set([RH_nome_pro_bono]))))
        outros_docentes=sort_list(L, simplify_strings(L))
        # list de nomes de todos os potenciais docentes 
        todos_docentes=[RH_nome_pro_bono]+docentes_em_contratacao+outros_docentes
        # ordenar df segundo lista todos_docentes
        df=reorder_and_filter_dataframe(df, RH_nome, todos_docentes)
        # acrescentar novas linhas para eventuais docentes não listados
        numero_docentes=len(df[RH_nome])
        for i in range(N_extra_nomes):
            df=insert_row_at_end(df,{RH_nome: 'a_designar_'+str(i+1), RH_posicao:'indicar se possível', RH_obs: 'Indicar nome de docente em falta e posição nas colunas respetivas. Aqui, indicar quem faz a porposta e justificar.' })
        
    # Criar drop-down menu para inserir nome responsável da UC
    if sheet_name==UC:
        # remover coluna área cientifica, etc
        df=df.drop(columns=[UC_area_cientifica,UC_numero_alunos])
        # re-ordenar UCs pelo nome, mas com cursos CNCG no fim
        df=add_suffix_to_duplicates(df,UC_uc)
        cncg=list(df[df[UC_ciclo_curso].str.contains(UC_ciclo_curso_curso, case=False, na=False)][UC_uc])
        cncg=sort_list(cncg, simplify_strings(cncg))
        uc_ciclos=list(set(list(df[UC_uc])).difference(set(cncg)))
        uc_ciclos=sort_list(uc_ciclos, simplify_strings(uc_ciclos))
        df=reorder_and_filter_dataframe(df, UC_uc, uc_ciclos+cncg)
        # Criar coluna responsável
        if UC_resp not in df.columns: 
            df.insert(idx_UC_uc-1,UC_resp,'') # automatizar .A largura da coluna tem que ser grande para se verem os nomes
        idx_UC_resp,letter_UC_resp=get_letter_from_column_name(df,UC_resp)
        # Criar coluna 'autor_sugestão'
        if UC_sugestoes not in df.columns: 
            df.insert(df.shape[1],UC_autor_sugestao,'') 
        idx_UC_autor_sugestao,letter_UC_autor_sugestao=get_letter_from_column_name(df,UC_autor_sugestao) 
        # Criar coluna 'sugestões'
        if UC_sugestoes not in df.columns: 
            df.insert(df.shape[1],UC_sugestoes,'') 
        idx_UC_sugestoes,letter_UC_sugestoes=get_letter_from_column_name(df,UC_sugestoes) 
        # validation responsáveis
        dv = DataValidation(type="list", formula1=f"{quote_sheetname('RH')}!${letter_RH_nome}$2:${letter_RH_nome}${numero_docentes+1+N_extra_nomes}") # 10 extra
        new_worksheet.add_data_validation(dv)
        dv.add(f"${letter_UC_resp}$2:{letter_UC_resp}${numero_ucs+1}") # creates drop-down menu. #automatizar
        # validation autor_sugestao
        dv.add(f"${letter_UC_autor_sugestao}$2:{letter_UC_autor_sugestao}${numero_ucs+1}") # creates drop-down menu.

    # Write the DataFrame to the new workbook
    #df.to_excel(new_workbook, sheet_name=sheet_name, index=False, startrow=0, header=True)
    df_to_excel_with_columns(df,new_worksheet,maxwidth=20,header=True,index=False,startrow=0, startcol=0)

    # Apply filters to the first row
    if '_meta' not in sheet_name:
        new_worksheet.auto_filter.ref = new_worksheet.dimensions
        new_worksheet.freeze_panes = "A2"
    
    # a ideia é desbloquear algumas células e a seguir bloquear toda a worksheet
    # desbloquear células dos responsáveis e das sugestões das UC
    if sheet_name==UC and UNPROTECT_OUTPUT_CELLS:
        stripe_cells(new_worksheet, fill_color=fill_light_yellow,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_resp,fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_sugestoes,fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_UC_autor_sugestao,fill_color=fill_green,border=thin_border)
    
    # Dar possibilidade de criar novos docentes
    if sheet_name==RH and UNPROTECT_OUTPUT_CELLS:
        stripe_cells(new_worksheet, fill_color=fill_light_yellow,border=thin_border)
        # docentes extra (a poderem ser adicionados)
        unlock_cells(new_worksheet,letter_RH_nome, min_row=numero_docentes+2, max_row=numero_docentes+N_extra_nomes+1, fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_RH_posicao, min_row=numero_docentes+2, max_row=numero_docentes+N_extra_nomes+1, fill_color=fill_green,border=thin_border)
        unlock_cells(new_worksheet,letter_RH_obs, min_row=numero_docentes+2, max_row=numero_docentes+N_extra_nomes+1, fill_color=fill_green,border=thin_border)

    if PROTECT_WORKSHEET:
        new_worksheet.protection.sheet = True
        new_worksheet.protection.enable()
        new_worksheet.protection = SheetProtection(sheet=True, objects=False, scenarios=False, formatCells=False, formatRows=False, formatColumns=False, insertColumns=True, insertRows=True, insertHyperlinks=True, deleteColumns=True, deleteRows=True, selectLockedCells=False, selectUnlockedCells=False, sort=False, autoFilter=False, pivotTables=True, password=None, algorithmName=None, saltValue=None, spinCount=None, hashValue=None) #(autoFilter=True, formatColumns=True)

# eliminar a worksheet 'Sheet' que foi criada automaticamente
if 'Sheet' in new_workbook.sheetnames:  # remove default sheet
    new_workbook.remove(new_workbook['Sheet'])

# workbook protection
new_workbook.security.workbookPassword = PASSWORD
new_workbook.security.lockStructure = True

# Save the new workbook
try: 
    new_workbook.save(output_file)
except PermissionError:
    print('File must be in use: close if first please')
new_workbook.close