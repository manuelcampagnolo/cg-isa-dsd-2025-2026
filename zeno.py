######################################################
# Manuel Campagnolo (abril 2023)
# ISA/ULIsboa
# script para preparar ficheiro Excel DSD ISA 2023-2024 -- TEMPLATE: não adaptar diretamente
#######################################################

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors

from copy import copy
import os
from fuzzywuzzy import fuzz # compare strings
import numpy as np

import pandas as pd
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

##############################################################################
# Funções
##############################################################################
# converter DataFrame para openpyxl WorkSheet (ao nível da célula)
from openpyxl.utils.dataframe import dataframe_to_rows
def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    # alternativa mais simples:
    # rows = ws.iter_rows()
    # for row in rows:
    #     for cell in row:
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             cell=ws.cell(row=r_idx, column=c_idx)
             cell.value = value
             # or cell = ws.cell(row=i, column=j, value=val)
             # target_cell.data_type = source_cell.data_type
             # target_cell._hyperlink = copy(source_cell.hyperlink)
             # cell.border = border
             # cell.font = Font(bold=True)
             # cell.number_format
             # cell.alignment
             # cell.protection
             # cell.protection = Protection(locked=False)
             # if cell.protection.locked ...
             # if isinstance(cell, openpyxl.cell.read_only.EmptyCell) ...
             # if isinstance(source_cell, openpyxl.cell.ReadOnlyCell) ....
             # my_data_validation.add(cell)

# idem, mas com formatação da largura das colunas para ajustar ao conteúdo
def df_to_excel_with_columns(df,ws,maxwidth=30,header=True,index=False):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = min(max_length+2, maxwidth) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_excel(df,ws,header=True,index=False)

#########################################################################
# Fim funções 
##########################################################################

# ficheiros de input
# ficheiro RH enviado pela Madalena 18 de maio 2023
fn_docentes='nomes_docentes_codigos_RH_17maio2023.xlsx'
fn_out='out.xlsx'

########################################################
# LER ficheiros Excel
######################## ler ficheiro Excel e verificar quais são as worksheets
wb = openpyxl.load_workbook(fn_docentes,data_only=True) # with data_only, it will only read values
# verificar sheetnames
wsnames=wb.sheetnames
print('ficheiro RH: ', wsnames)
##################################### possibilidade (1): abrir em openpyxl
# abrir worksheet
ws = wb[wsnames[0]]
# ler nomes das colunas colunas
print([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])
##################################### possibilidade (2): ler para DataFrame
dfdocs= pd.read_excel(fn_docentes, sheet_name=wsnames[0]) 
dfdocs.head()

# TRABALHAR EM PANDAS...


######################################################## dataframe para openpyxl
# CRIAR NOVO openpyxl WorkBook (target)
################################# criar workbook
wb_target = openpyxl.Workbook()
################################# criar worksheet 
ws_target=wb_target.create_sheet('my_sheet')
########################### Exportar DataFrame para openpyxl WorkSheet
#df_to_excel_with_columns(dfdocs, ws_target)
#####################################################

# TRABALHAR EM OPENPYXL ...

####################################### Criar lista para drop-down menu
mycol='Nome completo'
mylist = dfdocs[mycol]
l0=len(mylist)
# limpar
mylist = [item for item in mylist if item != "" or item is not None]
if (len(mylist) < l0): print('removed items')

####################################### Data Validation
# Example 1: create a data validation rule of type "whole" between the values 1 and 10 and apply it to cells A1 to A10 in the worksheet.
# data_validation = DataValidation(type="whole", operator="between", formula1=1, formula2=10)
# worksheet.add_data_validation(data_validation)
# data_validation.add(worksheet["A1:A10"])
# Example 2: create a data validation rule of type "list" with the allowed values "Value1", "Value2", and "Value3", and apply it to cells A1 to A10 in the worksheet. 
# data_validation = DataValidation(type="list", formula1='"Value1,Value2,Value3"')
# worksheet.add_data_validation(data_validation)
# data_validation.add(worksheet["A1:A10"])
# Example 3: same, but list of values comes from Python list
# values_string = ",".join(['"{}"'.format(value) for value in values_list])
# data_validation = DataValidation(type="list", formula1=values_string)
# worksheet.add_data_validation(data_validation)
# data_validation.add(worksheet["A1:A10"])

# values_string = ','.join(mylist) # mylist[0:9] funciona mas mylist[0:10] não !! Não parece haver nenhuma relação com espaços ou acentos
# values_string = '"{}"'.format(values_string)
# values_string='"A,B"' # funciona
# my_data_validation = DataValidation(type="list", formula1=values_string, allow_blank=True)

# alternativa para type="list"
# 1. Copiar os valores para outra worksheet
ws_aux=wb_target.create_sheet('InfoAux')
df_to_excel_with_columns(pd.DataFrame(mylist,columns=['Nomes docentes']), ws_aux)
coldoc='A'
N=len(mylist)
# 2. Criar data_validation a partir dos valores na worksheet
my_data_validation = DataValidation(type='list',formula1='=InfoAux!{}{}:{}{}'.format(coldoc,'$2',coldoc,'$'+str(1+N)), allow_blank=True) 
ws_target.add_data_validation(my_data_validation)

# outro data_validation
values_string='"A,B"' # funciona
my_data_validation_2 = DataValidation(type="list", formula1=values_string, allow_blank=True)
#my_data_validation_2 = DataValidation(type="whole", operator="between", formula1=1, formula2=10)
ws_target.add_data_validation(my_data_validation_2)

# escolher, associar menu e desproteger célula
cell=ws_target["A1"]
my_data_validation.add(cell) # só pode ser 1 célula
cell.protection = Protection(locked=False)
cell=ws_target["B1"]
my_data_validation_2.add(cell) # só pode ser 1 célula
cell.protection = Protection(locked=False)

########################################################
# PROTEGER E EXPORTAR :  de openpyxl para Excel
####################################################
# eliminar a worksheet original
if 'Sheet' in wb_target.sheetnames:  # remove default sheet
    wb_target.remove(wb_target['Sheet'])
####################################### Proteger células antes de gravar
# protege exceto células desprotegidas com cell.protection = Protection(locked=False)
ws_target.protection.sheet = True
ws_target.protection.password = 'minha_password'
ws_aux.protection.sheet = True
ws_aux.protection.password = 'minha_password'

##################### de openpyxl para Excel
wb_target.save(fn_out)
################################### Fechar ligação a openpyxl WorkBook
wb.close
wb_target.close
#########################################################
