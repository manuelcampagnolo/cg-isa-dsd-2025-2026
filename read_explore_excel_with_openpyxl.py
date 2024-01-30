import openpyxl
from pathlib import Path

p=Path().absolute() # or just p=Path('.').absolute()
# file path and name
fn= p / 'ficheiros_DSD' / 'reisa_mestrados_v1.xlsx'
# read workbook
wb = openpyxl.load_workbook(fn) #,data_only=True) # with data_only, it will only read values
# verificar sheetnames
wsnames=wb.sheetnames
print('ficheiro RH: ', wsnames)