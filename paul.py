######################################################
# Manuel Campagnolo (abril 2023)
# ISA/ULIsboa
# script para preparar ficheiro Excel DSD ISA 2023-2024 docência em outras escolas
# Usa um dicionário de colunas e opções do menu drop down para construir um ficheiro Excel
#######################################################


import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.styles import Border, Side

from copy import copy
import os
from fuzzywuzzy import fuzz # compare strings
import numpy as np

import pandas as pd
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

##############################################################################
# Funções
##############################################################################
# converter DataFrame para openpyxl WorkSheet (ao nível da célula)
from openpyxl.utils.dataframe import dataframe_to_rows
def df_to_openpyxl(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""
    # alternativa mais simples:
    # rows = ws.iter_rows()
    # for row in rows:
    #     for cell in row:
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow):
        for c_idx, value in enumerate(row, startcol):
             cell=ws.cell(row=r_idx+1, column=c_idx+1)
             cell.value = value
             # or cell = ws.cell(row=i, column=j, value=val)
             # target_cell.data_type = source_cell.data_type
             # target_cell._hyperlink = copy(source_cell.hyperlink)
             # cell.border = border
             # cell.font = Font(bold=True)
             # cell.number_format
             # cell.alignment
             # cell.protection
             # cell.protection = Protection(locked=False)
             # if cell.protection.locked ...
             # if isinstance(cell, openpyxl.cell.read_only.EmptyCell) ...
             # if isinstance(source_cell, openpyxl.cell.ReadOnlyCell) ....
             # my_data_validation.add(cell)

# idem, mas com formatação da largura das colunas para ajustar ao conteúdo
def df_to_openpyxl_with_columns(df,ws,maxwidth=30,header=True,index=False):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = min(max_length+2, maxwidth) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_openpyxl(df,ws,header=True,index=False)

#########################################################################
# Fim funções 
##########################################################################

# ficheiros de input
# ficheiro RH enviado pela Madalena 18 de maio 2023
fn_docentes='nomes_docentes_codigos_RH_17maio2023.xlsx'
fn_out='DSD_2023_2024_servico_externo_v6.xlsx'
nomeWS='Servico_externo'
Nmax=100 # número de linhas do ficheiro a preencher

########################################################
# LER ficheiros Excel
######################## ler ficheiro Excel e verificar quais são as worksheets
wb = openpyxl.load_workbook(fn_docentes,data_only=True) # with data_only, it will only read values
# verificar sheetnames
wsnames=wb.sheetnames
print('ficheiro RH: ', wsnames)
##################################### possibilidade (1): abrir em openpyxl
# abrir worksheet
ws = wb[wsnames[0]]
# ler nomes das colunas colunas
print([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])
##################################### possibilidade (2): ler para DataFrame
dfdocs= pd.read_excel(fn_docentes, sheet_name=wsnames[0]) 
dfdocs.head()

# TRABALHAR EM PANDAS...

############################################## lista docentes
mycol='Nome completo'
dfdocs['Grupo de empregados'].unique()
doc_invest=dfdocs.loc[(dfdocs['Grupo de empregados']=='Docentes')|(dfdocs['Grupo de empregados']=='Investigadores'),mycol]
nao_docentes=dfdocs.loc[dfdocs['Grupo de empregados']=='Não Docente',mycol]
[x for x in nao_docentes if 'Mariana' in x]
mylist = doc_invest.copy()
l0=len(mylist)
# limpar
mylist = [item for item in mylist if item != "" or item is not None]
if (len(mylist) < l0): print('removed items')
listadocentes=mylist.copy()

######################################################## dataframe para openpyxl
# CRIAR NOVO openpyxl WorkBook (target)
################################# criar workbook
wb_target = openpyxl.Workbook()
################################# criar worksheet 
ws_target=wb_target.create_sheet(nomeWS)
########################### Exportar DataFrame para openpyxl WorkSheet
#df_to_openpyxl_with_columns(dfdocs, ws_target)
#####################################################

# TRABALHAR EM OPENPYXL ...

####################################### Criar lista para drop-down menu
start_year=2023
num_years=10
academic_years = [f"{year}-{year+1}" for year in range(start_year, start_year + num_years)]
semestres = [f"{s}o" for s in range(1,11)]
SN=['Sim','Não']

mydict={
 'Nome do docente': listadocentes,
        'Nome da UC': None,
         'Nome do curso': None,
         'Nível': ['1º ciclo','2º ciclo','3º ciclo','Não conferente de grau'],
         'Semestre de funcionamento': ['1º','2º','extra-semestre'],
         'Parceria institucional estabelecida': ['Sim','Não'],
         'Com centro de custos ISA ou de empresas do ISA': ['Sim','Não'],
         'Ocorrência': ['Todos os anos', 'Não regular'],
         'Instituição responsável': None,
         'Link do curso': None,
         'Responsável do Curso': None,
         'Responsável de UC': None,
         'Email responsável UC': None,
         'Funcionamento': ['obrigatória','optativa'],
         'Funciona dependente do nº alunos': ['Sim','Não'],
         'Número mínimo de alunos para funcionar': None,
         'Número horas letivas':None
}

len(mydict)


# Create a border with specific line style and color
border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# cor light red, light yellow
fill_red = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
fill_green = PatternFill(start_color="C0FFCB", end_color="C0FFCB", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # alpha: 1st 2 characters


####################################### Criar Data Validation para cada item do dicionário que não é None
ws_aux=wb_target.create_sheet('InfoAux')
# 1. criar dataframe para cada item e copiar para ws_aux
maxwidth=30
startcol=0
for key, value in mydict.items():
    if value is not None:
        df=pd.DataFrame(value,columns=[key])
        df_to_openpyxl(df, ws_aux, header=True, index=False, startrow=0, startcol=startcol)
        startcol+=1

# 2a: change row height
for k in range(Nmax):
    ws_target.row_dimensions[k+1].height = 25 # mudar altura da linha

# 2. criar data validation para cada item e associar às células de target
startcol=0 # percorre as colunas em InfoAux
idx_target=0 # percorre as colunas em target
k=0
for key, value in mydict.items():
    col_target=get_column_letter(idx_target+1) # col letter in target
    cell=ws_target[col_target+'1'] # 1a linha
    cell.value=key
    cell.fill=fill_green
    cell.border = border
    if value is not None:
        cell.fill=fill_red # 1a linha ainda
        # column width
        lens = [len(element) for element in value]
        lens.append(len(key))
        width = min(max(lens)+2, maxwidth) # set a maximum width of 30
        col_aux=get_column_letter(startcol+1) # column index in InfoAux
        N=len(value)
        my_data_validation = DataValidation(type='list',formula1='=InfoAux!{}{}:{}{}'.format(col_aux,'$2',col_aux,'$'+str(1+N)), allow_blank=True) 
        ws_target.add_data_validation(my_data_validation)
        # column width
        ws_aux.column_dimensions[col_aux].width = width
        ws_target.column_dimensions[col_target].width = width
        # cell values
        for k in range(Nmax):
            cell=ws_target[get_column_letter(idx_target+1)+str(k+2)] # colocar a partir da 2a linha
            if idx_target==0 and k==0: cell.value='Clicar para selecionar docente'
            my_data_validation.add(cell) # só pode ser 1 célula
            cell.protection = Protection(locked=False)
            cell.border = border
            # pintar metade das linhas 
            if k%2 == 0: cell.fill=fill_yellow
        startcol+=1
    else:
        ws_target.column_dimensions[col_target].width = min(len(key)+2, maxwidth)
        for k in range(Nmax):
            cell=ws_target[get_column_letter(idx_target+1)+str(k+2)] # colocar a partir da 2a linha
            cell.protection = Protection(locked=False)
            cell.border = border
            if k%2 == 0: cell.fill=fill_yellow
    idx_target+=1

def df_to_openpyxl_with_columns(df,ws,maxwidth=30,header=True,index=False):
    for column in df.columns:
        # get the column letter
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        # determine the optimal width based on the contents of the column
        max_length = df[column].astype(str).map(len).max()
        width = min(max_length+2, maxwidth) # set a maximum width of 30
        # set the column width
        ws.column_dimensions[column_letter].width = width
        # write 
        df_to_openpyxl(df,ws,header=True,index=False)

########################################################
# PROTEGER E EXPORTAR :  de openpyxl para Excel
####################################################
# eliminar a worksheet original
if 'Sheet' in wb_target.sheetnames:  # remove default sheet
    wb_target.remove(wb_target['Sheet'])
####################################### Proteger células antes de gravar
# protege exceto células desprotegidas com cell.protection = Protection(locked=False)
ws_target.protection.sheet = True
ws_target.protection.password = 'minha_password'
ws_aux.protection.sheet = True
ws_aux.protection.password = 'minha_password'
# hide worksheet
ws_aux.sheet_state = 'hidden'
# Freeze the top row
ws_target.freeze_panes = "A2"

##################### de openpyxl para fichero Excel
wb_target.save(fn_out)
################################### Fechar ligação a openpyxl WorkBook
wb.close
wb_target.close
#########################################################
""" 
# https://stackoverflow.com/questions/54071663/how-to-add-drop-down-list-in-excel-cell-using-win32com-python
import win32com.client as win32

# Open Excel application
excel_app = win32.Dispatch("Excel.Application")
# Open workbook
path=r'C:\temp\cg-isa'
workbook = excel_app.Workbooks.Open(os.path.join(path,fn_out))
# Select worksheet
worksheet = workbook.Worksheets(nomeWS)
# the data validation in the range A1:A10 will use the values from the B1:B5 range as the validation list
# Define the range for data validation
validation_range = worksheet.Range("A1:A"+str(Nmax+1))
validation= worksheet.Cells(2,1).Validation
validation = validation_range.Validation
# Define the range for validation list
M=len(listadocentes)
col_aux='A'
formula='=InfoAux!{}{}:{}{}'.format(col_aux,'2',col_aux,str(1+M))
print(formula)
# Add data validation to the range


validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1=formula)

# Enable autocomplete
validation.InCellDropdown = True
validation.ShowInput = False
validation.ShowError = False



workbook.Save()
workbook.Close()
# Quit Excel application
excel_app.Quit()
 """