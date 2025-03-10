######################################################
# Manuel Campagnolo (abril 2023)
# ISA/ULIsboa
# script para preparar ficheiro Excel DSD ISA 2023-2024
#######################################################

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Protection
from copy import copy
import os
from fuzzywuzzy import fuzz # compare strings
import numpy as np

#fuzz.ratio('abscerg', 'abcderfg')

# folder not necessary when working in a folder
#folder=r"C:\Users\mlc\OneDrive - Universidade de Lisboa\Documents\profissional-isa-cv\cg-isa"
#pathIn=os.path.join(folder,"DSD_2324_.xlsx")
#pathOut=os.path.join(folder,"DSD_2324_new.xlsx")
#fnIn="DSD_2324_.xlsx"
fnDSDv1="DSD_2324_28abril.xlsx" #  "DSD_v1_teste.xlsx"; 
DSDv1=True # para copiar a informação já inserida 'preencher'+ folha 'info'
#fnIn="DSD_2324_ML_12abr2023__.xlsx"
fnIn="DSD_2324_ML_12abr2023_acrescentar_UC_no_final__.xlsx" # agora tem folha 'DocentesNovo'
#fnOut="DSD_2324_ML_12abr2023_new.xlsx"
fnOut="DSD_2324.xlsx"
fnResumo="resumo_DSD_2324.xlsx"

############################################################ funcoes
# devolve letra da coluna com nome (1a linha) da worksheet ws
def nomeColuna2letter(ws,nome):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx=headers.index(nome)
    return get_column_letter(idx+1)

# copy column named nameCol to a a column located at colLetter
def copyColumn(ws,nameCol,colLetter,colType=None):
    count=0
    for cell in ws['{}:{}'.format(nomeColuna2letter(ws,nameCol),nomeColuna2letter(ws,nameCol))]:
        count+=1
        if colType=='numeric': # tenta converter para numérico
            try: 
                new_cell=ws.cell(row=cell.row, column=column_index_from_string(colLetter))
                new_cell.value=int(cell.value)
                new_cell.number_format='0'
            except:
                ws.cell(row=cell.row, column=column_index_from_string(colLetter), value=cell.value)
        else:
            ws.cell(row=cell.row, column=column_index_from_string(colLetter), value=cell.value)

def copy_sheet(source_sheet, target_sheet, idx):
    rows_resp=copy_cells(source_sheet, target_sheet, idx)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)
    return rows_resp

def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = False #copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

# idx indica a coluna em que estão os responsáveis das UCs
def copy_cells(source_sheet, target_sheet,idx):
    delta=0
    rows_resp=[]
    for r, row in enumerate(source_sheet.iter_rows()):
        #print('main r',r)
        # linhas em que há docente indicado
        if source_sheet.cell(r+1,idx+1).value: 
            #print(source_sheet.cell(r+1,idx+1).value)
            # ou 1a linha ou linhas dos responsáveis:
            if r==0: # 1a linha do ficheiro
                copy_first_row(r+delta,row, source_sheet, target_sheet)
            else:
                # linhas dos responsáveis
                copy_row(r+delta,row, source_sheet, target_sheet,fill=True)
                rows_resp.append(r+delta)
                for k in range(N):
                    delta+=1
                    partial_copy_row(r+delta,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao) # row comes from source; r+delta is the index in target
        else:
            # linhas em branco
            copy_row(r+delta,row, source_sheet, target_sheet, fill=False)
        if r>40: break #<----------------------------------------------------------------------  remover
    print('delta: ',delta)
    return rows_resp

def copy_first_row(r,row, source_sheet, target_sheet):
     for c, cell in enumerate(row):
        source_cell = cell
        if c <= headers.index(colunas_red[-1]): # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                #target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def copy_row(r,row, source_sheet, target_sheet, fill):
    for c, cell in enumerate(row):
        if c==headers.index(colunas_red[-1]):
            # define default style
            red_cell=cell
    for c, cell in enumerate(row):
        source_cell = cell
        if c <= headers.index(colunas_red[-1]): # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
            if fill:
                target_cell._value = source_cell._value
                target_cell.data_type = source_cell.data_type
                target_cell.font = copy(red_cell.font)
                target_cell.border = copy(red_cell.border)
                target_cell.fill = copy(red_cell.fill)
                target_cell.number_format = copy(red_cell.number_format)
                #target_cell.protection = copy(red_cell.protection)
                target_cell.alignment = copy(red_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def partial_copy_row(r,row, source_sheet, target_sheet, columns_to_copy,coluna_validacao):
    for c, cell in enumerate(row):
        if c==headers.index(colunas_red[-1]):
            # define default style
            red_cell=cell
    for c, cell in enumerate(row):
        source_cell = cell
        if headers[c] in colunas_red: # pintar todas as células até à última a pintar
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell.font = copy(red_cell.font)
            target_cell.border = copy(red_cell.border)
            target_cell.fill = copy(red_cell.fill)
            target_cell.number_format = copy(red_cell.number_format)
            #target_cell.protection = copy(red_cell.protection)
            target_cell.alignment = copy(red_cell.alignment)
        if headers[c] in columns_to_copy:  
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                #target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
        if headers[c]==coluna_validacao: # drop-down menus
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1) # indices em cell começam em 1 ...!!!
            target_cell._value = VALIDATION_VALUE
            target_cell.data_type='s'
            #target_cell.protection = Protection(locked=False)


########################################################################################################
# number max docentes a inserir por UC e informação a passar
# input
# fnDSDv1="DSD_2324_old.xlsx" # copiar 'DSD (para preencher)'
# fnIn="DSD_2324_ML_12abr2023_acrescentar_UC_no_final__.xlsx" # Copiar 'DSD (informação UCs)' +  'DocentesNovo'
# output
N=15 # número máximo docentes
VALIDATION_VALUE='Inserir docente'
# worksheets de input
ws_name_preencher='DSD (para preencher)'
ws_name_info='DSD (informação UCs)'
if DSDv1: 
    ws_name_docentes='DocentesNovo' 
else: 
    ws_name_docentes='docentes'

# ws preencher
columns_to_copy='Nome da UC' # 'Nome' # designação UC
column_key='Inserir docentes na UC ' #'docentes ' # cuidado: tem um espaço a mais
coluna_validacao=column_key
coldoc='AK'
print(column_index_from_string(coldoc))
colunas_red=['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento',  'Responsável', 'Nome da UC', 'ciclo de estudos', 'Código UC', 'ano curricular', 'semestre', 'ECTS', 'semanas', 'Total horas da UC', 'Somatório', 'Horas em falta na UC']
column_codigo_UC_preencher='Código UC'
column_horas_em_falta_preencher='Horas em falta na UC'
column_somatorio_preencher= 'Somatório'
column_preencher_first='Total Horas Teóricas'
column_preencher_last='Total Horas  Outras'
column_new_horas_totais='Horas totais docente'
column_new_horas_semanais='Horas semanais docente'

info_cols_to_unprotect=['H','M','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AD','AE']
DAA='Docente a atribuir'

# ws info
column_total_horas_info='Total Horas previsto '
column_codigo_UC_info='Código UC'

# match strings
MIN_MATCH=80

#################################################################################################
# workbooks
########################################################### criar workbook (target)
# target work book and sheet
wb_target = openpyxl.Workbook()
target_sheet = wb_target.create_sheet(ws_name_preencher)
########################################################### current DSD
if DSDv1: 
    wb1 = openpyxl.load_workbook(fnDSDv1) #, data_only=True) # with data_only, it will only read values
# worksheet input # wb_source is "DSD_2324_ML_12abr2023_acrescentar_UC_no_final__.xlsx" # Copiar 'DSD (informação UCs)' +  'DocentesNovo'
wb_source = openpyxl.load_workbook(fnIn) #, data_only=True) # with data_only, it will only read values
wsnames=wb_source.sheetnames
print('ficheiro Madalena: ', wsnames)
# conteudo worksheet docentes
ws=wb_source[ws_name_preencher]
print([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])
ws=wb_source[ws_name_docentes]
print([c.value for c in next(ws.iter_rows(min_row=1, max_row=1))])
########################################################### workbook resumo_DSD_2324
wbr=openpyxl.Workbook()
wsr_ucs=wbr.create_sheet('UCs')
wsr_docentes=wbr.create_sheet('docentes')
###########################################################

# test if sheet names are right
# print(wsnames)
# for name in [ws_name_preencher,ws_name_info,ws_name_docentes]:
#     if name not in wsnames:
#         stop
#     source_sheet=wb_source[name]
#     print(source_sheet.max_row)
#     print(source_sheet.max_column)



################################################################################# criar main worksheet 'DSD (para preencher)'
source_sheet=wb_source[ws_name_preencher]

# coluna column_key -> idx
headers = [c.value for c in next(source_sheet.iter_rows(min_row=1, max_row=1))]
idx=headers.index(column_key)
idxred=headers.index(colunas_red[-1])
red_cell=source_sheet.cell(1,idxred)
colunas_horas_a_preencher=[]
acc=[]
for h in headers:
    acc.append(h)
    if column_preencher_first in acc and column_preencher_last not in acc:
        colunas_horas_a_preencher.append(h)
colunas_horas_a_preencher.append(column_preencher_last)

print(colunas_horas_a_preencher)


# Criar cópia
rows_resp=copy_sheet(source_sheet, target_sheet,idx)

# create drop-down validation list
# lista de docentes para validação (drop-down menu)
# lista de 'Posições': coluna B
val_sheet=wb_source[ws_name_docentes]
mylist = [c.value for c in val_sheet['A']]
mylistB = [c.value for c in val_sheet['B']]
# to remove None values in list
mylist = [i for i in mylist if i is not None]
mylistB = [i for i in mylistB if i is not None]
nomes_docentes=list(map(lambda x:x.strip(),mylist)) # eliminar \n
posicoes_docentes=list(map(lambda x:x.strip(),mylistB)) # eliminar \n
if len(mylist)!=len(mylistB): stop

nomes_docentes.append(DAA)
posicoes_docentes.append(' ')

max([len(x) for x in posicoes_docentes])

#docentes a remover
#nomes_docentes.remove('Adelino Mendes da Silva Paiva')
#nomes_docentes.remove('João António Ribeiro Ferreira Nunes')
len(nomes_docentes)


# escrever docentes em coluna coldoc e outras 2 colunas adicionais
for i in range(len(nomes_docentes)):
    # horas semanais
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc)-2)
    if i==0: 
        target_cell.value=column_new_horas_semanais
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else:
        target_cell.value=''
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    # horas totais
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc)-1)
    if i==0: 
        target_cell.value=column_new_horas_totais
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else:
        target_cell.value=''
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    # docentes
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc))
    if i==0: 
        target_cell.value=nomes_docentes[i]
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else: 
        target_cell.value=nomes_docentes[i]
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    # posições (categorias)
    target_cell=target_sheet.cell(i+1,column_index_from_string(coldoc)+1)
    if i==0: 
        target_cell.value=posicoes_docentes[i]
        target_cell.font = copy(red_cell.font)
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
        target_cell.number_format = copy(red_cell.number_format)
        #target_cell.protection = copy(red_cell.protection)
        target_cell.alignment = copy(red_cell.alignment)
    else: 
        target_cell.value=posicoes_docentes[i]
        target_cell.border = copy(red_cell.border)
        target_cell.fill = copy(red_cell.fill)
    #if i <5 : print(target_sheet['AZ{}'.format(i+2)].value)
#str1 = ','.join(nomes_docentes)                                    
#str1 = '"'+str1+'"'
#str1= '"Y,N"' # funciona!
#str1='"Vítor Manuel Delgado Alves,Docente a atribuir"' # funciona!
#str1='"Susete Maria Gonçalves Marques,Teresa de Jesus da Silva Matos Nolasco Crespo,Teresa Maria Gonçalves Quilhó Marques dos Santos,Teresa Paula Gonçalves Cruz (EU),Vítor Manuel Delgado Alves,Docente a atribuir"'
#data_val = DataValidation(type='list',formula1=str1) #, allow_blank=False)
data_val = DataValidation(type='list',formula1='={}{}:{}{}'.format(coldoc,'$2',coldoc,'$'+str(1+len(nomes_docentes))), allow_blank=True) #, allow_blank=False)

# criar validação em target_sheet
target_sheet.add_data_validation(data_val)

# abrid DSD_v1
if DSDv1: 
    #wb1 = openpyxl.load_workbook(fnDSDv1) #, data_only=True) # with data_only, it will only read values
    names1=wb1.sheetnames
    # test if sheet names are right
    print(names1)
    for name in [ws_name_preencher,ws_name_info]:
        if name not in names1:
            stop
            ws1=wb1[name]
            print(ws1.max_row)
            print(ws1.max_column)

low_matches=[]
# subsituir os nomes já preenchidos pelo novo nome da lista da RH (nomes_docentes)
def most_similar_name(nome):
    if nome==DAA:
        return DAA
    if nome==VALIDATION_VALUE: 
        return VALIDATION_VALUE
    else:
        L=[fuzz.ratio(nome,x) for x in nomes_docentes]
        if np.max(L) < MIN_MATCH: 
            low_matches.append((nome,np.max(L)))
            return(DAA)
        else: 
            return nomes_docentes[np.argmax(L)]

#fuzz.ratio('Jorge Gominho',"Jorge Manuel Barros D'Almeida Gominho")
#nome='Ana Cristina Delaunay Caperta'
#most_similar_name('Jorge Gominho')

# depois de fazer a cópia de source_sheet para target_sheet:
# indicar as células em que ficam os drop-down menus e desproteger as células a preencher
# if DSDv1: copiar valores das células de DSDv_1 para DSD_v2 (target sheet):
if True:
    # headers = [c.value for c in next(source_sheet.iter_rows(min_row=1, max_row=1))]
    # colocar drop-down menu para colunas em coluna_validacao (que têm valor VALIDATION_VALUE='VAL')
    # desproteger células de coluna_validacao e de colunas_horas_a_preencher
    if DSDv1:
        ws1=wb1[ws_name_preencher]
    idxval=headers.index(coluna_validacao) # where drop-manu will be
    idx_not_empty=0
    for k in range(target_sheet.max_row):
        ct=target_sheet.cell(k+1,idxval+1)
        if ct.value!=VALIDATION_VALUE: # originalmente é sempre == VALIDATION_VALUE ,não depende de DSDv1
            # for resumo_DSD2324 
            for colpreencher in colunas_horas_a_preencher+['Nome da UC', 'ciclo de estudos', 'ano curricular', 'semestre']+['Grandes Áreas Científicas (FOS)', 'Áreas Científicas (FOS)', 'Áreas Disicplinares', 'Departamento']:
                idx=headers.index(colpreencher)
                dt=target_sheet.cell(k+1,idx+1)
                cr=wsr_ucs.cell(idx_not_empty+1,idx+1)
                cr.value=dt.value
            idx_not_empty +=1
        else:
            data_val.add(ct)
            if DSDv1:
                c1=ws1.cell(k+1,idxval+1) # mesma célula de ws1 e target_sheet<---------------------------------------
                if not c1.protection.locked and c1.value is not None and c1.value != '':
                     if False: 
                         newname=most_similar_name(c1.value)
                     else: 
                         newname=c1.value
                     ct.value=newname
                     # for resumo_DSD2324 
                     if c1.value!=VALIDATION_VALUE and c1.value!=DAA:
                        cr=wsr_ucs.cell(idx_not_empty+1,idxval+1)
                        cr.value=newname
                        idx_not_empty +=1
            ct.protection = Protection(locked=False)
            for colpreencher in colunas_horas_a_preencher:
                idx=headers.index(colpreencher)
                dt=target_sheet.cell(k+1,idx+1)
                if DSDv1:
                    d1=ws1.cell(k+1,idx+1) # mesma célula
                    if not d1.protection.locked and d1.value is not None and d1.value != '':
                        dt.value=d1.value
                        # for resumo_DSD2324 
                        if c1.value!=VALIDATION_VALUE and c1.value!=DAA:
                            cr=wsr_ucs.cell(idx_not_empty+1,idx+1)
                            cr.value=d1.value
                dt.protection = Protection(locked=False)

# """ ct=target_sheet['E3']
# data_val.add(ct)
# ct.value
# c1=ws1['E3']
# c1.value
# ct.value=c1.value
# ct.value """
#print(d1.value)
wb1.close

# ajustar tamanho das colunas
target_sheet.column_dimensions[coldoc].width = max([len(x) for x in nomes_docentes])
target_sheet.column_dimensions[get_column_letter(column_index_from_string(coldoc)+1)].width = max([len(x) for x in posicoes_docentes])

#dim_holder = DimensionHolder(worksheet=target_sheet)
#dim_holder[coldoc] = ColumnDimension(target_sheet, min=col, max=col, width=30)
#target_sheet.column_dimensions = dim_holder


####################################################################### copiar worksheet 'DSD (informação UCs)'
if DSDv1: 
    source_sheet=wb1[ws_name_info]  # lê no ficheiro DSD já prépreenchido
else:
    source_sheet=wb_source[ws_name_info] # lê no ficheiro DA

# new sheet to create
info_sheet = wb_target.create_sheet(ws_name_info)

copy_sheet_attributes(source_sheet, info_sheet)
for r, row in enumerate(source_sheet.iter_rows()):
    for c, cell in enumerate(row):
        source_cell = cell
        if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
            continue
        target_cell = info_sheet.cell(column=c+1, row=r+1) # indices am cell começam em 1 ...!!!
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        if get_column_letter(c+1) in info_cols_to_unprotect:
            target_cell.protection = Protection(locked=False)
        #target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
        if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

#headers = [c.value for c in next(info_sheet.iter_rows(min_row=1, max_row=1))]

copyColumn(info_sheet,column_codigo_UC_info,'AY','numeric')
copyColumn(info_sheet,column_total_horas_info,'AZ','numeric')
# column_total_horas

# ##################################################################### criar fórmulas em ws_name_preencher
# '''
# column_codigo_UC_preencher='Código UC'
# column_horas_em_falta_preencher='Horas em falta na UC'
# column_key='Inserir docentes na UC ' #'docentes ' # cuidado: tem um espaço a mais
# coluna_validacao=column_key
# column_somatorio_preencher
# coldoc='AK'
# # ws info
# column_total_horas_info='Total Horas previsto '
# column_codigo_UC_info='Código UC'
# column_preencher_first='Total Horas Teóricas'
# column_preencher_last='Total Horas  Outras'
# '''

for r in rows_resp:
    d=target_sheet.cell(r+1,column_index_from_string(nomeColuna2letter(target_sheet, column_somatorio_preencher)))
    d.value='' #"=SUM({}{}:{}{})".format(nomeColuna2letter(target_sheet, column_preencher_first),r+2,nomeColuna2letter(target_sheet, column_preencher_last),r+1+N)
    # total horas de docência por docente para a UC
    for delta in range(N):
        k=r+1+delta+1
        e=target_sheet.cell(k,column_index_from_string(nomeColuna2letter(target_sheet, column_somatorio_preencher)))
        e.value="=SUM({}{}:{}{})".format(nomeColuna2letter(target_sheet, column_preencher_first),k,nomeColuna2letter(target_sheet, column_preencher_last),k)
    # horas em falta: total de horas que vem de info menos soma das horas dos docentes
    c=target_sheet.cell(r+1,column_index_from_string(nomeColuna2letter(target_sheet, column_horas_em_falta_preencher)))
    # código 1564 -- Sistemas de Produção Animal nos Trópicos
    if r+1<4948: # com as UC originais
        c.value="=VLOOKUP(int({}{}),'{}'!AY:AZ, 2, FALSE)-SUM({}{}:{}{})".format(nomeColuna2letter(target_sheet, column_codigo_UC_preencher),r+1,ws_name_info,nomeColuna2letter(target_sheet, column_preencher_first),r+2,nomeColuna2letter(target_sheet, column_preencher_last),r+1+N)
    else: # UC adicionada no fim do ficheiro fnIn
        c.value="=VLOOKUP(1564,'{}'!AY:AZ, 2, FALSE)-SUM({}{}:{}{})".format(ws_name_info,nomeColuna2letter(target_sheet, column_preencher_first),r+2,nomeColuna2letter(target_sheet, column_preencher_last),r+1+N)

#copyColumn(target_sheet,column_somatorio_preencher,'AZ')
for i in range(len(nomes_docentes)): #target_sheet.max_row):
    if i>0:
        g=target_sheet.cell(i+1,column_index_from_string(coldoc)-1)
        g.value="=SUMIF({}:{},{}{},{}:{})".format(nomeColuna2letter(target_sheet, column_key),nomeColuna2letter(target_sheet, column_key),coldoc,i+1,nomeColuna2letter(target_sheet, column_somatorio_preencher),nomeColuna2letter(target_sheet, column_somatorio_preencher))
        h=target_sheet.cell(i+1,column_index_from_string(coldoc)-2)
        h.value="=SUMIF({}:{},{}{},{}:{})/28".format(nomeColuna2letter(target_sheet, column_key),nomeColuna2letter(target_sheet, column_key),coldoc,i+1,nomeColuna2letter(target_sheet, column_somatorio_preencher),nomeColuna2letter(target_sheet, column_somatorio_preencher))
        h.number_format='0.00'

# =SUMIF(E:E,K2,AA) # adaptar

# column_index_from_string
#nomeColuna2letter(target_sheet, column_codigo_UC_preencher)
#=VLOOKUP(int(I2);'DSD (informação UCs)'!AY:AZ; 2; FALSE)
# =SUM(O3:Z12)


###################################################################### gravar wb_target
# eliminar a worksheet original
if 'Sheet' in wb_target.sheetnames:  # remove default sheet
    wb_target.remove(wb_target['Sheet'])
if 'Sheet' in wbr.sheetnames:  # remove default sheet
    wbr.remove(wbr['Sheet'])


# proteger
target_sheet.protection.sheet = True
target_sheet.protection.password = 'kathleen'
info_sheet.protection.sheet = True
info_sheet.protection.password = 'kathleen'

# gravar para ficheiro
wb_target.save(fnOut)
wbr.save(fnResumo)

wb_target.close
wbr.close

# what is this?
# isinstance(c, openpyxl.cell.read_only.EmptyCell)

print(low_matches)


