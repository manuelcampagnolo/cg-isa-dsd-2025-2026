import openpyxl
from openpyxl import *
import numpy as np

# for 1st filter to delete all rows after M-th
M=1000

fnIn="DSD_2324_ML_5abr2023.xlsx"
fnOut="DSD_2324_ML_5abr2023_.xlsx"

wb2 = openpyxl.load_workbook(fnIn)

############################################# filter 1
# remove all rows after M-th
for sheet in wb2.worksheets: 
    print ('Your currently in ', sheet)  
    max_row_in_sheet = sheet.max_row  
    max_col_in_sheet = sheet.max_column 
    print (max_row_in_sheet, max_col_in_sheet)
    # delete all rows from M on
    sheet.delete_rows(M,max_row_in_sheet-M)

####################################################### filter 2
# remove empty rows and columns from the reduced wb2
for sheet in wb2.worksheets: 
    # indices of empty rows
    indx = []
    for i in range(len(tuple(sheet.rows))):
        flag = False
        for cell in tuple(sheet.rows)[i]:
            if cell.value != None:
                flag = True
                break
        if flag == False:
            indx.append(i)
    indx.sort()
    for i in range(len(indx)):
        sheet.delete_rows(idx = indx[i]+1-i)

    max_row_in_sheet = sheet.max_row  # maximum enterd row
    max_col_in_sheet = sheet.max_column  # maximum entered column
    print (max_row_in_sheet, max_col_in_sheet)

wb2.save(fnOut)